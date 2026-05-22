"""
Capital Gains Automation - Multi-format Extractor v5
New: Dezerv/PMS Excel, ICICI PMS PDF, generic table parser, date range filter, password fix
"""

import re, zipfile, os, sys, json, argparse
from datetime import datetime
from pathlib import Path

import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# UTILS
# ─────────────────────────────────────────────

def parse_date(s):
    if s is None: return None
    if isinstance(s, datetime): return s
    if isinstance(s, pd.Timestamp): return s.to_pydatetime()
    if not isinstance(s, str):
        try: return pd.to_datetime(s).to_pydatetime()
        except: return None
    s = s.strip()
    if s.lower() in ("", "nan", "none", "nat", "n/a", "n.a.", "-"): return None
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d",
                "%d-%b-%Y", "%d-%b-%y", "%d %b %Y", "%d %b %y",
                "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try: return datetime.strptime(s, fmt)
        except: continue
    return None

def to_float(s):
    if s is None: return None
    if isinstance(s, (int, float)):
        v = float(s)
        return None if (v != v) else v  # handle NaN
    s = str(s).replace(",","").replace("₹","").replace("\xa0","").replace("\xad","").replace("(","").replace(")","").strip()
    if s in ("", "-", "N.A.", "N/A", "NA"): return None
    try: return float(s)
    except: return None

def password_from_filename(path):
    """Fix 7: password = last segment before the final dot extension.
    e.g. vansh.1.pass.pdf -> 'pass'
    e.g. Vansh final action.csv -> 'action' (last word)
    """
    name = Path(path).stem   # removes last extension
    # If name still has dots, take last dot-segment
    if '.' in name:
        return name.rsplit('.', 1)[-1]
    # Otherwise take last space-separated word
    parts = name.strip().split()
    return parts[-1] if parts else None


DEBT_GRANDFATHERING_DATE = datetime(2023, 3, 31)

def determine_gain_type(purchase_date, sale_date, is_equity=True):
    """Determine ST/LT based on holding period and asset class.
    
    Equity (issue 6): Long Term if days > 365
    Debt (issue 7):   Long Term ONLY if:
      - Purchase date on or before 31-Mar-2023 (pre-Finance Act 2023 grandfathering), AND
      - Held for more than 730 days (2 years)
    All other debt: Short Term (Finance Act 2023 removed LTCG for new debt purchases)
    """
    if not purchase_date or not sale_date:
        return "Short Term"
    days = (sale_date - purchase_date).days
    if is_equity:
        return "Long Term" if days > 365 else "Short Term"
    else:
        if purchase_date <= DEBT_GRANDFATHERING_DATE and days > 730:
            return "Long Term"
        return "Short Term"


def _get_pdf_passwords(pdf_path):
    """Get candidate passwords from filename.
    Rule: last dot-segment before final extension.
    e.g. vansh.1.pass.pdf -> ['pass']
    e.g. CAMS statement 1234.pdf -> ['1234']
    """
    stem = Path(pdf_path).stem
    pwds = []
    if '.' in stem:
        pwds.append(stem.rsplit('.', 1)[-1])
    parts = [p for p in stem.replace('_', ' ').split() if p]
    if parts and parts[-1] not in pwds:
        pwds.append(parts[-1])
    return pwds


def _pikepdf_decrypt(pdf_path, password):
    """Decrypt PDF using pikepdf (handles AES-128/256 that pdfplumber can't).
    Returns path to decrypted temp file, or None on failure.
    """
    try:
        import pikepdf
        out = pdf_path + '._dec.pdf'
        with pikepdf.open(pdf_path, password=password) as pk:
            pk.save(out)
        return out
    except Exception:
        return None


def open_pdf(pdf_path, password=None):
    """Open a PDF for text extraction, handling password protection automatically.

    For password-protected files, rename with the password as the last
    dot-segment before .pdf:
        CAMS.mypassword.pdf   ← 'mypassword' will be tried
        report.1234.pdf       ← '1234' will be tried

    Strategy per password candidate:
      1. Try pdfplumber directly (works for RC4 / unprotected)
      2. Try pikepdf decrypt → pdfplumber (works for AES-128/256)
    """
    candidates = list(_get_pdf_passwords(pdf_path))
    if password and password not in candidates:
        candidates.insert(0, password)
    candidates.append(None)  # always try unprotected

    last_err = None
    dec_files = []   # temp decrypted files to clean up

    try:
        for pwd in candidates:
            # ── Attempt 1: pdfplumber direct ──────────────────────────────
            opened = None
            try:
                kwargs = {'password': pwd} if pwd else {}
                opened = pdfplumber.open(pdf_path, **kwargs)
                text = (opened.pages[0].extract_text() or '').strip()
                if text:          # got real content → success
                    return opened
                opened.close(); opened = None
            except Exception as e:
                last_err = e
                if opened:
                    try: opened.close()
                    except: pass

            if not pwd:
                continue   # pikepdf can't help without a password

            # ── Attempt 2: pikepdf decrypt → pdfplumber ───────────────────
            dec = _pikepdf_decrypt(pdf_path, pwd)
            if dec:
                dec_files.append(dec)
                try:
                    opened = pdfplumber.open(dec)
                    text = (opened.pages[0].extract_text() or '').strip()
                    if text:
                        return opened   # caller closes this
                    opened.close(); opened = None
                except Exception as e:
                    last_err = e
                    if opened:
                        try: opened.close()
                        except: pass

        raise Exception(
            f"Cannot read this PDF. "
            f"If it is password-protected, rename it as: filename.PASSWORD.pdf "
            f"(e.g. CAMS.abc123.pdf where abc123 is the password). "
            f"Last error: {last_err}"
        )
    finally:
        for p in dec_files:
            try: os.unlink(p)
            except: pass

def extract_client_name(text):
    patterns = [
        r'NAME\s*:\s*([A-Za-z][A-Za-z\s\.]+?)(?:\s+STATUS|\s+PAN)',
        r'(?:Client Name|Name|Investor Name|Account)\s*[:\-]\s*([A-Z][A-Z\s\.]+?)(?:\s*\n|\s*PAN|\s*Account)',
        r'SANTHI SRI VOORA|SANJAY DEVUDU',  # fallback specific
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            name = m.group(0) if m.lastindex is None else m.group(1)
            name = re.sub(r'\s+', ' ', name).strip()
            if 3 < len(name) < 60:
                return name.title()
    # Try generic name extraction
    m = re.search(r'(?:SANTHI|SANDEEP|SANJAY|SUDAKSHINA|NIMMARAJU)\s+[A-Z\s]+', text.upper())
    if m:
        return m.group(0).strip().title()
    return None

def source_name_from_file(file_path):
    return Path(file_path).stem

def apply_grandfathering(sale_amt, fmv_total, purchase_amt, transfer_expenses=0):
    """Apply Section 112A grandfathering formula.
    Formula: Sale - IF(ISBLANK(J), M, MAX(MIN(H,J),M)) - K
    where H=sale_amt, J=fmv_total, M=purchase_amt, K=transfer_expenses
    """
    te = transfer_expenses or 0
    if fmv_total and fmv_total > 0:
        effective_cost = max(min(sale_amt, fmv_total), purchase_amt)
    else:
        effective_cost = purchase_amt
    return round(sale_amt - effective_cost - te, 4)


def _determine_share_type(name, asset_type=None):
    """
    Returns "Transaction Tax" (equity) or "Unlisted Security" (debt/other).
    
    Priority rules:
      1. Explicit "unlisted" keyword → Unlisted Security
      2. Explicit equity signals (share, equity MF, fof/fund of fund) → Transaction Tax
      3. Pure debt instrument keywords → Unlisted Security  
      4. asset_type guidance
      5. ETF alone is ambiguous — only debt-type ETFs → Unlisted Security
      6. Default → Transaction Tax
    """
    nl = (name or "").lower()
    al = str(asset_type or "").lower()

    # Rule 1: explicit unlisted
    if "unlisted" in nl or "unlisted" in al:
        return "Unlisted Security"

    # Rule 2a: international/overseas FOFs are non-equity for Indian tax purposes
    _NON_EQUITY_FOF_KW = ["nasdaq", "nyse", "global x", "global electric",
                          "passive fof", "etf fof", "international fund",
                          "overseas fund", "global fund", "us equity fof"]
    if any(kw in nl for kw in _NON_EQUITY_FOF_KW):
        return "Unlisted Security"

    # Rule 2b: explicit equity signals win
    _EQUITY_KW = ["share", "fund of fund", "fof", "equity fund", "equity mutual",
                  "large cap", "mid cap", "small cap", "multicap", "flexi cap",
                  "bluechip", "blue chip", "contra", "thematic", "sectoral",
                  "elss", "index fund", "nifty", "sensex",
                  "emerging bluechip"]
    if any(kw in nl for kw in _EQUITY_KW):
        return "Transaction Tax"

    # Rule 3: pure debt/fixed-income keywords → Unlisted Security
    _DEBT_KW = ["liquid", "arbitrage", "ultra-short term", "ultra short term",
                "non-equity", "money market", "duration", "debenture",
                "overnight", "gilt", "banking and psu", "credit risk",
                "corporate bond", "floater", "fixed maturity", "fixed income",
                "debt fund", "short term fund", "low duration"]
    if any(kw in nl for kw in _DEBT_KW):
        return "Unlisted Security"

    # Debt bond specifically (but NOT "bond fund" which is debt, NOR "bond etf")
    if "bond" in nl and "fund" not in nl:
        return "Unlisted Security"

    # Rule 4: asset_type guidance — check "non-equity" before "equity" to avoid substring match
    if "debt" in al:
        return "Unlisted Security"
    if "non-equity" in al or al == "non equity":
        return "Unlisted Security"
    if "equity" in al:
        return "Transaction Tax"

    # Rule 5: standalone "etf" — debt ETFs are Unlisted, equity ETFs are Transaction Tax
    # If ETF and no equity signal already caught: check for debt ETF markers
    if "etf" in nl:
        _DEBT_ETF_KW = ["liquid etf", "debt etf", "gilt etf", "bond etf"]
        if any(kw in nl for kw in _DEBT_ETF_KW):
            return "Unlisted Security"
        # All other ETFs (equity index ETFs) → Transaction Tax
        return "Transaction Tax"

    # Rule 6: default equity
    return "Transaction Tax"


def make_row(source, name, isin, sale_date, qty, sale_price, sale_amt,
             purchase_date, purchase_amt, gain_type, client_name=None, asset_type=None,
             market_price_31jan2018=None, total_market_value_31jan2018=None,
             transfer_expenses=None):
    """Standardize row creation."""
    return {
        "source": source, "share_name": name,
        "share_type": _determine_share_type(name, asset_type),
        "isin": isin or "",
        "sale_date": sale_date, "quantity": qty,
        "sale_price": sale_price, "sale_amount": sale_amt,
        "market_price_31jan2018": market_price_31jan2018,
        "total_market_value_31jan2018": total_market_value_31jan2018,
        "transfer_expenses": transfer_expenses,
        "purchase_date": purchase_date,
        "purchase_amount": purchase_amt, "fmv_unquoted": None,
        "gain_type": gain_type,
        "mutual_fund_col": "Equity Mutual Fund",
        "section": "112A", "scheme_type": "Equity",
        "client_name": client_name,
        "asset_type": asset_type or "",
    }



def build_validation_report(rows, pdf_st=None, pdf_lt=None, pdf_total=None, parse_failures=0, parser_name=None):
    """Universal validation: compute extracted totals, compare vs PDF-stated totals (if available).

    Returns a dict with extracted/stated totals, gap, and cause analysis.
    Works even when pdf_st/pdf_lt/pdf_total are None — in that case only extraction stats are reported.
    """
    ext_st = ext_lt = 0.0
    zero_cost, intraday = [], []

    for r in rows:
        sale  = r.get('sale_amount') or 0.0
        cost  = r.get('purchase_amount') or 0.0
        gain  = sale - cost
        gtype = r.get('gain_type', '')
        pdate = r.get('purchase_date')
        sdate = r.get('sale_date')

        if gtype == 'Short Term':
            ext_st += gain
        else:
            ext_lt += gain

        if cost == 0.0 and sale > 0:
            zero_cost.append({'name': r.get('share_name'), 'sale': round(sale, 2)})
        if pdate and sdate and getattr(pdate, 'date', lambda: pdate)() == getattr(sdate, 'date', lambda: sdate)():
            intraday.append({'name': r.get('share_name'), 'gain': round(gain, 2),
                             'date': str(getattr(sdate, 'date', lambda: sdate)())})

    ext_total = ext_st + ext_lt
    gap_total = round(pdf_total - ext_total, 2) if pdf_total is not None else None
    gap_pct   = round(abs(gap_total) / abs(pdf_total) * 100, 2) if (pdf_total and gap_total is not None) else None

    # Per-category gaps (ST and LT individually)
    gap_st = round(pdf_st - ext_st, 2) if pdf_st is not None else None
    gap_lt = round(pdf_lt - ext_lt, 2) if pdf_lt is not None else None
    gap_st_pct = round(abs(gap_st) / abs(pdf_st) * 100, 2) if (pdf_st and gap_st is not None) else None
    gap_lt_pct = round(abs(gap_lt) / abs(pdf_lt) * 100, 2) if (pdf_lt and gap_lt is not None) else None

    causes = []
    if zero_cost:
        causes.append({
            'type': 'ZERO_COST_BONUS_SHARES',
            'count': len(zero_cost),
            'note': 'Purchase cost is 0 — bonus/rights/demerger shares. Gain = full sale value. '
                    'Correct per PDF.',
            'rows': zero_cost[:5],
        })
    if intraday:
        causes.append({
            'type': 'INTRADAY_TRADES',
            'count': len(intraday),
            'note': 'Buy and sell on same date. Broker includes in ST G/L but these are '
                    'speculative income, not capital gains — correctly excluded from output.',
            'rows': intraday,
        })
    if parse_failures:
        causes.append({
            'type': 'PARSE_FAILURES',
            'count': parse_failures,
            'note': 'Rows the parser could not reconstruct — dates/amounts unreadable. '
                    'Each failure may contribute to missing gain.',
        })
    # Flag when ST and LT diverge significantly even if totals cancel (gain_type misclassification)
    if (gap_st_pct is not None and gap_st_pct > 1.0) or (gap_lt_pct is not None and gap_lt_pct > 1.0):
        _st_str = f"{gap_st_pct:.1f}% (₹{gap_st:,.2f})" if gap_st_pct is not None else "n/a"
        _lt_str = f"{gap_lt_pct:.1f}% (₹{gap_lt:,.2f})" if gap_lt_pct is not None else "n/a"
        causes.append({
            'type': 'GAIN_TYPE_MISMATCH',
            'note': (f'ST gap: {_st_str} | LT gap: {_lt_str}. '
                     f'Even if the total is close, rows may be misclassified between ST and LT. '
                     f'Verify "Shares w/o STT" or borderline-365-day holdings.'),
        })
    if gap_pct is not None and gap_pct > 1.0:
        causes.append({
            'type': 'AMOUNT_DISCREPANCY',
            'note': f'Total gain gap is {gap_pct:.1f}% (₹{gap_total:,.2f}). '
                    f'Review zero-cost rows and parse failures above.',
        })

    return {
        'parser':           parser_name,
        'pdf_st_gain':      pdf_st,
        'pdf_lt_gain':      pdf_lt,
        'pdf_total_gain':   pdf_total,
        'extracted_st_gain':    round(ext_st, 2),
        'extracted_lt_gain':    round(ext_lt, 2),
        'extracted_total_gain': round(ext_total, 2),
        'gap':              gap_total,
        'gap_pct':          gap_pct,
        'gap_st':           gap_st,
        'gap_lt':           gap_lt,
        'gap_st_pct':       gap_st_pct,
        'gap_lt_pct':       gap_lt_pct,
        'rows_extracted':   len(rows),
        'collapsed_parse_failures': parse_failures,
        'causes':           causes,
    }


# ─────────────────────────────────────────────
# DEZERV / CONSOLIDATED CAPITAL GAIN EXCEL
# ─────────────────────────────────────────────

class DezervExcelParser:
    """Parses Dezerv/PMS consolidated capital gain Excel statements.
    Format: Account | Strategy | Asset Type | Security | Sale Date | Qty | Rate | Sale Amt |
            Purchase Date | Purchase Rate | Price 31-Jan-18 | Purchase Amt | Indexed Cost |
            Days Held | ST Gain/Loss | LT Gain/Loss | LT After Index
    """
    def parse(self, file_path, source_name=None):
        self.source_name = source_name or source_name_from_file(file_path)
        self.client_name = None
        rows = []
        xl = pd.ExcelFile(file_path)

        # Look for 'Consolidated' or first sheet
        sheet = 'Consolidated' if 'Consolidated' in xl.sheet_names else xl.sheet_names[0]
        df = pd.read_excel(file_path, sheet_name=sheet, header=None)

        # Extract client name from title row
        title = str(df.iloc[0, 0]) if pd.notna(df.iloc[0, 0]) else ""
        m = re.search(r'–\s+(.+?)(?:\s*\||\s*$)', title)
        if m: self.client_name = m.group(1).strip()

        # Find header row
        hdr_row = None
        for i, row in df.iterrows():
            vals = [str(v).strip().lower() for v in row if pd.notna(v)]
            if 'security' in vals and 'sale date' in vals and 'sale amount' in vals:
                hdr_row = i; break
        if hdr_row is None: return rows

        hdr_pos = df.index.get_loc(hdr_row)
        headers = [str(v).strip().lower() if pd.notna(v) else "" for v in df.iloc[hdr_pos]]

        col = {}
        for i, h in enumerate(headers):
            if h == 'account': col['account'] = i
            elif h == 'strategy': col['strategy'] = i
            elif h == 'asset type': col['asset_type'] = i
            elif h == 'security': col['name'] = i
            elif h == 'sale date': col['sale_date'] = i
            elif h == 'qty sold' or h == 'quantity': col['qty'] = i
            elif h == 'sale rate': col['sale_rate'] = i
            elif h == 'sale amount': col['sale_amt'] = i
            elif h == 'purchase date': col['purchase_date'] = i
            elif h == 'purchase rate': col['purchase_rate'] = i
            elif h == 'purchase amt': col['purchase_amt'] = i
            elif h == 'indexed cost': col['indexed_cost'] = i
            elif h == 'days held': col['days'] = i
            elif h == 'st gain/loss': col['st_gain'] = i
            elif h == 'lt gain/loss': col['lt_gain'] = i
            elif h == 'lt after index': col['lt_after_index'] = i
            elif h == 'st gain/loss': col['st_gain'] = i
            elif h == 'lt gain/loss': col['lt_gain'] = i
            elif h == 'isin' or h == 'isin code': col['isin'] = i

        if 'name' not in col or 'sale_date' not in col: return rows

        def g(row, field):
            idx = col.get(field)
            if idx is None or idx >= len(row): return None
            v = row.iloc[idx]
            return None if pd.isna(v) else v

        for pos in range(hdr_pos + 1, len(df)):
            row = df.iloc[pos]
            if row.isna().all(): continue

            name = str(g(row, 'name') or '').strip()
            if not name or name.lower() in ('nan','','security','sub-total') or 'sub-total' in name.lower():
                continue

            sale_date = parse_date(str(g(row, 'sale_date') or ''))
            purchase_date = parse_date(str(g(row, 'purchase_date') or ''))
            if not sale_date: continue

            sale_amt = to_float(g(row, 'sale_amt'))
            purchase_amt = to_float(g(row, 'purchase_amt'))
            if not sale_amt: continue
            if not purchase_amt: purchase_amt = 0.0

            qty = to_float(g(row, 'qty'))
            sale_rate = to_float(g(row, 'sale_rate'))
            days = to_float(g(row, 'days'))
            asset_type = str(g(row, 'asset_type') or '').strip()

            is_eq = asset_type.lower() not in ('debt', 'non-equity mf', 'debt mf', 'non equity mf')
            gain_type = determine_gain_type(purchase_date, sale_date, is_equity=is_eq)

            # For Long Term rows: use Indexed Cost as purchase_amount
            # so Capital Gain = Sale - Indexed Cost = LT After Index (tax-relevant figure)
            # For Short Term rows: use raw Purchase Amt
            indexed_cost = to_float(g(row, 'indexed_cost'))
            if gain_type == "Long Term" and indexed_cost and indexed_cost > 0:
                effective_purchase = indexed_cost
            else:
                effective_purchase = purchase_amt

            rows.append(make_row(
                self.source_name, name, str(g(row, 'isin') or '').strip(),
                sale_date, qty, sale_rate, sale_amt,
                purchase_date, effective_purchase, gain_type,
                self.client_name, asset_type
            ))

        return rows


# ─────────────────────────────────────────────
# ICICI PRUDENTIAL PMS PDF PARSER
# ─────────────────────────────────────────────

class ICICIPMSParser:
    """Parses ICICI Prudential PMS Capital Gain/Loss Statement PDFs.
    Format: Security | Sale Date | Sale Qty | Sale Rate | Sale Amt | 
            Purchase Date | Purchase Rate | Price 31-Jan | Purchase Amt | Indexed Cost | Days | ST | LT | LT After Index
    """
    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        with open_pdf(pdf_path) as pdf:
            pages_text = [p.extract_text() or "" for p in pdf.pages]
        full_text = "\n".join(pages_text)
        self.client_name = self._extract_client(full_text)
        return self._parse_all_pages(pages_text)

    def _extract_client(self, text):
        m = re.search(r'Account\s*:\s*\d+\s+\w+\s*-\s*([A-Z\s]+?)(?:ICICI|$|\n)', text)
        if m: return m.group(1).strip().title()
        return None

    def _parse_all_pages(self, pages_text):
        rows = []
        current_strategy = None
        current_account = None

        for page_text in pages_text:
            lines = page_text.split('\n')

            for line in lines:
                line = line.strip()

                # Strategy header line
                acct_m = re.match(r'Account\s*:\s*(\d+)\s+\S+\s*-\s*([A-Z\s]+)', line, re.IGNORECASE)
                if acct_m:
                    current_account = acct_m.group(1)
                    continue

                # Strategy name line (ICICI Prudential PMS Liquid Strategy etc)
                if re.match(r'ICICI Prudential PMS', line, re.IGNORECASE):
                    current_strategy = line.strip()
                    continue

                # Skip header/total/summary lines
                if re.match(r'(Total|Security|Sale Date|Purchase|Capital Gain|This is|NSDL-|From\s+\d)', line, re.IGNORECASE):
                    continue
                if not line: continue

                # Data row: Security-ISIN Date Qty Rate Amt Date Rate N.A. Amt Amt Days ST LT LT
                # Try to parse: ends with numbers for ST/LT columns
                # Format: "Security Name-ISIN DD/MM/YY qty rate sale_amt DD/MM/YY rate N.A. purch_amt idx days st lt lt_idx"
                row_data = self._parse_data_line(line)
                if row_data:
                    name, isin, sale_date, qty, sale_rate, sale_amt, \
                    purchase_date, purchase_rate, purchase_amt, days, st_gain, lt_gain = row_data

                    gain_type = determine_gain_type(purchase_date, sale_date, is_equity=True)
                    src = f"{self.source_name}"
                    if current_strategy:
                        src = self.source_name

                    rows.append(make_row(
                        src, name, isin, sale_date, qty, sale_rate, sale_amt,
                        purchase_date, purchase_amt, gain_type,
                        self.client_name, current_strategy
                    ))

        return rows

    def _parse_data_line(self, line):
        """Parse a single data line from ICICI PMS statement."""
        # Lines look like:
        # "ICICI PRUDENTIAL S&P BSE LIQUID RATE ETF- 14/02/24 1,000 999.965 999,965.00 01/02/24 1,000.18 N.A. 1,000,180.00 1,000,180.00 13 -215.00 0.00 0.00"
        # "INF109KC1KT9"  <- sometimes ISIN is on separate line, already merged by extract_text

        # Extract dates
        dates = re.findall(r'\d{2}/\d{2}/\d{2,4}', line)
        if len(dates) < 2: return None

        # Split on first date to get security name + ISIN
        first_date_pos = line.find(dates[0])
        name_part = line[:first_date_pos].strip()

        # Extract ISIN from name part
        isin = ""
        isin_m = re.search(r'(IN[A-Z0-9]{10})', name_part)
        if isin_m:
            isin = isin_m.group(1)
            name_part = name_part[:isin_m.start()].strip()

        # Clean name
        name = re.sub(r'[-\s]+$', '', name_part).strip()
        if not name or len(name) < 3: return None

        # Get numbers after removing dates and name
        remaining = line[first_date_pos:]
        remaining_no_dates = re.sub(r'\d{2}/\d{2}/\d{2,4}', ' ', remaining)
        remaining_no_dates = remaining_no_dates.replace('N.A.', ' ').replace('N.A', ' ')
        nums = [to_float(n) for n in re.findall(r'-?[\d,]+\.?\d*', remaining_no_dates) if to_float(n) is not None]

        if len(nums) < 6: return None

        sale_date = parse_date(dates[0])
        purchase_date = parse_date(dates[1])
        if not sale_date or not purchase_date: return None

        # Layout after dates removed: qty sale_rate sale_amt purchase_rate purchase_amt indexed_cost days st lt lt_idx
        qty = nums[0]
        sale_rate = nums[1]
        sale_amt = nums[2]
        purchase_rate = nums[3]
        purchase_amt = nums[4]
        # indexed_cost = nums[5] (skip)
        days = nums[6] if len(nums) > 6 else None
        st_gain = nums[7] if len(nums) > 7 else 0
        lt_gain = nums[8] if len(nums) > 8 else 0

        if not sale_amt or sale_amt <= 0: return None

        return name, isin, sale_date, qty, sale_rate, sale_amt, \
               purchase_date, purchase_rate, purchase_amt, days, st_gain, lt_gain


# ─────────────────────────────────────────────
# CAMS PDF PARSER
# ─────────────────────────────────────────────

class CAMSParser:
    # Last line of CAMS statement: "Total  <ST>  <LT>"
    _TOTAL_RE = re.compile(r'^Total\s+([-\d,]+\.?\d*)\s+([-\d,]+\.?\d*)\s*$', re.M)

    def _extract_pdf_summary(self, pdf_path):
        """Read the grand-total ST/LT line from the CAMS statement last page."""
        try:
            with open_pdf(pdf_path) as pdf:
                # Total line appears on the last page
                text = pdf.pages[-1].extract_text() or ''
            m = self._TOTAL_RE.search(text)
            if m:
                st = float(m.group(1).replace(',', ''))
                lt = float(m.group(2).replace(',', ''))
                return round(st, 2), round(lt, 2), round(st + lt, 2)
        except Exception:
            pass
        return None, None, None

    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        with open_pdf(pdf_path) as pdf:
            full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        self.client_name = extract_client_name(full_text)
        rows = self._parse_text(full_text)
        pdf_st, pdf_lt, pdf_total = self._extract_pdf_summary(pdf_path)
        self.last_validation = build_validation_report(
            rows, pdf_st=pdf_st, pdf_lt=pdf_lt, pdf_total=pdf_total,
            parse_failures=0, parser_name='CAMSParser',
        )
        return rows

    def _parse_text(self, text):
        rows = []
        lines = text.split("\n")
        current_scheme = None
        current_isin = None
        current_scheme_type = "Equity"

        redemption_re = re.compile(
            r'(?:Redemption|Switch.Out)\s+(\d{2}-\w{3}-\d{4})\s+([\d,\.]+)\s+([\d,\.]+)\s+([\d,\.]+)'
        )

        for i, line in enumerate(lines):
            line = line.strip()
            if re.search(r'Debt|Liquid|Money Market|Overnight|Ultra Short', line, re.I):
                current_scheme_type = "Debt"
            elif re.search(r'(?<!Non-)Equity|ELSS|Large|Mid|Small|Multi|Flexi|Balanced|Hybrid', line, re.I):
                current_scheme_type = "Equity"

            # ISIN detection
            isin_m = re.search(r'ISIN\s*:\s*(IN\w+)', line)
            if isin_m:
                current_isin = isin_m.group(1)
                scheme_m = re.match(r'(.+?)\s*,\s*ISIN', line)
                if scheme_m:
                    current_scheme = re.sub(r'\s*\(.*?\)', '', scheme_m.group(1)).strip()
                continue

            isin2 = re.search(r'\b(IN[A-Z0-9]{10})\b', line)
            if isin2 and re.search(r'Fund|ETF|Scheme|Plan', line, re.I):
                current_isin = isin2.group(1)
                current_scheme = re.sub(r'\s*\(.*?\)', '', line).split('ISIN')[0].strip()
                continue

            redemption_match = redemption_re.search(line)
            if redemption_match and current_scheme and current_isin:
                sale_date = parse_date(redemption_match.group(1))
                redeemed_units = to_float(redemption_match.group(2))
                sale_amount = to_float(redemption_match.group(3))
                sale_price = to_float(redemption_match.group(4))

                purchase_lines = []
                inline_pi = re.search(
                    r'(?:Switch In|Switch-In|Equity Switch In|SIP Purchase|Purchase|IDCW Reinvest(?:ment)?|Transfer In)\s+'
                    r'(\d{2}-\w{3}-\d{4}|\d{2}-\w+-\d{4})\s+([\d,\.]+)\s+([\d,\.]+)\s+([\d,\.]+)', line
                )
                if inline_pi:
                    purchase_lines.append(line[inline_pi.start():])

                j = i + 1
                while j < len(lines):
                    l2 = lines[j].strip()
                    if re.match(r'(Redemption|Switch.Out)\s+', l2) or re.match(r'Total\s+', l2): break
                    if re.match(r'(Switch In|Switch-In|Equity Switch In|SIP Purchase|Purchase|Systematic|IDCW Reinvest|Transfer In|Bonus)\s+', l2, re.I):
                        purchase_lines.append(l2)
                    j += 1

                for pl in purchase_lines:
                    pm = re.match(
                        r'(?:Switch In|Switch-In|Equity Switch In|SIP Purchase|Purchase|Systematic(?:\s+Investment)?|IDCW Reinvest(?:ment)?|Transfer In|Bonus)\s+'
                        r'(\d{2}-\w{3}-\d{4}|\d{2}-\w+-\d{4})\s+([\d,\.]+)\s+([\d,\.]+)\s+([\d,\.]+)'
                        r'(?:\s+([\d,\.]+)\s+([\d,\.]+))?', pl
                    )
                    if not pm: continue
                    purchase_date = parse_date(pm.group(1))
                    lot_units = to_float(pm.group(3))
                    unit_cost = to_float(pm.group(4))
                    gf_nav = to_float(pm.group(6)) if pm.group(6) else None
                    if not lot_units or not unit_cost: continue

                    # Apply Section 112A grandfathering: use Jan-31-2018 FMV when
                    # purchase pre-dates the cutoff and GF cost exceeds original cost
                    GF_CUTOFF = datetime(2018, 1, 31)
                    effective_unit_cost = unit_cost
                    market_price_31jan2018 = None
                    total_market_value_31jan2018 = None
                    if gf_nav and purchase_date and purchase_date <= GF_CUTOFF and gf_nav > unit_cost:
                        effective_unit_cost = gf_nav
                        market_price_31jan2018 = gf_nav
                        total_market_value_31jan2018 = round(lot_units * gf_nav, 4)

                    lot_sale_amount = (lot_units / redeemed_units * sale_amount) if redeemed_units else 0
                    lot_purchase_amount = lot_units * effective_unit_cost
                    is_equity = (current_scheme_type == "Equity")
                    gain_type = determine_gain_type(purchase_date, sale_date, is_equity)

                    rows.append(make_row(
                        self.source_name, current_scheme, current_isin,
                        sale_date, lot_units, sale_price, round(lot_sale_amount, 4),
                        purchase_date, round(lot_purchase_amount, 4), gain_type, self.client_name,
                        market_price_31jan2018=market_price_31jan2018,
                        total_market_value_31jan2018=total_market_value_31jan2018,
                    ))
        return rows


# ─────────────────────────────────────────────
# NIRMAL BANG PARSER
# ─────────────────────────────────────────────

class NirmalBangParser:
    """Parses Nirmal Bang Short-Term/Long-Term P&L reports.
    Emits one row per individual purchase lot (not per scrip summary).
    """
    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        with open_pdf(pdf_path) as pdf:
            full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        self.client_name = extract_client_name(full_text)
        rows = self._parse_text(full_text)
        # NirmalBang statement totals cover the full broker period, not just extracted lots,
        # so PDF vs extracted comparison would be misleading — extraction stats only.
        self.last_validation = build_validation_report(
            rows, pdf_st=None, pdf_lt=None, pdf_total=None,
            parse_failures=0, parser_name='NirmalBangParser',
        )
        return rows

    def _parse_text(self, text):
        rows = []
        # Normalize soft hyphens (U+00AD) to nothing — they split numbers across lines
        text = text.replace('­', '')
        lines = text.split("\n")
        pending = []   # lot rows buffered until we find their ISIN
        orphan_nums = []  # large numbers from a preceding line (split bought/sold totals)

        skip = {'Bought Quantity', 'Bought Amount', 'Short Term PL', 'Long Term PL',
                'Speculative PL', 'Final Total', 'Capital Gain Type', 'Buy Date',
                'Client Code', 'Client Name', 'Date From', 'NIRMAL BANG',
                'Compliance', 'SEBI Regn', 'Page No', 'Disclaimer',
                'Scrip Code', 'Quantity Cut off', 'Os Purchase', 'Os Sales',
                'Short Term -', 'Long Term -', '^ -', 'No. of',
                'B-2 /', 'CIN Number'}

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # ISIN line → assign to buffered rows and flush
            isin_m = re.search(r'ISIN Code\s*:\s*([A-Z]{2}[A-Z0-9]{10})', line)
            if isin_m:
                isin = isin_m.group(1)
                for r in pending:
                    r["isin"] = isin
                    rows.append(r)
                pending = []
                orphan_nums = []
                continue

            if any(s in line for s in skip):
                orphan_nums = []
                continue

            # Detect orphan-number lines: no letters, no date, contains large numbers
            # These are bought/sold totals that Nirmal Bang PDFs sometimes place on a
            # separate line above the transaction due to column overflow + soft hyphens.
            has_date    = bool(re.search(r'\d{2}/\d{2}/\d{4}', line))
            has_letters = bool(re.search(r'[A-Za-z]', line))
            if not has_date and not has_letters:
                candidates = [to_float(n) for n in re.findall(r'[\d,]+\.?\d*', line)
                              if to_float(n) and to_float(n) > 1000]
                if candidates:
                    orphan_nums = candidates
                    continue

            txn = self._parse_lot_line(line, orphan_nums=orphan_nums)
            orphan_nums = []
            if txn:
                pending.append(txn)

        # Flush remaining (shouldn't happen but safety)
        rows.extend(pending)
        return rows

    def _parse_lot_line(self, line, orphan_nums=None):
        """Parse a single lot-level P&L line.

        Formats seen:
          BuyDate Code Name Days Qty BuyRate BoughtTotal SellRate SoldTotal SellDate [Gains...]
          BuyDate Code Name Days Qty CutoffRate BuyRate BoughtTotal SellRate SoldTotal SellDate [Gains...]
          BuyDate Code Name Days Qty CutoffRate BuyRate BoughtTotal SellDate Gains...  (sell amt in gains)
          (orphan_nums): bought/sold totals were on a preceding line due to PDF column overflow
        """
        # Need exactly 2 dates (buy and sell)
        dates = re.findall(r'\d{2}/\d{2}/\d{4}', line)
        if len(dates) != 2:
            return None

        buy_date  = parse_date(dates[0])
        sell_date = parse_date(dates[1])
        if not buy_date or not sell_date or sell_date < buy_date:
            return None

        first_date_end  = line.index(dates[0]) + len(dates[0])
        sell_date_start = line.rfind(dates[1])

        middle = line[first_date_end:sell_date_start].strip()
        after  = line[sell_date_start + len(dates[1]):].strip()

        # middle: code(digits) name(text) days(digits) nums...
        m = re.match(r'(\d+)\s+(.+?)\s+(\d+)\s+(.*)', middle)
        if not m:
            return None

        scrip_name = m.group(2).strip()
        scrip_name = re.sub(r'\s+\d+(\.\d+)?$', '', scrip_name).strip()

        days = to_float(m.group(3))
        rest = m.group(4)

        # All numbers in the amounts section (middle)
        nums = [to_float(n) for n in re.findall(r'[\d,]+\.?\d*', rest) if to_float(n) is not None]

        # Numbers after sell date (gains, but first positive may be sold_total)
        after_nums = [to_float(g) for g in re.findall(r'-?[\d,]+\.?\d*', after) if to_float(g) is not None]

        # Find amounts — pass after_nums and any orphan totals from preceding line
        result = self._find_amounts(nums, after_nums, orphan_nums=orphan_nums or [])
        if not result:
            return None

        qty, buy_rate, bought_total, sell_rate, sold_total = result

        if not sold_total or sold_total <= 0:
            return None   # open / unsold position

        gain_type = determine_gain_type(buy_date, sell_date, is_equity=True)

        # For cases where broker computed gain uses a different cost basis
        # (e.g. grandfathered cost after splits/bonuses), use gain from after_nums
        # to recalculate purchase_amount so CG = sold - purchase is correct
        # Detect: all after_nums are positive (not losses), and after[0] ≠ sold_total
        effective_purchase = bought_total
        if after_nums and sold_total == bought_total and after_nums[0] > 0:
            # Case C: sell=buy price, gain stored in after — use it to set purchase
            broker_gain = after_nums[0]
            effective_purchase = round(sold_total - broker_gain, 4)

        return make_row(
            self.source_name, scrip_name, '',
            sell_date, qty, sell_rate, sold_total,
            buy_date, effective_purchase, gain_type, self.client_name
        )

    def _find_amounts(self, nums, after_nums=None, orphan_nums=None):
        """Detect qty, buy_rate, bought_total, sell_rate, sold_total.

        Column formats (with cutoff = split-adjusted / grandfathered reference price):
          [qty, buy_r, bought, sell_r, sold]                  no cutoff, sell in middle
          [qty, cutoff, buy_r, bought, sell_r, sold]          with cutoff, sell in middle
          [qty, cutoff, buy_r, bought]  + after=[sold, ...]   sell in after section
          [qty, cutoff, buy_r, bought]  + after=[gain]        sell = bought (same price)
        Key invariant: qty * buy_rate ≈ bought_total
        """
        if len(nums) < 3:
            return None

        after_nums  = after_nums  or []
        orphan_nums = orphan_nums or []
        tol = lambda b: max(2, b * 0.02)   # 2% tolerance for Indian comma formatting

        def try_find(q, r, b, rest_nums, extra_after):
            """Given qty/buy_rate/bought confirmed, find sell_rate & sold_total."""
            if not (b > 0 and abs(r * q - b) < tol(b)):
                return None
            # Case A: sell_rate + sold both in rest_nums
            if len(rest_nums) >= 2:
                s_r, s_b = rest_nums[0], rest_nums[1]
                if s_b > 0:
                    return q, r, b, s_r, s_b
            # Case A': single rest_num is sell_rate → compute sold from qty × sell_rate
            if len(rest_nums) == 1 and rest_nums[0] > 0:
                s_r = rest_nums[0]
                s_b = q * s_r
                return q, r, b, s_r, s_b
            # Case B: sold_total is first number in after_nums
            # ONLY when after has >= 2 numbers (sold + gain/tax)
            # With 1 after number it's the gain, not sold_total
            if len(rest_nums) == 0 and len(extra_after) >= 2 and extra_after[0] > 0:
                s_b = extra_after[0]
                s_r = s_b / q if q > 0 else r
                return q, r, b, s_r, s_b
            # Case C: sell = bought (same price), single after number = gain
            # Purchase is reconstructed later using (sold - gain)
            return q, r, b, r, b

        # Try with cutoff at pos [0,1,2,3,...]: q=0, cutoff=1, buy=2, bought=3
        if len(nums) >= 4:
            q = nums[0]
            if 0 < q <= 1_000_000:
                r, b = nums[2], nums[3]
                res = try_find(q, r, b, nums[4:], after_nums)
                if res:
                    return res

        # Try without cutoff: q=0, buy=1, bought=2
        if len(nums) >= 3:
            q = nums[0]
            if 0 < q <= 1_000_000:
                r, b = nums[1], nums[2]
                res = try_find(q, r, b, nums[3:], after_nums)
                if res:
                    return res

        # Zero-cutoff fallback (ANGEL ONE / bonus-share format):
        # [qty, 0.0, buy_rate, sell_rate, ...] — cutoff field is literally 0
        if len(nums) >= 4 and nums[1] == 0.0 and nums[2] > 0 and nums[3] > 0:
            q, buy_r, sell_r = nums[0], nums[2], nums[3]
            if 0 < q <= 1_000_000:
                return q, buy_r, q * buy_r, sell_r, q * sell_r

        # Orphan fallback: bought/sold totals were on a preceding PDF line
        # orphan_nums[0] = bought_total, orphan_nums[1] = sold_total
        if len(orphan_nums) >= 2 and len(nums) >= 2:
            q = nums[0]
            if 0 < q <= 1_000_000:
                bought_t = orphan_nums[0]
                sold_t   = orphan_nums[1]
                # Find buy_rate: the value in nums satisfying qty × buy_rate ≈ bought_total
                for buy_r in nums[1:]:
                    if buy_r > 0 and abs(buy_r * q - bought_t) < tol(bought_t):
                        sell_r = sold_t / q if q > 0 else buy_r
                        return q, buy_r, bought_t, sell_r, sold_t

        return None



# ─────────────────────────────────────────────
# PURNARTHA INVESTMENT ADVISERS PARSER
# ─────────────────────────────────────────────

class PurnarthaParser:
    """Parses Purnartha Investment Advisers Capital Gain/Loss statements.
    Table-based format: SecurityName-ISIN SaleDate Qty Rate SaleAmt PurchDate Rate N.A PurchAmt EffCost Days ST LT EffGainLT
    """
    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        self.client_name = None
        rows = []
        with open_pdf(pdf_path) as pdf:
            full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)
            # Extract client name from "Account : 5911789 SANIGAPALLY HARINATH REDDY -PURJC114"
            m = re.search(r'Account\s*:\s*\d+\s+([A-Z][A-Z\s]+?)(?:\s*-PURJC|\s*From\s|\s*\n)', full_text, re.I)
            if m:
                self.client_name = m.group(1).strip().title()
            for page in pdf.pages:
                for table in (page.extract_tables() or []):
                    rows.extend(self._parse_table(table))
        self.last_validation = self._build_validation(rows, pdf_path)
        return rows

    def _parse_table(self, table):
        rows = []
        if not table: return rows
        current_asset_type = "Equity"   # default: Listed/Equity section
        for row in table:
            if not row: continue
            cell = str(row[0] or '').strip()
            if not cell: continue

            # ── Section headers — track but don't parse as data ──
            if "Unlisted Shares" in cell or "Non-Equity MF" in cell or "Debt Instrument" in cell:
                # "Shares w/o STT" = listed equity sold off-market → 365-day LT threshold
                # This sub-category is included in the combined header but must use equity rules.
                current_asset_type = "Equity" if "w/o STT" in cell else "Debt"
                continue
            if "Listed Shares" in cell or "Equity Mutual Fund" in cell:
                current_asset_type = "Equity"
                continue

            # Skip other header/summary/section lines
            skip = ['Security', 'Sale Date', 'Capital Gain', 'Total', 'ST Gain',
                    'LT Gain', 'HDFC-IN', 'SEBI', 'This is', 'www.', 'Account :', 'From ', 'Purnartha']
            if any(cell.startswith(s) or s in cell for s in skip):
                continue
            parsed = self._parse_data_row(cell, current_asset_type)
            if parsed:
                rows.append(parsed)
        return rows

    def _parse_data_row(self, cell, asset_type="Equity"):
        """Parse: SecurityName-ISIN DD/MM/YY qty rate amt DD/MM/YY rate N.A amt eff_cost days st lt eff_lt"""
        dates = re.findall(r'\d{2}/\d{2}/\d{2,4}', cell)
        if len(dates) < 2:
            return None
        first_date_pos = cell.find(dates[0])
        name_isin = cell[:first_date_pos].strip()

        # Extract ISIN from end of name segment
        isin_m = re.search(r'-(IN[A-Z0-9]{10,12})\s*$', name_isin)
        if isin_m:
            isin = isin_m.group(1)
            name = name_isin[:isin_m.start()].strip()
        else:
            isin = ''
            name = name_isin

        name = re.sub(r'\s+', ' ', name).strip()
        if not name or len(name) < 3:
            return None

        sale_date = parse_date(dates[0])
        purchase_date = parse_date(dates[1])
        if not sale_date or not purchase_date:
            return None

        # Parse numbers after the name block
        rest = cell[first_date_pos:]
        rest_clean = re.sub(r'\d{2}/\d{2}/\d{2,4}', '', rest)
        rest_clean = rest_clean.replace('N.A', ' ').replace('N.A.', ' ')
        nums = [to_float(n) for n in re.findall(r'-?[\d,]+\.?\d*', rest_clean) if to_float(n) is not None]

        if len(nums) < 5:
            return None

        # Layout: qty, sale_rate, sale_amt, purchase_rate, purchase_amt[, eff_cost, days, st, lt, eff_lt]
        qty        = nums[0]
        sale_rate  = nums[1]
        sale_amt   = nums[2]
        # nums[3] = purchase_rate
        purchase_amt = nums[4]
        days = nums[6] if len(nums) > 6 else None

        if not sale_amt or sale_amt <= 0:
            return None
        if not purchase_amt:
            purchase_amt = 0.0

        is_equity_like = (asset_type != "Debt")
        gain_type = determine_gain_type(purchase_date, sale_date, is_equity=is_equity_like)

        return make_row(
            self.source_name, name, isin,
            sale_date, qty, sale_rate, sale_amt,
            purchase_date, purchase_amt, gain_type, self.client_name,
            asset_type
        )

    # "ST Gain/Loss  q1 q2 q3 q4 q5  TOTAL  sale  eff_cost"
    _ST_GAIN_RE = re.compile(
        r'ST\s+Gain/Loss\s+'
        r'(-?[\d,]+\.?\d*)\s+(-?[\d,]+\.?\d*)\s+(-?[\d,]+\.?\d*)\s+'
        r'(-?[\d,]+\.?\d*)\s+(-?[\d,]+\.?\d*)\s+(-?[\d,]+\.?\d*)',
        re.I,
    )
    _LT_GAIN_RE = re.compile(
        r'LT\s+Gain/Loss\s+'
        r'(-?[\d,]+\.?\d*)\s+(-?[\d,]+\.?\d*)\s+(-?[\d,]+\.?\d*)\s+'
        r'(-?[\d,]+\.?\d*)\s+(-?[\d,]+\.?\d*)\s+(-?[\d,]+\.?\d*)',
        re.I,
    )

    def _extract_pdf_summary(self, pdf_path):
        """Sum ST/LT totals from all Capital Gain/Loss Summary pages (each section has its own page)."""
        try:
            st_total = lt_total = 0.0
            with open_pdf(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text() or ''
                    m_st = self._ST_GAIN_RE.search(text)
                    m_lt = self._LT_GAIN_RE.search(text)
                    if m_st:
                        st_total += float(m_st.group(6).replace(',', ''))
                    if m_lt:
                        lt_total += float(m_lt.group(6).replace(',', ''))
            total = round(st_total + lt_total, 2)
            return (round(st_total, 2), round(lt_total, 2), total) if (st_total or lt_total) else (None, None, None)
        except Exception:
            return None, None, None

    def _build_validation(self, rows, pdf_path):
        pdf_st, pdf_lt, pdf_total = self._extract_pdf_summary(pdf_path)
        return build_validation_report(
            rows, pdf_st=pdf_st, pdf_lt=pdf_lt, pdf_total=pdf_total,
            parse_failures=0, parser_name='PurnarthaParser',
        )


# ─────────────────────────────────────────────
# HDFC BANK REALIZED GAIN/LOSS PARSER
# ─────────────────────────────────────────────

class HDFCBankParser:
    """Parses HDFC Bank Realized Gain/Loss Report PDFs.
    Multi-line table cells per row:
      col0  = Name (multi-line)
      col1  = Instrument Cat
      col2  = ISIN\nRTA Code
      col3  = Folio\nUnits Sold
      col4  = Sell Date
      col5  = Sell Amount\nSell Price
      col6  = Purc Date
      col7  = Purc Cost  Abs.(G/L)\nPurc Price  Holding Days
      col9  = LTCG
      col10 = STCG
    """
    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        self.client_name = None
        rows = []
        with open_pdf(pdf_path) as pdf:
            # Extract client name from page 1 text
            try:
                txt1 = pdf.pages[0].extract_text() or ''
                m = re.search(r'\n([A-Z][A-Z\s]{4,})\n', txt1)
                if m:
                    self.client_name = m.group(1).strip().title()
            except Exception:
                pass
            current_asset_type = "Equity"
            for page in pdf.pages:
                for table in (page.extract_tables() or []):
                    rows.extend(self._parse_table(table, current_asset_type))
        return rows

    def _parse_table(self, table, current_asset_type="Equity"):
        rows = []
        if not table: return rows
        for row in table:
            if not row or all(v is None or str(v).strip() == '' for v in row):
                continue
            first = str(row[0] or '').strip()
            # Track asset type from section headers
            if re.search(r'Debt\s*[-–]\s*Mutual|Debt.*Fund', first, re.I):
                current_asset_type = "Debt"
            elif re.search(r'Equity\s*[-–]\s*Mutual|Equity.*Fund', first, re.I):
                current_asset_type = "Equity"
            # Skip headers / section labels / total rows
            skip_kw = ['Instrument Name', 'Equity - Mutual', 'Debt - Mutual',
                       'Online -', 'Offline', 'Total :', 'Total:', 'Disclaimer',
                       'HDFC BANK', 'Realized Gain', 'Page:', 'Customer ID']
            if any(kw in first for kw in skip_kw): continue
            if len(row) < 8: continue

            # ── Name ──
            name = re.sub(r'\s+', ' ', str(row[0] or '').replace('\n', ' ')).strip()
            name = re.sub(r'[-\s]+$', '', name).strip()
            if not name or len(name) < 3: continue

            # ── ISIN ──
            isin_cell = str(row[2] or '').strip()
            isin = isin_cell.split('\n')[0].strip()
            if not re.match(r'IN[A-Z0-9]{10}', isin): isin = ''

            # ── Units ──
            folio_units = str(row[3] or '').strip()
            units_parts = folio_units.split('\n')
            units = to_float(units_parts[1]) if len(units_parts) > 1 else to_float(units_parts[0])

            # ── Sell date ──
            sell_date = parse_date(str(row[4] or '').strip())
            if not sell_date: continue

            # ── Sell amount & price ──
            sell_parts = str(row[5] or '').split('\n')
            sell_amt   = to_float(sell_parts[0])
            sell_price = to_float(sell_parts[1]) if len(sell_parts) > 1 else None
            if not sell_amt or sell_amt <= 0: continue

            # ── Purchase date ──
            purc_date = parse_date(str(row[6] or '').strip())

            # ── Purchase cost & holding days ──
            purc_nums = [to_float(n) for n in re.findall(r'-?[\d,]+\.?\d*', str(row[7] or '')) if to_float(n) is not None]
            purc_cost     = purc_nums[0] if purc_nums else 0.0
            holding_days  = purc_nums[3] if len(purc_nums) >= 4 else None

            # ── Gain type from LTCG / STCG columns ──
            ltcg = to_float(str(row[9] or '')) if len(row) > 9 else None
            stcg = to_float(str(row[10] or '')) if len(row) > 10 else None

            if ltcg and ltcg != 0:
                gain_type = "Long Term"
            elif stcg and stcg != 0:
                gain_type = "Short Term"
            elif holding_days is not None:
                is_equity = current_asset_type == "Equity"
                gain_type = determine_gain_type(purc_date, sell_date, is_equity) if purc_date else "Short Term"
            else:
                gain_type = determine_gain_type(purc_date, sell_date, current_asset_type == "Equity") if purc_date else "Short Term"

            rows.append(make_row(
                self.source_name, name, isin,
                sell_date, units, sell_price, sell_amt,
                purc_date, purc_cost if purc_cost else 0.0,
                gain_type, self.client_name
            ))
        return rows


class MiraeAssetParser:
    """Parses Mirae Asset Capital Gains Statement PDFs.

    Layout: Section A (Subscriptions) | Section B (Redemptions) | Section C (Gains)
    All three sections appear on the SAME row in pdfplumber's extracted text.

    Section A columns (left side):
      Trxn.Type | Date | Current Units | Source Units | Original Purchase Cost |
      **Original Purchase Amount | [Grandfathered NAV | Grandfathered Cost Value |]
      IT Applicable NAV | IT Applicable Cost Value

    Section B columns (right side):
      Trxn.Type | Date | Units | Amount | Price | Tax Perc | Total Tax

    Section C: Short Term | Indexed Cost | LT With Index | LT Without Index

    Non-GF rows: 10 numeric columns before sale amount → sale_amt at nums[7]
    GF rows:     12 numeric columns before sale amount → sale_amt at nums[9]
    IT Applicable Cost Value is used as purchase_amount (it reflects the
    grandfathered cost basis where applicable).
    """

    # Lines that are definitely NOT data rows
    _SKIP_RE = re.compile(
        r'^(Total\s*[:\s]|Section\s+[ABC]|Trxn\.?\s*Type|Source\s+Scheme|'
        r'Scheme Name|Capital Gain|Grandfathered|Status\s*:|Folio|PAN\s*No|'
        r'Page\s+\d|For the period|This statement|Note\s*:|Income Dist|'
        r'Amount includes|\*\*Amount|For Units|Short Term|Long Term|'
        r'IT Applicable|GrandFathered|Investment Gain|Building on)',
        re.IGNORECASE,
    )

    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        with open_pdf(pdf_path) as pdf:
            full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        self.client_name = self._extract_client(full_text)
        return self._parse_text(full_text)

    def _extract_client(self, text):
        m = re.search(r'Name\s*:\s*([A-Z][A-Z\s]+?)(?:\s*\n|\s+Folio|\s+PAN)',
                      text, re.IGNORECASE)
        if m:
            name = m.group(1).strip()
            if 3 < len(name) < 70:
                return name.title()
        return extract_client_name(text)

    _NON_EQUITY_FOF_KW = ["nasdaq", "nyse", "global x", "global electric",
                          "passive fof", "etf fof", "international fund",
                          "overseas fund", "global fund", "us equity fof"]

    def _parse_text(self, text):
        rows = []
        current_scheme   = None
        current_isin     = None
        current_is_equity = True
        current_asset_type = None

        for raw in text.split("\n"):
            line = raw.strip()
            if not line:
                continue

            # ── Scheme header ─────────────────────────────────────
            isin_m = re.search(r'\(\s*(IN[A-Z0-9]{10})\s*\)', line)
            if isin_m and re.search(r'Fund|ETF|Plan|Scheme', line, re.I):
                current_isin   = isin_m.group(1)
                current_scheme = re.sub(r'\s*\(.*?\)\s*$', '', line).strip()
                nl = current_scheme.lower()
                if any(kw in nl for kw in self._NON_EQUITY_FOF_KW):
                    current_is_equity  = False
                    current_asset_type = "Non-Equity"
                else:
                    current_is_equity  = True
                    current_asset_type = None
                continue

            if not current_scheme:
                continue

            # ── Skip known non-data lines ─────────────────────────
            if self._SKIP_RE.match(line):
                continue

            # ── Must have exactly 2 dates (purchase + sale) ───────
            dates = re.findall(r'\d{2}-\d{2}-\d{4}', line)
            if len(dates) < 2:
                continue
            purchase_date = parse_date(dates[0])
            sale_date     = parse_date(dates[1])
            if not purchase_date or not sale_date:
                continue
            # Sanity: purchase must precede sale
            if sale_date <= purchase_date:
                continue

            # ── Extract all numbers from the line ─────────────────
            clean = re.sub(r'\d{2}-\d{2}-\d{4}', '', line)
            clean = re.sub(r'[A-Za-z/]+', ' ', clean)
            nums = [to_float(n) for n in re.findall(r'[\d,]+\.?\d*', clean)
                    if to_float(n) is not None]

            if len(nums) < 8:
                continue

            # ── Map columns ───────────────────────────────────────
            # Non-GF layout (10 nums before sale section):
            #   [0]units [1]src_units [2]purch_nav [3]purch_amt
            #   [4]it_nav [5]it_cost  [6]sale_units [7]sale_amt
            #   [8]sale_nav ...
            #
            # GF layout (12 nums before sale section):
            #   [0]units [1]src_units [2]purch_nav [3]purch_amt
            #   [4]gf_nav [5]gf_cost  [6]it_nav [7]it_cost
            #   [8]sale_units [9]sale_amt [10]sale_nav ...
            #
            # Heuristic: if nums[5] ≈ nums[7] (IT cost ≈ next block)
            # it's non-GF; if nums[5] ≠ nums[3] (GF cost differs from
            # original purchase) it's GF.
            # Simpler: IT Applicable Cost Value is always our purchase_amount.
            # Non-GF: nums[5] = IT cost = same as nums[3]  → sale_amt = nums[7]
            # GF:     nums[7] = IT cost (different)         → sale_amt = nums[9]

            purchase_units  = nums[0]
            orig_purch_amt  = nums[3]

            # Detect GF: nums[4] and nums[5] are GF nav/cost (present only in GF)
            # Key signal: in GF rows nums[5] != nums[3] (GF cost ≠ original cost)
            # and nums[5] > nums[3] (grandfathered cost is higher for pre-2018 buys)
            is_gf = (len(nums) >= 10 and
                     abs(nums[5] - nums[3]) > 1.0 and
                     nums[5] > 0)

            if is_gf:
                it_cost    = nums[7]    # IT Applicable Cost Value (grandfathered basis)
                sale_units = nums[8]
                sale_amt   = nums[9]
                sale_nav   = nums[10] if len(nums) > 10 else None
                # Use IT cost (grandfathered) as purchase amount
                purchase_amt = it_cost
                # FMV for grandfathering
                fmv_total = nums[5]   # Grandfathered Cost Value
                fmv_price = nums[4]   # Grandfathered NAV
            else:
                it_cost      = nums[5]   # IT Applicable Cost Value = orig purch amt
                sale_units   = nums[6]
                sale_amt     = nums[7]
                sale_nav     = nums[8] if len(nums) > 8 else None
                purchase_amt = orig_purch_amt
                fmv_total    = None
                fmv_price    = None

            if not purchase_amt or purchase_amt <= 0:
                continue
            if not sale_amt or sale_amt <= 0:
                continue

            gain_type = determine_gain_type(purchase_date, sale_date, is_equity=current_is_equity)

            rows.append(make_row(
                self.source_name, current_scheme, current_isin,
                sale_date, purchase_units, sale_nav, round(sale_amt, 4),
                purchase_date, round(purchase_amt, 4), gain_type,
                self.client_name,
                asset_type=current_asset_type,
                market_price_31jan2018=fmv_price,
                total_market_value_31jan2018=fmv_total,
            ))

        return rows




# ─────────────────────────────────────────────
# ZERODHA EXCEL PARSER
# ─────────────────────────────────────────────

class ZerodhaExcelParser:
    def parse(self, excel_path, source_name=None):
        self.source_name = source_name or source_name_from_file(excel_path)
        self.client_name = None
        rows = []
        xl = pd.ExcelFile(excel_path)
        for sheet in xl.sheet_names:
            sl = sheet.lower()
            if any(x in sl for x in ['future','option','commodity','currency','summary','f&o']): continue
            try:
                df = pd.read_excel(excel_path, sheet_name=sheet, header=None)
                rows.extend(self._parse_sheet(df))
            except: continue
        return rows

    def _parse_sheet(self, df):
        rows = []
        hdr_idx = None
        key_cols = {'symbol','scrip','stock','share name','isin','security','script'}
        for i, row in df.iterrows():
            vals = {str(v).lower().strip() for v in row if pd.notna(v)}
            if vals & key_cols and ('sell date' in vals or 'sale date' in vals or 'exit date' in vals):
                hdr_idx = i; break
        if hdr_idx is None: return rows

        hdr_pos = df.index.get_loc(hdr_idx)
        headers = [str(v).strip().lower() if pd.notna(v) else "" for v in df.iloc[hdr_pos]]
        col = {}
        for i, h in enumerate(headers):
            if any(x in h for x in ['scrip','symbol','stock','share name','script','security']): col.setdefault('name',i)
            elif h in ('isin','isin code'): col['isin'] = i
            elif any(x in h for x in ['buy date','purchase date','entry date']): col['buy_date'] = i
            elif any(x in h for x in ['sell date','sale date','exit date']): col['sell_date'] = i
            elif any(x in h for x in ['buy amount','purchase amount','buy value','buy amt']): col['buy_amt'] = i
            elif any(x in h for x in ['sell amount','sale amount','sell value','sell amt']): col['sell_amt'] = i
            elif any(x in h for x in ['buy qty','quantity','qty','shares']): col.setdefault('qty',i)
            elif any(x in h for x in ['buy price','buy rate','purchase price']): col['buy_price'] = i
            elif any(x in h for x in ['sell price','sell rate','sale price']): col['sell_price'] = i
            elif 'short term' in h and 'pl' in h: col['st_pl'] = i
            elif 'long term' in h and 'pl' in h: col['lt_pl'] = i
            elif any(x in h for x in ['holding','days','period']): col['days'] = i

        if 'name' not in col or 'sell_date' not in col: return rows

        def g(row, field):
            idx = col.get(field)
            if idx is None or idx >= len(row): return None
            v = row.iloc[idx]; return None if pd.isna(v) else v

        for pos in range(hdr_pos + 1, len(df)):
            row = df.iloc[pos]
            if row.isna().all(): continue
            name = str(g(row,'name') or '').strip()
            if not name or name.lower() in ('nan','','total'): continue
            sell_date = parse_date(str(g(row,'sell_date') or ''))
            buy_date = parse_date(str(g(row,'buy_date') or ''))
            if not sell_date: continue
            qty = to_float(g(row,'qty'))
            sell_amt = to_float(g(row,'sell_amt'))
            buy_amt = to_float(g(row,'buy_amt'))
            sell_price = to_float(g(row,'sell_price'))
            buy_price = to_float(g(row,'buy_price'))
            days = to_float(g(row,'days'))
            st_pl = to_float(g(row,'st_pl'))
            lt_pl = to_float(g(row,'lt_pl'))

            if sell_amt is None and qty and sell_price: sell_amt = qty * sell_price
            if buy_amt is None and qty and buy_price: buy_amt = qty * buy_price
            # Allow buy_amt=0 (bonus/rights) and sell_amt=0 (delisted/write-off)
            # Only skip if amounts are truly missing (None), not zero
            if sell_amt is None or buy_amt is None: continue

            if days is not None: gain_type = "Long Term" if days > 365 else "Short Term"
            elif lt_pl and lt_pl != 0: gain_type = "Long Term"
            elif buy_date and sell_date: gain_type = "Long Term" if (sell_date-buy_date).days > 365 else "Short Term"
            else: gain_type = "Short Term"

            rows.append(make_row(self.source_name, name, str(g(row,'isin') or '').strip(),
                                sell_date, qty, sell_price, sell_amt, buy_date, buy_amt, gain_type, self.client_name))
        return rows


# ─────────────────────────────────────────────
# UPSTOX EXCEL PARSER
# ─────────────────────────────────────────────

class UpstoxExcelParser:
    def parse(self, excel_path, source_name=None):
        self.source_name = source_name or source_name_from_file(excel_path)
        self.client_name = None
        rows = []
        xl = pd.ExcelFile(excel_path)
        for sheet in xl.sheet_names:
            sl = sheet.lower()
            if any(x in sl for x in ['future','option','commodity','currency','summary','f&o']): continue
            try:
                df = pd.read_excel(excel_path, sheet_name=sheet, header=None)
                rows.extend(self._parse_sheet(df))
            except: continue
        return rows

    def _parse_sheet(self, df):
        rows = []
        hdr_idx = None
        for i, row in df.iterrows():
            vals = [str(v).lower().strip() for v in row if pd.notna(v)]
            if 'scrip name' in vals and ('buy date' in vals or 'sell date' in vals):
                hdr_idx = i; break
        if hdr_idx is None: return rows

        hdr_pos = df.index.get_loc(hdr_idx)
        headers = [str(v).strip().lower() if pd.notna(v) else "" for v in df.iloc[hdr_pos]]
        col = {}
        for i, h in enumerate(headers):
            if h == 'scrip name': col['name'] = i
            elif h in ('isin','isin code'): col['isin'] = i
            elif 'buy date' in h: col['buy_date'] = i
            elif 'buy rate' in h or 'buy price' in h: col['buy_rate'] = i
            elif 'buy amt' in h or 'buy amount' in h: col['buy_amt'] = i
            elif 'sell date' in h: col['sell_date'] = i
            elif 'sell rate' in h or 'sell price' in h: col['sell_price'] = i
            elif 'sell amt' in h or 'sell amount' in h: col['sell_amt'] = i
            elif h in ('qty','quantity'): col['qty'] = i
            elif h == 'days' or 'holding' in h: col['days'] = i
            elif 'short term' in h: col['short_term'] = i
            elif 'long term' in h: col['long_term'] = i
            elif 'speculation' in h: col['speculation'] = i

        if 'name' not in col or 'sell_date' not in col: return rows

        def g(row, field):
            idx = col.get(field)
            if idx is None or idx >= len(row): return None
            v = row.iloc[idx]; return None if pd.isna(v) else v

        for pos in range(hdr_pos + 1, len(df)):
            row = df.iloc[pos]
            if row.isna().all(): continue
            name = str(g(row,'name') or '').strip()
            if not name or name.lower() in ('nan','','scrip name'): continue
            sell_date = parse_date(str(g(row,'sell_date') or ''))
            buy_date = parse_date(str(g(row,'buy_date') or ''))
            if not sell_date: continue

            qty = to_float(g(row,'qty'))
            sell_amt = to_float(g(row,'sell_amt'))
            buy_amt = to_float(g(row,'buy_amt'))
            sell_price = to_float(g(row,'sell_price'))
            buy_rate = to_float(g(row,'buy_rate'))
            days = to_float(g(row,'days'))
            short_term = to_float(g(row,'short_term'))
            long_term = to_float(g(row,'long_term'))
            speculation = to_float(g(row,'speculation'))

            if speculation and speculation != 0 and not short_term and not long_term: continue
            if sell_amt is None and qty and sell_price: sell_amt = qty * sell_price
            if buy_amt is None and qty and buy_rate: buy_amt = qty * buy_rate
            if not sell_amt or not buy_amt: continue

            if days is not None: gain_type = "Long Term" if days > 365 else "Short Term"
            elif long_term and long_term != 0: gain_type = "Long Term"
            elif buy_date and sell_date: gain_type = "Long Term" if (sell_date-buy_date).days > 365 else "Short Term"
            else: gain_type = "Short Term"

            rows.append(make_row(self.source_name, name, str(g(row,'isin') or '').strip(),
                                sell_date, qty, sell_price, sell_amt, buy_date, buy_amt, gain_type, self.client_name))
        return rows


# ─────────────────────────────────────────────
# MF ADDA / STP DEBT PARSER
# ─────────────────────────────────────────────

class MFAddaParser:
    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        with open_pdf(pdf_path) as pdf:
            full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        self.client_name = extract_client_name(full_text)
        return self._parse_text(full_text)

    def _parse_text(self, text):
        rows = []
        lines = text.split("\n")
        current_scheme = None; current_isin = None

        for line in lines:
            line = line.strip()
            isin_m = re.search(r'\(ISIN:\s*(INF\w+)\)', line)
            if isin_m:
                current_isin = isin_m.group(1)
                current_scheme = re.sub(r'\s*\(ISIN:.*?\)', '', line).strip()
                current_scheme = re.sub(r'\s*\(\d{10,}\)', '', current_scheme).strip()
                continue
            if not current_scheme: continue
            if not re.match(r'\d{2}-\d{2}-\d{4}', line): continue
            if 'STP Out' not in line and 'Redemption' not in line: continue

            dates = re.findall(r'\d{2}-\d{2}-\d{4}', line)
            if len(dates) < 2: continue
            purchase_date = parse_date(dates[0]); sale_date = parse_date(dates[1])

            line_no_dates = re.sub(r'\d{2}-\d{2}-\d{4}', '', line)
            nums = [to_float(n) for n in re.findall(r'[\d,]+\.?\d*', line_no_dates) if to_float(n) is not None]
            if len(nums) < 4: continue

            purchase_amount = nums[0]; units = nums[1]; sell_amt = nums[3]
            actual_gain = nums[-1]
            if not sell_amt or sell_amt == 0: sell_amt = purchase_amount + actual_gain
            days = (sale_date - purchase_date).days if sale_date and purchase_date else 0
            gain_type = "Long Term" if days > 365 else "Short Term"
            sale_price = (sell_amt / units) if units and units > 0 else None

            rows.append(make_row(self.source_name, current_scheme, current_isin,
                                sale_date, units, sale_price, round(sell_amt,4),
                                purchase_date, round(purchase_amount,4), gain_type, self.client_name))
        return rows


# ─────────────────────────────────────────────
# GENERIC EXCEL/CSV PARSER
# ─────────────────────────────────────────────

class GenericExcelParser:
    def parse(self, file_path, source_name=None):
        self.source_name = source_name or source_name_from_file(file_path)
        self.client_name = None
        try:
            if file_path.lower().endswith('.csv'):
                df = pd.read_csv(file_path, header=None, encoding='utf-8', on_bad_lines='skip')
            else:
                df = pd.read_excel(file_path, header=None)
            return self._parse_df(df)
        except Exception as e:
            print(f"[WARN] GenericExcelParser: {e}", file=sys.stderr)
            return []

    def _parse_df(self, df):
        rows = []
        hdr_idx = None
        for i, row in df.iterrows():
            vals = [str(v).lower().strip() for v in row if pd.notna(v)]
            has_date = any(any(k in v for k in ['date','dt']) for v in vals)
            has_amt = any(any(k in v for k in ['amount','value','amt','proceeds']) for v in vals)
            has_name = any(any(k in v for k in ['name','scrip','stock','symbol','fund','scheme','security']) for v in vals)
            if has_date and has_amt and has_name: hdr_idx = i; break
        if hdr_idx is None: return rows

        hdr_pos = df.index.get_loc(hdr_idx)
        headers = [str(v).strip().lower() if pd.notna(v) else "" for v in df.iloc[hdr_pos]]
        col = {}
        for i, h in enumerate(headers):
            if any(k in h for k in ['scrip','symbol','stock','name','fund','scheme','security']): col.setdefault('name',i)
            elif 'isin' in h: col['isin'] = i
            elif any(k in h for k in ['buy date','purchase date','entry','acquisition']): col['buy_date'] = i
            elif any(k in h for k in ['sell date','sale date','exit','redemption date']): col['sell_date'] = i
            elif any(k in h for k in ['buy amount','purchase amount','cost','invested']): col['buy_amt'] = i
            elif any(k in h for k in ['sell amount','sale amount','proceeds','redemption amount']): col['sell_amt'] = i
            elif any(k in h for k in ['quantity','qty','units','shares']): col.setdefault('qty',i)
            elif any(k in h for k in ['gain type','term','type']): col['gain_type'] = i

        if 'name' not in col: return rows

        def g(row, field):
            idx = col.get(field)
            if idx is None or idx >= len(row): return None
            v = row.iloc[idx]; return None if pd.isna(v) else v

        for pos in range(hdr_pos+1, len(df)):
            row = df.iloc[pos]
            if row.isna().all(): continue
            name = str(g(row,'name') or '').strip()
            if not name or name.lower() in ('nan','','total','grand total'): continue
            sell_date = parse_date(str(g(row,'sell_date') or ''))
            buy_date = parse_date(str(g(row,'buy_date') or ''))
            if not sell_date and not buy_date: continue
            sell_amt = to_float(g(row,'sell_amt'))
            buy_amt = to_float(g(row,'buy_amt'))
            qty = to_float(g(row,'qty'))
            if not sell_amt or not buy_amt: continue
            gain_type_raw = str(g(row,'gain_type') or '').lower()
            if 'long' in gain_type_raw: gain_type = "Long Term"
            elif 'short' in gain_type_raw: gain_type = "Short Term"
            elif buy_date and sell_date: gain_type = "Long Term" if (sell_date-buy_date).days > 365 else "Short Term"
            else: gain_type = "Short Term"

            rows.append(make_row(self.source_name, name, str(g(row,'isin') or '').strip(),
                                sell_date, qty, None, sell_amt, buy_date, buy_amt, gain_type, self.client_name))
        return rows


# ─────────────────────────────────────────────
# GENERIC PDF PARSER (fallback)
# ─────────────────────────────────────────────

class GenericPDFParser:
    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        with open_pdf(pdf_path) as pdf:
            full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        self.client_name = extract_client_name(full_text)
        rows = []
        with open_pdf(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    rows.extend(self._parse_table(table))
        if not rows:
            rows = self._text_fallback(full_text)
        return rows

    def _parse_table(self, table):
        rows = []
        if not table or len(table) < 2: return rows
        hdr = None; data_start = 1
        for i, row in enumerate(table):
            vals = [str(v).lower().strip() if v else "" for v in row]
            if any('date' in v for v in vals) and any(any(k in v for k in ['amount','value','price']) for v in vals):
                hdr = vals; data_start = i+1; break
        if not hdr: return rows

        for row in table[data_start:]:
            if not row or all(v is None for v in row): continue
            dates=[]; amounts=[]; name=None
            for cell in row:
                if not cell: continue
                cell_str = str(cell).strip()
                d = parse_date(cell_str)
                if d: dates.append(d)
                n = to_float(cell_str)
                if n and abs(n) > 100: amounts.append(n)
                if not name and len(cell_str) > 5 and not re.match(r'[\d/\-,\.]+$', cell_str):
                    name = cell_str
            if len(dates) >= 1 and len(amounts) >= 2 and name:
                buy_date = dates[0] if len(dates)>1 else None
                sell_date = dates[-1]
                buy_amt = min(abs(a) for a in amounts[:2])
                sell_amt = max(abs(a) for a in amounts[:2])
                gain_type = "Long Term" if (buy_date and sell_date and (sell_date-buy_date).days>365) else "Short Term"
                rows.append(make_row(self.source_name, name, "", sell_date, None, None, sell_amt, buy_date, buy_amt, gain_type, self.client_name))
        return rows

    def _text_fallback(self, text):
        rows = []
        for line in text.split('\n'):
            line = line.strip()
            dates = re.findall(r'\d{2}[-/]\d{2}[-/]\d{2,4}', line)
            if len(dates) < 2: continue
            amounts = re.findall(r'[\d,]+\.\d{2}', line)
            if len(amounts) < 2: continue
            name_m = re.match(r'([A-Z][A-Z &\.]+?)\s+\d', line)
            name = name_m.group(1).strip() if name_m else "Unknown"
            buy_date = parse_date(dates[0]); sell_date = parse_date(dates[-1])
            nums = [to_float(a) for a in amounts]
            if len(nums) < 2: continue
            buy_amt = min(nums[:2]); sell_amt = max(nums[:2])
            gain_type = "Long Term" if (buy_date and sell_date and (sell_date-buy_date).days>365) else "Short Term"
            rows.append(make_row(self.source_name, name, "", sell_date, None, None, sell_amt, buy_date, buy_amt, gain_type, self.client_name))
        return rows



# ─────────────────────────────────────────────
# ANGEL BROKING CSV PARSER
# ─────────────────────────────────────────────

class AngelBrokingCSVParser:
    """Parses Angel Broking EQ Capital Gains CSV.
    Format: Stock Symbol | ISIN | Qty | Sale Date | Sale Rate | Sale Value |
            Sale Expenses | Purchase Date | Purchase Rate | Price as on 31st Jan 2018 |
            Purchase Price Considered | Purchase Value | Purchase Expenses | Profit/Loss
    """
    def parse(self, file_path, source_name=None):
        self.source_name = source_name or source_name_from_file(file_path)
        self.client_name = None
        rows = []
        try:
            df = pd.read_csv(file_path, header=None, encoding='utf-8-sig', on_bad_lines='skip')
        except Exception:
            df = pd.read_csv(file_path, header=None, encoding='latin-1', on_bad_lines='skip')

        # Extract client name
        for i, row in df.iterrows():
            cell0 = str(row.iloc[0]).strip().lower()
            if cell0 == 'name':
                self.client_name = str(row.iloc[1]).strip().title()
                break

        # Find header row
        hdr_idx = None
        for i, row in df.iterrows():
            vals = [str(v).strip().lower() for v in row if pd.notna(v)]
            if 'stock symbol' in vals and 'sale date' in vals:
                hdr_idx = i; break
        if hdr_idx is None:
            return rows

        hdr = [str(v).strip().lower() if pd.notna(v) else '' for v in df.iloc[hdr_idx]]
        col = {}
        for i, h in enumerate(hdr):
            if 'stock symbol' in h: col.setdefault('name', i)
            elif h == 'isin': col['isin'] = i
            elif h == 'qty': col['qty'] = i
            elif 'sale date' in h: col['sale_date'] = i
            elif 'sale rate' in h: col['sale_rate'] = i
            elif 'sale value' in h: col['sale_value'] = i
            elif 'sale expense' in h: col['sale_exp'] = i
            elif 'purchase date' in h: col['buy_date'] = i
            elif 'purchase rate' in h: col['buy_rate'] = i
            elif '31st jan' in h or '31/01/2018' in h: col['fmv_price'] = i
            elif 'purchase value' in h: col['buy_value'] = i
            elif 'purchase expense' in h: col['buy_exp'] = i

        current_gain_type = 'Short Term'
        GF_CUTOFF = datetime(2018, 1, 31)

        def g(row, field):
            idx = col.get(field)
            if idx is None or idx >= len(row): return None
            v = row.iloc[idx]
            return None if pd.isna(v) else v

        for pos in range(hdr_idx + 1, len(df)):
            row = df.iloc[pos]
            first = str(row.iloc[0] if pd.notna(row.iloc[0]) else '').strip()

            if 'short term' in first.lower(): current_gain_type = 'Short Term'; continue
            if 'long term'  in first.lower(): current_gain_type = 'Long Term';  continue
            if not first or 'sub total' in first.lower() or 'grand total' in first.lower():
                continue

            name = str(g(row, 'name') or '').strip()
            if not name or name.lower() in ('nan', ''): continue

            sale_date = parse_date(str(g(row, 'sale_date') or ''))
            buy_date  = parse_date(str(g(row, 'buy_date')  or ''))
            if not sale_date: continue

            qty        = to_float(g(row, 'qty'))
            sale_rate  = to_float(g(row, 'sale_rate'))
            sale_value = to_float(g(row, 'sale_value'))
            buy_rate   = to_float(g(row, 'buy_rate'))
            buy_value  = to_float(g(row, 'buy_value'))
            sale_exp   = to_float(g(row, 'sale_exp'))  or 0
            buy_exp    = to_float(g(row, 'buy_exp'))   or 0
            fmv_price  = to_float(g(row, 'fmv_price'))

            if not sale_value:
                if qty and sale_rate: sale_value = qty * sale_rate
                else: continue
            if not buy_value or buy_value == 0:
                if qty and buy_rate and buy_rate != 0: buy_value = qty * buy_rate
                else: buy_value = 0.0

            # FMV on 31-Jan-2018 applies only for pre-cutoff purchases
            fmv_total = None
            if fmv_price and fmv_price > 0 and qty and buy_date and buy_date <= GF_CUTOFF:
                fmv_total = round(fmv_price * qty, 4)

            transfer_exp = (sale_exp + buy_exp) if (sale_exp or buy_exp) else None

            rows.append(make_row(
                self.source_name, name, str(g(row, 'isin') or '').strip(),
                sale_date, qty, sale_rate, sale_value,
                buy_date, buy_value, current_gain_type, self.client_name,
                market_price_31jan2018=fmv_price if fmv_total else None,
                total_market_value_31jan2018=fmv_total,
                transfer_expenses=transfer_exp,
            ))
        return rows


# ─────────────────────────────────────────────
# CONSOLIDATED MF REPORT PARSER
# ─────────────────────────────────────────────

class ConsolidatedMFParser:
    """Parses consolidated capital gains XLSX from various MF aggregators.
    Sheet: TransactionDetails
    Cols: Fund | FundName | Folio | SchemeName | BuyType | BuyDate | Units |
          GrandfatheredNAV | GrandFatheredCostValue | SellType | SellDate | Units |
          SellAmount | SellPrice | TaxPerc | Tax | ShortTerm | IndexedCost |
          LTWithIndex | LTWithoutIndex
    """
    def parse(self, file_path, source_name=None):
        self.source_name = source_name or source_name_from_file(file_path)
        self.client_name = None
        rows = []
        xl = pd.ExcelFile(file_path)
        sheet = next((s for s in xl.sheet_names if s.lower() == 'transactiondetails'), xl.sheet_names[0])
        df = pd.read_excel(file_path, sheet_name=sheet, header=None)

        # Extract client name
        for i, row in df.iterrows():
            cell = str(row.iloc[0] if pd.notna(row.iloc[0]) else '')
            if 'Name :' in cell or 'Name:' in cell:
                self.client_name = re.sub(r'Name\s*:\s*', '', cell).strip().title()
                break

        # Find header row
        hdr_idx = None
        for i, row in df.iterrows():
            vals = [str(v).strip().lower() for v in row if pd.notna(v)]
            if 'fund' in vals and 'units' in vals and any('date' in v for v in vals):
                hdr_idx = i; break
        if hdr_idx is None: return rows

        hdr = [str(v).strip().lower() if pd.notna(v) else '' for v in df.iloc[hdr_idx]]
        # Map columns by position since headers can repeat (buy/sell)
        col = {}
        buy_date_found = False
        sell_type_found = False
        for i, h in enumerate(hdr):
            if 'scheme name' in h or 'fund name' in h: col.setdefault('scheme', i)
            elif 'isin' in h: col['isin'] = i
            elif h == 'folio number' or h == 'folio': col['folio'] = i
            elif ('date' in h) and not buy_date_found:
                col['buy_date'] = i; buy_date_found = True
            elif ('date' in h) and buy_date_found:
                col['sell_date'] = i
            elif 'current units' in h or (h == 'units' and 'buy_units' not in col):
                col['buy_units'] = i
            elif 'grandfathered nav' in h or 'grandfather' in h: col.setdefault('gf_nav', i)
            elif 'grandfathered cost' in h or 'grandfather' in h and 'cost' in h: col['gf_cost'] = i
            elif 'short term' in h: col['st_gain'] = i
            elif 'lt without' in h or 'long term without' in h: col['lt_gain'] = i
            elif 'long term' in h and 'index' not in h: col.setdefault('lt_gain', i)
            elif 'amount' in h and 'sell' not in col.get('scheme_name',''):
                col['sell_amt'] = i

        # Fallback positional mapping if keyword matching missed cols
        # Col layout (0-indexed): 0=Fund, 1=FundName, 2=Folio, 3=SchemeName,
        # 4=BuyType, 5=BuyDate, 6=Units, 7=GfNAV, 8=GfCostVal,
        # 9=SellType, 10=SellDate, 11=SellUnits, 12=SellAmt, 13=SellPrice,
        # 14=TaxPerc, 15=Tax, 16=ShortTerm, 17=IndexedCost, 18=LTWithIdx, 19=LTWithoutIdx
        SCHEME_COL = 3; BUY_DATE_COL = 5; UNITS_COL = 6
        GF_NAV_COL = 7; GF_COST_COL = 8
        SELL_DATE_COL = 10; SELL_AMT_COL = 12
        ST_GAIN_COL = 16; LT_GAIN_COL = 19

        def gp(row, c):
            if c >= len(row): return None
            v = row.iloc[c]
            return None if pd.isna(v) else v

        for pos in range(hdr_idx + 1, len(df)):
            row = df.iloc[pos]
            if row.isna().all(): continue

            scheme = str(gp(row, SCHEME_COL) or '').strip()
            if not scheme or scheme.lower() in ('nan','','scheme name','fund name'): continue
            # Strip ISIN from scheme name (e.g. "INF200K01QX4/SBI Large Cap Fund")
            isin_m = re.match(r'(IN[A-Z0-9]{10,12})\s*/\s*(.*)', scheme)
            if isin_m:
                isin = isin_m.group(1)
                scheme = isin_m.group(2).strip()
            else:
                isin = ''

            buy_date  = parse_date(str(gp(row, BUY_DATE_COL) or ''))
            sell_date = parse_date(str(gp(row, SELL_DATE_COL) or ''))
            if not sell_date: continue

            units     = to_float(gp(row, UNITS_COL))
            gf_cost   = to_float(gp(row, GF_COST_COL)) or 0
            sell_amt  = to_float(gp(row, SELL_AMT_COL))
            st_gain   = to_float(gp(row, ST_GAIN_COL)) or 0
            lt_gain   = to_float(gp(row, LT_GAIN_COL)) or 0
            gf_nav    = to_float(gp(row, GF_NAV_COL))

            if not sell_amt or sell_amt <= 0: continue

            total_cg = st_gain + lt_gain
            # Back-calculate purchase amount from the broker-computed gain
            purchase_amt = round(sell_amt - total_cg, 4)

            gain_type = "Long Term" if abs(lt_gain) > abs(st_gain) else "Short Term"

            # FMV on 31-Jan-2018 only for pre-2018 purchases
            GF_CUTOFF = datetime(2018, 1, 31)
            fmv_total = gf_cost if (gf_cost > 0 and buy_date and buy_date <= GF_CUTOFF) else None

            rows.append(make_row(
                self.source_name, scheme, isin,
                sell_date, units, None, sell_amt,
                buy_date, purchase_amt, gain_type, self.client_name,
                market_price_31jan2018=gf_nav if fmv_total else None,
                total_market_value_31jan2018=fmv_total,
            ))
        return rows


# ─────────────────────────────────────────────
# SHAREKHAN SUMMARY PARSER
# ─────────────────────────────────────────────

class SharekhanSummaryParser:
    """Parses Sharekhan Capital Gain Summary XLSX (security-level summary).
    Sheets: Assets, Equity, Mutual Fund, Derivatives
    Equity cols: SECURITY | ISIN | BUY QTY | BUY VALUE | SELL QTY | SELL VALUE |
                 SPECULATIVE | SHORT TERM | LONG TERM
    MF cols: SCHEME | ISIN | PURCH QTY | PURCH VALUE | INDEXED VALUE |
             REDEEM QTY | REDEEM VALUE | ST CG | LT CG | INDEXED GAIN
    """
    def parse(self, file_path, source_name=None):
        self.source_name = source_name or source_name_from_file(file_path)
        self.client_name = None
        rows = []
        xl = pd.ExcelFile(file_path)

        for sheet in xl.sheet_names:
            sl = sheet.lower()
            if sl in ('assets', 'derivatives'): continue
            try:
                df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                if not self.client_name:
                    for i in range(min(10, len(df))):
                        cell = str(df.iloc[i, 0] if pd.notna(df.iloc[i, 0]) else '')
                        # Pattern: "DEMIS131 - Capital Gain..." — extract client code
                        m = re.search(r'([A-Z0-9]+)\s*-\s*Capital\s+Gain', cell, re.I)
                        if m:
                            self.client_name = m.group(1).strip()
                            break
                rows.extend(self._parse_sheet(df, sl))
            except Exception:
                continue
        return rows

    def _parse_sheet(self, df, sheet_lower):
        rows = []
        # Find header row
        hdr_idx = None
        key_terms = {'security', 'scheme name', 'scheme', 'isin'}
        for i, row in df.iterrows():
            vals = {str(v).strip().lower() for v in row if pd.notna(v)}
            if vals & key_terms and ('short term' in vals or 'short term ' in ' '.join(vals)):
                hdr_idx = i; break
        if hdr_idx is None: return rows

        hdr = [str(v).strip().lower() if pd.notna(v) else '' for v in df.iloc[hdr_idx]]
        col = {}
        for i, h in enumerate(hdr):
            if h in ('security', 'scheme name', 'scheme'): col['name'] = i
            elif h == 'isin': col['isin'] = i
            elif 'sell transaction value' in h or 'redeem transaction value' in h or 'sell value' in h:
                col['sell_val'] = i
            elif 'buy transaction value' in h or 'purchase transaction value' in h or 'buy value' in h:
                col['buy_val'] = i
            elif 'short term' in h and 'cg' not in h: col['st_gain'] = i
            elif 'long term' in h and 'index' not in h: col['lt_gain'] = i
            elif 'capital gain' in h and 'short' in h: col['st_gain'] = i
            elif 'capital gain' in h and 'long' in h: col['lt_gain'] = i

        if 'name' not in col: return rows

        def g(row, field):
            idx = col.get(field)
            if idx is None or idx >= len(row): return None
            v = row.iloc[idx]; return None if pd.isna(v) else v

        for pos in range(hdr_idx + 1, len(df)):
            row = df.iloc[pos]
            if row.isna().all(): continue
            name = str(g(row, 'name') or '').strip()
            if not name or name.lower() in ('nan', '', 'total', 'grand total'): continue
            if re.match(r'^(total|sub.?total|grand)', name, re.I): continue

            sell_val = to_float(g(row, 'sell_val'))
            buy_val  = to_float(g(row, 'buy_val'))
            st_gain  = to_float(g(row, 'st_gain')) or 0
            lt_gain  = to_float(g(row, 'lt_gain')) or 0

            if not sell_val or sell_val == 0: continue

            isin = str(g(row, 'isin') or '').strip()

            # Create separate rows per gain type
            for gain_type, cg in [('Short Term', st_gain), ('Long Term', lt_gain)]:
                if cg == 0: continue
                total_abs = abs(st_gain) + abs(lt_gain)
                if total_abs == 0: continue
                frac = abs(cg) / total_abs
                sv   = round(sell_val * frac, 4)
                bv   = round(sv - cg, 4)
                if bv < 0: bv = 0
                rows.append(make_row(
                    self.source_name, name, isin,
                    None, None, None, sv,
                    None, bv, gain_type, self.client_name,
                ))
        return rows


# ─────────────────────────────────────────────
# GROWW STOCKS PARSER
# ─────────────────────────────────────────────

class GrowwStocksParser:
    """Parses Groww Stocks P&L Report XLSX.
    Sheet: Trade Level (header at first row containing 'Stock name' and 'Buy date')
    Cols: Stock name | ISIN | Quantity | Buy date | Buy price | Buy value |
          Sell date | Sell price | Sell value | Realised P&L | Remark
    """
    def parse(self, file_path, source_name=None):
        self.source_name = source_name or source_name_from_file(file_path)
        self.client_name = None
        rows = []
        xl = pd.ExcelFile(file_path)
        sheet = next((s for s in xl.sheet_names if 'trade level' in s.lower()), xl.sheet_names[0])
        df = pd.read_excel(file_path, sheet_name=sheet, header=None)

        # Client name from top rows
        for i in range(min(5, len(df))):
            row = df.iloc[i]
            cell0 = str(row.iloc[0] if pd.notna(row.iloc[0]) else '').lower()
            if cell0 == 'name':
                self.client_name = str(row.iloc[1] if pd.notna(row.iloc[1]) else '').strip().title()
                break

        # Find header row with 'Stock name' and 'Buy date'
        hdr_idx = None
        for i, row in df.iterrows():
            vals = [str(v).strip().lower() for v in row if pd.notna(v)]
            if 'stock name' in vals and 'buy date' in vals and 'sell date' in vals:
                hdr_idx = i; break
        if hdr_idx is None: return rows

        hdr = [str(v).strip().lower() if pd.notna(v) else '' for v in df.iloc[hdr_idx]]
        col = {}
        for i, h in enumerate(hdr):
            if h == 'stock name': col['name'] = i
            elif h == 'isin': col['isin'] = i
            elif h == 'quantity': col['qty'] = i
            elif h == 'buy date': col['buy_date'] = i
            elif h == 'buy price': col['buy_price'] = i
            elif h == 'buy value': col['buy_value'] = i
            elif h == 'sell date': col['sell_date'] = i
            elif h == 'sell price': col['sell_price'] = i
            elif h == 'sell value': col['sell_value'] = i

        def g(row, field):
            idx = col.get(field)
            if idx is None or idx >= len(row): return None
            v = row.iloc[idx]; return None if pd.isna(v) else v

        for pos in range(hdr_idx + 1, len(df)):
            row = df.iloc[pos]
            if row.isna().all(): continue
            first_cell = str(row.iloc[0] if pd.notna(row.iloc[0]) else '').strip()
            # Stop at the unrealised trades section — those are open positions, not capital gains
            if 'unrealised' in first_cell.lower() or 'unrealized' in first_cell.lower():
                break
            name = str(g(row, 'name') or '').strip()
            if not name or name.lower() in ('nan', '', 'stock name'): continue

            sell_date = parse_date(str(g(row, 'sell_date') or ''))
            buy_date  = parse_date(str(g(row, 'buy_date')  or ''))
            if not sell_date: continue

            qty        = to_float(g(row, 'qty'))
            sell_price = to_float(g(row, 'sell_price'))
            sell_value = to_float(g(row, 'sell_value'))
            buy_price  = to_float(g(row, 'buy_price'))
            buy_value  = to_float(g(row, 'buy_value'))

            if not sell_value and qty and sell_price: sell_value = qty * sell_price
            if not buy_value  and qty and buy_price:  buy_value  = qty * buy_price
            if not sell_value or sell_value <= 0: continue
            if not buy_value: buy_value = 0.0

            gain_type = determine_gain_type(buy_date, sell_date, is_equity=True)
            rows.append(make_row(
                self.source_name, name, str(g(row, 'isin') or '').strip(),
                sell_date, qty, sell_price, sell_value,
                buy_date, buy_value, gain_type, self.client_name,
            ))
        return rows


# ─────────────────────────────────────────────
# MF CAPITAL GAINS REPORT PARSER (Zerodha/Generic with Grandfathered NAV)
# ─────────────────────────────────────────────

class MFCapGainsParser:
    """Parses MF Capital Gains Report XLSX with Grandfathered NAV column.
    Sheet: Capital gains (header row contains 'Grandfathered Nav')
    Cols: Scheme Name | Scheme Code | Category | Folio | Purchase TxId |
          Purchase Date | Matched Qty | Purchase Price | Redeem TxId |
          Redeem Date | Grandfathered Nav | Redeem Price | ST CG | LT CG
    """
    def parse(self, file_path, source_name=None):
        self.source_name = source_name or source_name_from_file(file_path)
        self.client_name = None
        rows = []
        xl = pd.ExcelFile(file_path)
        sheet = next((s for s in xl.sheet_names if 'capital gain' in s.lower()), xl.sheet_names[0])
        df = pd.read_excel(file_path, sheet_name=sheet, header=None)

        # Client name
        for i, row in df.iterrows():
            for j, v in enumerate(row):
                if str(v).strip().lower() == 'name':
                    nxt = row.iloc[j+1] if j+1 < len(row) else None
                    if nxt and str(nxt).strip() not in ('nan',''):
                        self.client_name = str(nxt).strip().title()

        # Find header row with 'Grandfathered Nav'
        hdr_idx = None
        for i, row in df.iterrows():
            vals = [str(v).strip().lower() for v in row if pd.notna(v)]
            if 'scheme name' in vals and 'grandfathered nav' in vals:
                hdr_idx = i; break
        if hdr_idx is None: return rows

        hdr = [str(v).strip().lower() if pd.notna(v) else '' for v in df.iloc[hdr_idx]]
        col = {}
        for i, h in enumerate(hdr):
            if h == 'scheme name': col['scheme'] = i
            elif h == 'scheme code': col['scheme_code'] = i
            elif h == 'folio number': col['folio'] = i
            elif h == 'purchase date': col['buy_date'] = i
            elif h == 'matched quantity': col['qty'] = i
            elif h == 'purchase price': col['buy_price'] = i
            elif h == 'redeem date': col['sell_date'] = i
            elif h == 'grandfathered nav': col['gf_nav'] = i
            elif h == 'redeem price': col['sell_price'] = i
            elif 'short term' in h: col['st_gain'] = i
            elif 'long term' in h: col['lt_gain'] = i

        GF_CUTOFF = datetime(2018, 1, 31)

        def g(row, field):
            idx = col.get(field)
            if idx is None or idx >= len(row): return None
            v = row.iloc[idx]; return None if pd.isna(v) else v

        for pos in range(hdr_idx + 1, len(df)):
            row = df.iloc[pos]
            if row.isna().all(): continue

            scheme = str(g(row, 'scheme') or '').strip()
            if not scheme or scheme.lower() in ('nan', '', 'scheme name', 'total'): continue

            buy_date  = parse_date(str(g(row, 'buy_date')  or ''))
            sell_date = parse_date(str(g(row, 'sell_date') or ''))
            if not sell_date: continue

            qty        = to_float(g(row, 'qty'))
            buy_price  = to_float(g(row, 'buy_price'))
            sell_price = to_float(g(row, 'sell_price'))
            gf_nav     = to_float(g(row, 'gf_nav'))
            st_gain    = to_float(g(row, 'st_gain')) or 0
            lt_gain    = to_float(g(row, 'lt_gain')) or 0

            if not qty or not sell_price: continue
            sell_amt  = round(qty * sell_price, 4)
            if not buy_price: buy_price = 0
            buy_amt   = round(qty * buy_price, 4)
            if sell_amt <= 0: continue

            gain_type = "Long Term" if abs(lt_gain) > 0 else "Short Term"

            # Grandfathering: only for pre-2018 purchases with valid gf_nav
            fmv_total = None
            fmv_price = None
            if gf_nav and gf_nav > 0 and qty and buy_date and buy_date <= GF_CUTOFF:
                fmv_price = gf_nav
                fmv_total = round(qty * gf_nav, 4)

            rows.append(make_row(
                self.source_name, scheme, '',
                sell_date, qty, sell_price, sell_amt,
                buy_date, buy_amt, gain_type, self.client_name,
                market_price_31jan2018=fmv_price,
                total_market_value_31jan2018=fmv_total,
            ))
        return rows


# ─────────────────────────────────────────────
# OPULENCE WEALTH PARSER
# ─────────────────────────────────────────────

class OpulenceWealthParser:
    """Parses Opulence Wealth Realized Capital Gain Report XLSX.
    Cols: Scheme | ISIN | Folio | Purchase Date | Purchase Type | Purchase Price |
          Purchase Amount | Purchase Amount AS ON 31/01/2018 | Adjusted Purchase Price |
          Cost of Acquisition | Units Sold | Sell Date | Sell Type | Sell Price |
          Sell Amount | Actual Capital Gain | Taxable Capital Gain | STT | TDS
    """
    def parse(self, file_path, source_name=None):
        self.source_name = source_name or source_name_from_file(file_path)
        self.client_name = None
        rows = []
        xl = pd.ExcelFile(file_path)
        # Process only the first non-summary sheet (avoid duplicate data across sheets)
        for sheet in xl.sheet_names:
            if sheet.lower() in ('sheet2', 'summary'): continue
            try:
                df = pd.read_excel(file_path, sheet_name=sheet, header=None)
                sheet_rows = self._parse_sheet(df)
                if sheet_rows:
                    rows.extend(sheet_rows)
                    break   # stop after first sheet that yields data
            except Exception:
                continue
        return rows

    def _parse_sheet(self, df):
        rows = []
        # Extract client name — "Client: RAJEEV KASHYAP (ABOPK8600K)" is in col 0 of row 2
        for i in range(min(6, len(df))):
            for col_idx in range(min(4, df.shape[1])):
                cell = str(df.iloc[i, col_idx] if pd.notna(df.iloc[i, col_idx]) else '')
                m = re.search(r'Client\s*:\s*([A-Z][A-Z\s]+?)(?:\s*\(|\s*$)', cell, re.I)
                if m and not self.client_name:
                    self.client_name = m.group(1).strip().title()
                    break

        # Find header row with 'SCHEME' and 'SELL DATE'
        hdr_idx = None
        for i, row in df.iterrows():
            vals = [str(v).strip().lower() for v in row if pd.notna(v)]
            if 'scheme' in vals and ('sell date' in vals or 'sell amount' in vals):
                hdr_idx = i; break
        if hdr_idx is None: return rows

        hdr = [str(v).strip().lower() if pd.notna(v) else '' for v in df.iloc[hdr_idx]]
        col = {}
        for i, h in enumerate(hdr):
            if h == 'scheme': col['scheme'] = i
            elif h == 'isin': col['isin'] = i
            elif h == 'purchase date': col['buy_date'] = i
            elif h == 'purchase amount': col['buy_amt'] = i
            elif 'purchase amount as on' in h or '31/01/2018' in h: col['fmv_total'] = i
            elif h == 'cost of acquisition': col['cost_of_acq'] = i
            elif h == 'units sold': col['qty'] = i
            elif h == 'sell date': col['sell_date'] = i
            elif h == 'sell price': col['sell_price'] = i
            elif h == 'sell amount': col['sell_amt'] = i
            elif h == 'taxable capital gain': col['taxable_cg'] = i

        if 'scheme' not in col or 'sell_date' not in col: return rows

        GF_CUTOFF = datetime(2018, 1, 31)
        current_section = 'Equity'

        def g(row, field):
            idx = col.get(field)
            if idx is None or idx >= len(row): return None
            v = row.iloc[idx]; return None if pd.isna(v) else v

        for pos in range(hdr_idx + 1, len(df)):
            row = df.iloc[pos]
            if row.isna().all(): continue

            first = str(row.iloc[0] if pd.notna(row.iloc[0]) else '').strip()
            # Section headers
            if re.match(r'^(MUTUAL FUNDS|EQUITY|DEBT|SHORT TERM|LONG TERM)', first, re.I):
                if 'debt' in first.lower(): current_section = 'Debt'
                elif 'equity' in first.lower(): current_section = 'Equity'
                continue

            scheme = str(g(row, 'scheme') or '').strip()
            if not scheme or scheme.lower() in ('nan', '', 'scheme', 'total'): continue
            if re.match(r'^(total|sub.?total|grand)', scheme, re.I): continue

            buy_date  = parse_date(str(g(row, 'buy_date')  or ''))
            sell_date = parse_date(str(g(row, 'sell_date') or ''))
            if not sell_date: continue

            qty       = to_float(g(row, 'qty'))
            sell_price= to_float(g(row, 'sell_price'))
            sell_amt  = to_float(g(row, 'sell_amt'))
            buy_amt   = to_float(g(row, 'buy_amt'))
            fmv_total = to_float(g(row, 'fmv_total'))
            taxable_cg= to_float(g(row, 'taxable_cg'))

            if not sell_amt or sell_amt <= 0: continue
            if not buy_amt: buy_amt = 0.0

            # FMV only applies for pre-2018 purchases
            fmv_for_row = None
            if fmv_total and fmv_total > 0 and buy_date and buy_date <= GF_CUTOFF:
                if abs(fmv_total - buy_amt) > 0.01:  # only if different from purchase
                    fmv_for_row = fmv_total

            is_equity = current_section == 'Equity'
            gain_type = determine_gain_type(buy_date, sell_date, is_equity)

            rows.append(make_row(
                self.source_name, scheme, str(g(row, 'isin') or '').strip(),
                sell_date, qty, sell_price, sell_amt,
                buy_date, buy_amt, gain_type, self.client_name,
                total_market_value_31jan2018=fmv_for_row,
            ))
        return rows


# ─────────────────────────────────────────────
# MOTILAL OSWAL — P&L Transaction Statement PDF
# ─────────────────────────────────────────────

class MotilalOswalPnLParser:
    """Parses Motilal Oswal 'Profit and Loss Transaction Statement (All)' PDF.

    Column layout (equity only):
      Instrument Name | Qty | Purchase Date | Purchase Price | Purchase Cost |
      Purchase Value | Sell Date | Sell Price | Sell Value | Holding Period |
      G/L | sellstt | PURSTT | STT | ISIN

    Detection keywords: 'Profit and Loss Transaction Statement' + 'motilaloswal'
    """

    # Indian ISINs always start with "IN" and always contain digits.
    # Simple pattern + digit check avoids false matches on "INTERNATIONAL",
    # "INFRASTRUCTURE" (long words starting with IN, no digits).
    _ISIN_RE  = re.compile(r'(IN[A-Z0-9]{10})')
    _FULL_DT  = re.compile(r'(\d{1,2})[\s ]*(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[\s ]*(\d{4})', re.I)
    _PART_DT  = re.compile(r'(\d{1,2})[\s ]*(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b', re.I)
    _YEAR_RE  = re.compile(r'\b(20\d{2})\b')
    _SKIP_ROW = re.compile(
        r'Profit|and Loss|Transaction Statement|Purchase Purchase|Instrument Name|'
        r'Date Price Cost|Holding Long|Short Term|Long Term|Family |Client |'
        r'Trading Code|ACCRUED|PRODUCT|Summary|BR/BA|SB Name|SPEC GAIN|Equity\s+0',
        re.I
    )
    _MONTHS = {m: i for i, m in enumerate(
        ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec'], 1)}

    def _find_isin(self, text):
        """Return the first real ISIN match (has ≥3 digits — rejects company-name words
        like INTERNATIONAL, INFRASTRUCTURE that start with IN but have no digits)."""
        for m in self._ISIN_RE.finditer(str(text or '')):
            if sum(c.isdigit() for c in m.group(1)) >= 3:
                return m
        return None

    # Page-1 Summary table: "Equity ... ST_GAIN LT_GAIN ... TOTAL ..."
    _SUMMARY_RE = re.compile(
        r'Equity\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+'
        r'([\d.,]+)\s+([\d.,]+)\s+[\d.,]+\s+([\d.,]+)',
        re.I
    )

    def _extract_pdf_summary(self, pdf_path):
        """Read page-1 Summary table and return (st_gain, lt_gain, total_gain)."""
        try:
            with open_pdf(pdf_path) as pdf:
                text = pdf.pages[0].extract_text() or ''
            m = self._SUMMARY_RE.search(text)
            if m:
                def _n(s): return float(s.replace(',', ''))
                return _n(m.group(1)), _n(m.group(2)), _n(m.group(3))
        except Exception:
            pass
        return None, None, None

    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        self.client_name = None
        rows = []
        with open_pdf(pdf_path) as pdf:
            first_text = pdf.pages[0].extract_text() or ""
            self.client_name = self._extract_client(first_text)
            for page in pdf.pages:
                rows.extend(self._parse_page(page))
        return rows

    def _extract_client(self, text):
        for pat in [r'Client\s+([A-Z][A-Z\s]+?),', r'Client\s+([A-Z][A-Z\s]+?)\s*\(']:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                name = m.group(1).strip()
                if 3 < len(name) < 70:
                    return name.title()
        return None

    def _mk_date(self, day, mon, yr):
        try:
            return datetime(int(yr), self._MONTHS[str(mon).lower()], int(day))
        except Exception:
            return None

    def _parse_date_full(self, s):
        m = self._FULL_DT.search(str(s or '').replace(' ', ' ').replace('\n', ' '))
        return self._mk_date(m.group(1), m.group(2), m.group(3)) if m else None

    def _to_f(self, s):
        s = re.sub(r'[^0-9.\-]', '', str(s or ''))
        try:
            return float(s)
        except Exception:
            return None

    def _parse_page(self, page):
        rows = []
        tables = page.extract_tables()
        if not tables:
            return rows
        for table in tables:
            for row in table:
                if not row or not row[0]:
                    continue
                cell0 = str(row[0] or '')
                if self._SKIP_ROW.search(cell0):
                    continue
                filled = sum(1 for c in row if c is not None and str(c).strip())
                has_isin = any(self._find_isin(str(c or '')) for c in row)
                if not has_isin:
                    continue
                if filled >= 8:
                    r = self._parse_clean(row)
                elif filled <= 2:
                    r = self._parse_collapsed(cell0)
                    if r is None:
                        self._collapsed_failures = getattr(self, '_collapsed_failures', 0) + 1
                else:
                    r = None
                if r and r.get('sell_value'):
                    rows.append(r)
        return rows

    def _parse_clean(self, row):
        row = list(row)
        while len(row) < 16:
            row.append(None)
        isin_m = self._find_isin(str(row[14] or '') + ' ' + str(row[13] or ''))
        if not isin_m:
            return None
        sell_date = self._parse_date_full(row[6])
        if not sell_date:
            return None
        purch_date = self._parse_date_full(row[2])
        name = str(row[0] or '').replace('\n', ' ').replace(' ', ' ').strip()
        return {
            'name': name, 'isin': isin_m.group(1),
            'purch_date': purch_date, 'sell_date': sell_date,
            'sell_price': self._to_f(row[7]),
            'sell_value': self._to_f(row[8]),
            'pur_cost':   self._to_f(row[4]),
        }

    def _parse_collapsed(self, text):
        """
        Collapsed cells (all data in col[0]) have the structure:
          [part_dates_or_name]\n
          NAME [qty] [full_purch_date] nums... STT{ISIN}\n
          [years_or_name_cont]
        """
        text = text.replace(' ', ' ')
        parts = [p.strip() for p in text.split('\n') if p.strip()]

        # Find data line (contains ISIN)
        data_line = next((p for p in parts if self._find_isin(p)), None)
        if not data_line:
            joined = text.replace('\n', ' ')
            if self._find_isin(joined):
                data_line = joined
                parts = [data_line]
            else:
                return None

        isin_m = self._find_isin(data_line)
        if not isin_m:
            return None
        isin = isin_m.group(1)

        other = [p for p in parts if p != data_line]

        # Categorise other parts: date-fragment lines vs year-only lines vs name lines
        date_frags = []   # list of ('part', day, mon) or ('full', day, mon, yr)
        year_pools = []
        name_parts = []

        for p in other:
            full_dates = self._FULL_DT.findall(p)
            part_dates = self._PART_DT.findall(p)
            years = self._YEAR_RE.findall(p)
            # Line has date content (not just long company-name words)?
            has_only_date = part_dates and not re.search(r'[A-Z]{4,}', re.sub(
                r'\d{1,2}\s*(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)',
                '', p, flags=re.I).strip(), re.I)

            if full_dates:
                for fd in full_dates:
                    date_frags.append(('full', fd[0], fd[1], fd[2]))
            elif part_dates and has_only_date:
                for pd in part_dates:
                    date_frags.append(('part', pd[0], pd[1]))
                year_pools.extend(years)
            elif years and not part_dates and len(p.split()) <= 3:
                year_pools.extend(years)
            else:
                # Name fragment — also harvest any partial dates
                for pd in self._PART_DT.findall(p):
                    date_frags.append(('part', pd[0], pd[1]))
                year_pools.extend(self._YEAR_RE.findall(p))
                name_parts.append(p)

        # Full and partial dates in the data line itself
        full_in_data = self._FULL_DT.findall(data_line)
        data_no_full = self._FULL_DT.sub('', data_line)
        part_in_data = self._PART_DT.findall(data_no_full)
        if not full_in_data:
            year_pools.extend(self._YEAR_RE.findall(data_line))

        # Resolve purch_date and sell_date
        purch_date = sell_date = None
        parts_full_frags = [(d[1], d[2], d[3]) for d in date_frags if d[0] == 'full' and len(d) == 4]
        parts_part_frags = [(d[1], d[2]) for d in date_frags if d[0] == 'part']

        if len(full_in_data) >= 2:
            purch_date = self._mk_date(*full_in_data[0])
            sell_date  = self._mk_date(*full_in_data[1])
        elif len(full_in_data) == 1:
            purch_date = self._mk_date(*full_in_data[0])
            if parts_part_frags and year_pools:
                sell_date = self._mk_date(parts_part_frags[0][0], parts_part_frags[0][1], year_pools[-1])
            elif part_in_data and year_pools:
                sell_date = self._mk_date(part_in_data[0][0], part_in_data[0][1], year_pools[-1])
        elif len(parts_full_frags) >= 2:
            purch_date = self._mk_date(*parts_full_frags[0])
            sell_date  = self._mk_date(*parts_full_frags[1])
        elif len(parts_part_frags) >= 2 and len(year_pools) >= 2:
            purch_date = self._mk_date(parts_part_frags[0][0], parts_part_frags[0][1], year_pools[0])
            sell_date  = self._mk_date(parts_part_frags[1][0], parts_part_frags[1][1], year_pools[1])
        elif len(parts_part_frags) == 1 and year_pools:
            sell_date = self._mk_date(parts_part_frags[0][0], parts_part_frags[0][1], year_pools[-1])
        elif len(parts_full_frags) == 1 and parts_part_frags and year_pools:
            purch_date = self._mk_date(*parts_full_frags[0])
            sell_date  = self._mk_date(parts_part_frags[0][0], parts_part_frags[0][1], year_pools[-1])

        if not sell_date:
            return None

        # Name: prefix from name_parts + leading name in data line
        name_m = re.match(r'^([A-Z][A-Z\s&\(\)\./\-#]+?)(?=\s+\d)', data_line.strip(), re.I)
        name_from_data = re.sub(r'\s+', ' ', name_m.group(1)).strip() if name_m else ''
        full_name = ' '.join(name_parts + ([name_from_data] if name_from_data else [])).strip()
        if not full_name:
            return None

        # Numbers: strip name, all dates, ISIN from data line
        data_clean = data_line[len(name_from_data):].strip() if name_from_data else data_line
        data_clean = self._FULL_DT.sub('', data_clean)
        data_clean = self._PART_DT.sub('', data_clean)
        data_clean = re.sub(self._ISIN_RE.pattern, '', data_clean)
        nums = re.findall(r'-?\d+(?:\.\d+)?', data_clean)
        if len(nums) < 4:
            return None

        return {
            'name':       full_name,
            'isin':       isin,
            'purch_date': purch_date,
            'sell_date':  sell_date,
            'sell_price': self._to_f(nums[4]) if len(nums) > 4 else None,
            'sell_value': self._to_f(nums[5]) if len(nums) > 5 else self._to_f(nums[4]),
            'pur_cost':   self._to_f(nums[2]) if len(nums) > 2 else None,
        }

    def _build_row(self, r):
        """Convert dict from parse helpers to make_row() output."""
        purch_date = r.get('purch_date')
        sell_date  = r.get('sell_date')
        gain_type  = determine_gain_type(purch_date, sell_date, is_equity=True) if purch_date else 'Short Term'
        return make_row(
            self.source_name,
            r['name'],
            r['isin'],
            sell_date,
            None,
            r.get('sell_price'),
            r.get('sell_value'),
            purch_date,
            r.get('pur_cost') or 0.0,
            gain_type,
            self.client_name,
        )

    def _parse_page_rows(self, page):
        raw = self._parse_page(page)
        return [self._build_row(r) for r in raw]

    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        self.client_name = None
        self._collapsed_failures = 0
        rows = []
        with open_pdf(pdf_path) as pdf:
            first_text = pdf.pages[0].extract_text() or ""
            self.client_name = self._extract_client(first_text)
            for page in pdf.pages:
                raw = self._parse_page(page)
                for r in raw:
                    row = self._build_row(r)
                    if row:
                        rows.append(row)
        self.last_validation = self._build_validation(rows, pdf_path)
        return rows

    def _build_validation(self, rows, pdf_path):
        """Delegate to universal build_validation_report using this parser's PDF summary."""
        pdf_st, pdf_lt, pdf_total = self._extract_pdf_summary(pdf_path)
        return build_validation_report(
            rows, pdf_st=pdf_st, pdf_lt=pdf_lt, pdf_total=pdf_total,
            parse_failures=self._collapsed_failures,
            parser_name='MotilalOswalPnLParser',
        )


# ─────────────────────────────────────────────
# MOTILAL OSWAL — GrandFather GainLoss PDF
# ─────────────────────────────────────────────

class MotilalOswalParser:
    """Parses Motilal Oswal 'Equity GrandFather GainLoss Details' PDF.

    Column order in the PDF table:
      Scrip | ISIN | Qty | Purchase Date | Sell Date | Holding |
      Purchase Price | 31-Jan-2018 High | GF Price | Sell Price |
      Value At Cost | Value At Sell | Purstt | Sellstt | GF Gain | Abs Gain | Tax Impl
    """

    GF_CUTOFF = datetime(2018, 1, 31)

    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        self.client_name = None
        rows = []

        with open_pdf(pdf_path) as pdf:
            first_text = pdf.pages[0].extract_text() or ""
            self.client_name = self._extract_client(first_text)

            for page in pdf.pages:
                page_rows = self._parse_page(page)
                rows.extend(page_rows)

        return rows

    # ── client name ──────────────────────────────────────────────
    def _extract_client(self, text):
        for pat in [
            r'Client\s+([A-Z][A-Z\s]+?),',
            r'Client\s+([A-Z][A-Z\s]+?)\s*\(',
            r'Client\s*:\s*([A-Z][A-Z\s]+?)(?:\s*\(|\s*$)',
        ]:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                name = m.group(1).strip()
                if 3 < len(name) < 70:
                    return name.title()
        return None

    # ── page-level dispatcher ─────────────────────────────────────
    def _parse_page(self, page):
        # Strategy 1: pdfplumber table extraction
        try:
            tables = page.extract_tables()
            if tables:
                candidate_rows = []
                for tbl in tables:
                    candidate_rows.extend(self._parse_table(tbl))
                if candidate_rows:
                    return candidate_rows
        except Exception:
            pass

        # Strategy 2: word-position reconstruction
        try:
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if words:
                return self._parse_words(words)
        except Exception:
            pass

        return []

    # ── Strategy 1: table cells ───────────────────────────────────
    @staticmethod
    def _clean_cell(v):
        """Clean a pdfplumber cell value, handling numbers split across column lines."""
        if v is None:
            return ''
        s = str(v).strip()
        # Fix "- 16263.0 2" → "-16263.02" (negative sign separated by newline)
        s = re.sub(r'-\s+(\d)', r'-\1', s)
        # Fix "114317. 92" → "114317.92" (decimal point separated)
        s = re.sub(r'(\d+)\.\s+(\d+)', r'\1.\2', s)
        # Fix "76211.9 5" → "76211.95" (digit(s) after decimal split off)
        s = re.sub(r'(\d+\.\d+)\s+(\d{1,2})\b', r'\1\2', s)
        # Fix split whole-number portions "76211 95" — only when first part ends with digit
        # and second is 1-2 digits (continuation of last digit group from prior column wrap)
        # Be conservative: only join if separated by single space and no other words
        s = re.sub(r'^(-?\d{3,})\s+(\d{1,2})$', r'\1\2', s)
        # Normalise remaining newlines → space
        s = s.replace('\n', ' ').strip()
        return s

    def _parse_table(self, table):
        rows = []
        if not table:
            return rows
        for raw_row in table:
            if not raw_row:
                continue
            cells = [self._clean_cell(v) for v in raw_row]
            r = self._row_from_cells(cells)
            if r:
                rows.append(r)
        return rows

    def _row_from_cells(self, cells):
        """Map a list of cells to a make_row call.
        Expected column positions (0-indexed):
          0=Scrip  1=ISIN  2=Qty  3=Purchase Date  4=Sell Date
          5=Holding  6=Purchase Price  7=31Jan Price  8=GF Price
          9=Sell Price  10=Value Cost  11=Value Sell
          12=Purstt  13=Sellstt  14=GF Gain  15=Abs Gain  16=Tax
        """
        if len(cells) < 12:
            return None

        # Skip header/footer rows
        c0 = cells[0].lower()
        if any(kw in c0 for kw in ['scrip', 'isin', 'quantity', 'total', 'disclaimer', 'purchase']):
            return None

        name = cells[0].strip()
        if not name or len(name) < 2:
            return None

        # ISIN: look in cells[1] first, else scan all cells
        isin = ''
        for c in cells[:4]:
            m = re.search(r'(IN[A-Z0-9]{10})', c.replace(' ', ''))
            if m:
                isin = m.group(1)
                break

        qty        = to_float(cells[2]) if len(cells) > 2  else None
        purch_date = parse_date(cells[3]) if len(cells) > 3  else None
        sell_date  = parse_date(cells[4]) if len(cells) > 4  else None
        fmv_price  = to_float(cells[7]) if len(cells) > 7  else None   # 31-Jan price/unit
        sell_price = to_float(cells[9]) if len(cells) > 9  else None
        val_cost   = to_float(cells[10]) if len(cells) > 10 else None  # Value At Cost Price
        val_sell   = to_float(cells[11]) if len(cells) > 11 else None  # Value At Sell Price
        tax_impl   = cells[16].strip() if len(cells) > 16 else (cells[-1].strip() if cells else '')

        if not sell_date:
            return None
        if not val_sell or val_sell <= 0:
            return None
        if not val_cost:
            val_cost = 0.0

        gain_type = 'Long Term' if 'long' in tax_impl.lower() else 'Short Term'
        if not tax_impl and purch_date and sell_date:
            gain_type = determine_gain_type(purch_date, sell_date, is_equity=True)

        # Grandfathering only for pre-2018 purchases
        fmv_total = None
        if fmv_price and purch_date and purch_date <= self.GF_CUTOFF:
            fmv_total = fmv_price * (qty or 1)

        return make_row(
            self.source_name, name, isin, sell_date, qty, sell_price, val_sell,
            purch_date, val_cost, gain_type, self.client_name,
            market_price_31jan2018=fmv_price if fmv_total else None,
            total_market_value_31jan2018=fmv_total,
        )

    # ── Strategy 2: word-position reconstruction ──────────────────
    def _parse_words(self, words):
        """Group words by y-position into logical rows, then parse."""
        if not words:
            return []

        # Bucket words into rows by top coordinate (5-pt tolerance)
        buckets = {}
        for w in words:
            key = round(w['top'] / 5) * 5
            buckets.setdefault(key, []).append(w)

        # Sort each bucket by x, build text lines
        text_lines = []
        for key in sorted(buckets):
            line_words = sorted(buckets[key], key=lambda w: w['x0'])
            text_lines.append(' '.join(w['text'] for w in line_words))

        # Reconstruct the full page text and parse it
        full = '\n'.join(text_lines)
        return self._parse_text_block(full)

    # ── text-block parser (used by strategy 2) ───────────────────
    def _parse_text_block(self, text):
        """Find records anchored by ISIN patterns in raw extracted text.

        Approach:
          1. Normalise the text (collapse whitespace/newlines within each token)
          2. Walk the token stream; each ISIN starts a new record
          3. Within a record collect: qty, purchase date, sell date,
             all numeric values, and tax label
        """
        rows = []

        # Normalise: join split ISINs (INE009A \n 01021 -> INE009A01021)
        text = re.sub(r'(IN[A-Z0-9]{5,10})\s+([A-Z0-9]{1,7})\b', lambda m:
            m.group(1) + m.group(2) if re.match(r'IN[A-Z0-9]{10}$', m.group(1) + m.group(2))
            else m.group(0), text)

        # Split into tokens preserving structure
        # Treat each line as a unit
        lines = [l.strip() for l in text.splitlines() if l.strip()]

        # --- pass 1: find all ISIN lines and their positions ---
        isin_re = re.compile(r'\b(IN[A-Z0-9]{10})\b')
        isin_positions = []  # list of (line_idx, isin)
        for idx, line in enumerate(lines):
            m = isin_re.search(line.replace(' ', ''))
            if m:
                isin_positions.append((idx, m.group(1)))

        if not isin_positions:
            return rows

        # --- pass 2: for each ISIN, gather surrounding lines ---
        for rec_idx, (isin_line_idx, isin) in enumerate(isin_positions):
            # Company name: line(s) immediately before the ISIN line
            name_lines = []
            look_back = max(0, isin_line_idx - 3)
            for i in range(look_back, isin_line_idx):
                l = lines[i].strip()
                if l and not re.search(r'\d{2}[\-/]\d{2}|\d{4}|^\d+$|IN[A-Z]', l):
                    name_lines.append(l)
            name = ' '.join(name_lines).strip()
            if not name:
                name = lines[isin_line_idx - 1].strip() if isin_line_idx > 0 else ''

            # Data lines: from isin_line_idx to next ISIN (or end)
            next_isin_line = isin_positions[rec_idx + 1][0] if rec_idx + 1 < len(isin_positions) else len(lines)
            data_lines = lines[isin_line_idx:next_isin_line]
            data_text = ' '.join(data_lines)

            # Clean: rejoin split numbers like "76211.9 5" or "- 16263"
            data_text = re.sub(r'(\d+\.\d)\s+(\d)\b', r'\1\2', data_text)  # "76211.9 5" -> "76211.95"
            data_text = re.sub(r'-\s+(\d)', r'-\1', data_text)              # "- 16263"   -> "-16263"
            data_text = re.sub(r'LongTer\s*m\b', 'LongTerm', data_text, flags=re.I)
            data_text = re.sub(r'ShortTer\s*m\b', 'ShortTerm', data_text, flags=re.I)

            # Extract dates: "DD Mon YYYY" or "DD-Mon-YYYY"
            date_re = re.compile(
                r'(\d{1,2})\s*(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*(\d{4})',
                re.IGNORECASE)
            dates_found = date_re.findall(data_text)
            purch_date = parse_date(f"{dates_found[0][0]} {dates_found[0][1]} {dates_found[0][2]}") if len(dates_found) > 0 else None
            sell_date  = parse_date(f"{dates_found[1][0]} {dates_found[1][1]} {dates_found[1][2]}") if len(dates_found) > 1 else None

            if not sell_date:
                continue

            # Extract tax label
            gain_type = 'Long Term'
            if re.search(r'LongTerm', data_text, re.I):
                gain_type = 'Long Term'
            elif re.search(r'ShortTerm', data_text, re.I):
                gain_type = 'Short Term'
            elif purch_date:
                gain_type = determine_gain_type(purch_date, sell_date, is_equity=True)

            # Extract all numbers (including negatives)
            num_re = re.compile(r'-?[\d,]+\.?\d*')
            # Remove date digits and ISIN from data_text to avoid confusion
            clean = date_re.sub('', data_text)
            clean = isin_re.sub('', clean)
            nums = [to_float(n) for n in num_re.findall(clean) if to_float(n) is not None]

            # Expected positional order of numbers in each record (after ISIN):
            # [qty, holding_days, purch_price, jan18_price, gf_price, sell_price,
            #  val_cost, val_sell, purstt, sellstt, gf_gain, abs_gain]
            # We need at minimum 8 numbers to extract useful data
            if len(nums) < 5:
                continue

            qty        = nums[0] if len(nums) > 0 else None
            # holding   = nums[1]  -- skip
            purch_price= nums[2] if len(nums) > 2 else None
            jan18_price= nums[3] if len(nums) > 3 else None  # 31-Jan-2018 high
            # gf_price  = nums[4]  -- grandfathered price per unit
            sell_price = nums[5] if len(nums) > 5 else None
            val_cost   = nums[6] if len(nums) > 6 else None   # Value At Cost Price (total)
            val_sell   = nums[7] if len(nums) > 7 else None   # Value At Sell Price (total)

            if not val_sell or val_sell <= 0:
                continue
            if not val_cost:
                val_cost = 0.0

            # Validate val_sell roughly ≈ qty * sell_price (within 5%)
            if qty and sell_price and val_sell:
                expected = qty * sell_price
                if expected > 0 and abs(val_sell - expected) / expected > 0.10:
                    # Numbers may be shifted; try to find val_sell by scanning
                    for ni in range(5, len(nums)):
                        if qty and sell_price:
                            exp = qty * sell_price
                            if exp > 0 and abs(nums[ni] - exp) / exp < 0.05:
                                val_sell = nums[ni]
                                val_cost = nums[ni - 1] if ni > 0 else val_cost
                                break

            # Grandfathering only for pre-2018 purchases
            fmv_total = None
            if jan18_price and purch_date and purch_date <= self.GF_CUTOFF:
                fmv_total = jan18_price * (qty or 1)

            if not name or len(name) < 2:
                continue

            rows.append(make_row(
                self.source_name, name, isin, sell_date, qty, sell_price, val_sell,
                purch_date, val_cost, gain_type, self.client_name,
                market_price_31jan2018=jan18_price if fmv_total else None,
                total_market_value_31jan2018=fmv_total,
            ))

        return rows


# ─────────────────────────────────────────────
# FRANKLIN TEMPLETON PARSER
# Format: "Investment Gain / (Loss) Statement"
# ─────────────────────────────────────────────

class FranklinParser:
    """Parses Franklin Templeton Investment Gain/(Loss) Statement PDFs.

    Actual pdfplumber text layout (columns are merged left-to-right):
      Redemption line:
        "Redemption Of DATE UNITS SALE_AMT PRICE STT_VALSwitch In From - DATE
         PURCH_U REEM_U UNIT_COST NA LT_GAIN 0.00 0.00 0.00"
        (note: STT value is glued directly to "Switch" with no space)

      Extra lot line (same redemption):
        "Switch In From - DATE PURCH_U REEM_U UNIT_COST NA LT_GAIN"

      Total line: "Total UNITS SALE_TOTAL LT_TOTAL ..."
    """

    _SCHEME_RE = re.compile(
        r'(.+?),?\s*ISIN:\s*(IN[A-Z0-9]{10})', re.IGNORECASE)

    # Full redemption + first lot in one line.
    # "0.00Switch" — STT is glued to "Switch" with no space.
    _RDEM_RE = re.compile(
        r'Redemption Of\s+(\d{1,2}-\w{3}-\d{4})'      # G1 sale_date
        r'\s+([\d,]+\.\d+)'                             # G2 units
        r'\s+([\d,]+\.\d+)'                             # G3 sale_amount
        r'\s+([\d.]+)'                                  # G4 price
        r'\s+[\d.]+Switch In From\s*[-–]\s*'            # STT (glued to Switch)
        r'(\d{1,2}-\w{3}-\d{4})'                       # G5 lot_purchase_date
        r'\s+([\d,]+\.?\d*)'                            # G6 lot_purch_units
        r'\s+([\d,]+\.?\d*)'                            # G7 lot_reem_units
        r'\s+([\d.]+)'                                  # G8 lot_unit_cost
        r'\s+NA\s+([\d,]+\.\d+)',                       # G9 lot_lt_gain
        re.IGNORECASE,
    )

    # Extra lot lines continuing the previous redemption
    _LOT_RE = re.compile(
        r'^Switch In From\s*[-–]\s*(\d{1,2}-\w{3}-\d{4})'  # G1 purchase_date
        r'\s+([\d,]+\.?\d*)'                                 # G2 purch_units
        r'\s+([\d,]+\.?\d*)'                                 # G3 reem_units
        r'\s+([\d.]+)'                                       # G4 unit_cost
        r'\s+NA\s+([\d,]+\.\d+)',                            # G5 lt_gain
        re.IGNORECASE,
    )

    _TOTAL_RE = re.compile(
        r'^Total\s+[\d,]+\.?\d*\s+([\d,]+\.\d+)', re.IGNORECASE)

    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        with open_pdf(pdf_path) as pdf:
            full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        self.client_name = self._extract_client(full_text)
        return self._parse(full_text)

    def _extract_client(self, text):
        m = re.search(
            r'Name\s*:\s*([A-Z][A-Z\s]+?)(?:\s+Status\s*:|\s+PAN\s*:|\s*\n)',
            text)
        if m:
            name = m.group(1).strip()
            if 3 < len(name) < 70:
                return name.title()
        return None

    def _parse(self, text):
        rows         = []
        cur_scheme   = None
        cur_isin     = None
        cur_rdem     = None   # active redemption
        done_rdems   = []     # completed redemptions for this scheme

        for raw in text.split('\n'):
            line = raw.strip()
            if not line:
                continue

            # ── new scheme ────────────────────────────────────────
            sm = self._SCHEME_RE.search(line)
            if sm:
                if cur_rdem:
                    done_rdems.append(cur_rdem); cur_rdem = None
                rows.extend(self._flush(done_rdems, cur_scheme, cur_isin))
                done_rdems = []
                cur_scheme = sm.group(1).strip().rstrip(',').strip()
                cur_isin   = sm.group(2)
                continue

            # ── redemption header (+ first lot) ───────────────────
            rm = self._RDEM_RE.search(line)
            if rm:
                if cur_rdem:
                    done_rdems.append(cur_rdem)
                cur_rdem = {
                    'sale_date':      parse_date(rm.group(1)),
                    'units':          to_float(rm.group(2)),
                    'sale_price':     to_float(rm.group(4)),
                    'sale_amount':    to_float(rm.group(3)),
                    'lt_gain':        to_float(rm.group(9)) or 0.0,
                    'purchase_dates': [],
                }
                pd_ = parse_date(rm.group(5))
                if pd_:
                    cur_rdem['purchase_dates'].append(pd_)
                continue

            # ── extra lot line ────────────────────────────────────
            if cur_rdem is not None:
                lm = self._LOT_RE.match(line)
                if lm:
                    cur_rdem['lt_gain'] += to_float(lm.group(5)) or 0.0
                    pd_ = parse_date(lm.group(1))
                    if pd_:
                        cur_rdem['purchase_dates'].append(pd_)
                    continue

            # ── scheme total → flush ──────────────────────────────
            if self._TOTAL_RE.match(line):
                if cur_rdem:
                    done_rdems.append(cur_rdem); cur_rdem = None
                rows.extend(self._flush(done_rdems, cur_scheme, cur_isin))
                done_rdems = []; cur_scheme = None; cur_isin = None
                continue

        # final flush
        if cur_rdem:
            done_rdems.append(cur_rdem)
        rows.extend(self._flush(done_rdems, cur_scheme, cur_isin))
        return rows

    def _flush(self, rdems, scheme, isin):
        rows = []
        if not rdems or not scheme:
            return rows
        for r in rdems:
            sale_date   = r.get('sale_date')
            sale_amount = r.get('sale_amount')
            lt_gain     = r.get('lt_gain') or 0.0
            if not sale_date or not sale_amount:
                continue

            # purchase_amount derived from explicitly stated gain
            purchase_amount = round(max(sale_amount - lt_gain, 0.0), 2)

            pdates = [d for d in r.get('purchase_dates', []) if d]
            purchase_date = min(pdates) if pdates else None

            if purchase_date and sale_date:
                gain_type = determine_gain_type(purchase_date, sale_date, is_equity=True)
            else:
                gain_type = "Long Term"

            rows.append(make_row(
                self.source_name, scheme, isin or '',
                sale_date, r.get('units'), r.get('sale_price'), sale_amount,
                purchase_date, purchase_amount, gain_type, self.client_name,
            ))
        return rows



# ─────────────────────────────────────────────
# PDF TABLE FALLBACK PARSER (no LLM, no internet)
# ─────────────────────────────────────────────

class PDFTableFallbackParser:
    """Extracts capital gains rows from unknown PDFs using pdfplumber's native
    table detection. No LLM required — works on any machine with no extra setup.
    Uses header keywords + value heuristics to map columns to standard fields.
    """

    _DATE_RE  = re.compile(r'^\d{2}[-/]\d{2}[-/]\d{4}$|^\d{1,2}-[A-Za-z]{3}-\d{4}$', re.I)
    _NUM_RE   = re.compile(r'^[\d,]+\.?\d*$')
    _SKIP_ROW = re.compile(r'^(total|grand total|sub.?total|net|nil|opening|closing|balance)$', re.I)

    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        self.client_name = None
        rows = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                full_text = "\n".join(p.extract_text() or "" for p in pdf.pages[:2])
                self.client_name = extract_client_name(full_text)
                for page in pdf.pages:
                    for settings in [
                        {"vertical_strategy": "lines",  "horizontal_strategy": "lines"},
                        {"vertical_strategy": "text",   "horizontal_strategy": "text"},
                    ]:
                        tables = page.extract_tables(table_settings=settings)
                        if not tables:
                            continue
                        for tbl in tables:
                            rows.extend(self._process_table(tbl))
                        if rows:
                            break
        except Exception as e:
            print(f"[WARN] PDFTableFallbackParser: {e}", file=sys.stderr)

        # Deduplicate across pages
        seen, unique = set(), []
        for r in rows:
            k = (str(r.get("sale_date")), str(r.get("purchase_date")),
                 (r.get("share_name") or "").lower()[:40],
                 round(float(r.get("sale_amount") or 0), 0))
            if k not in seen:
                seen.add(k)
                unique.append(r)
        return unique

    def _process_table(self, table):
        if not table or len(table) < 2:
            return []
        clean = [[str(c or "").strip().replace("\n", " ") for c in row] for row in table]
        if len(clean[0]) < 3:
            return []

        # Find header row — first row with at least 2 non-numeric cells
        header_idx = 0
        for i, row in enumerate(clean[:4]):
            if sum(1 for c in row if c and not self._NUM_RE.match(c.replace(",", ""))) >= 2:
                header_idx = i
                break

        header    = [c.lower() for c in clean[header_idx]]
        data_rows = clean[header_idx + 1:]
        if not data_rows:
            return []

        col_map   = self._detect_columns(header, data_rows)
        required  = {"name", "sale_date", "purchase_date", "sale_amt", "purchase_amt"}
        if not required.issubset(col_map.values()):
            return []

        field_col = {v: k for k, v in col_map.items()}
        rows = []
        for row in data_rows:
            if len(row) <= max(field_col.values()):
                continue
            name = row[field_col["name"]].strip()
            if not name or self._SKIP_ROW.match(name):
                continue

            sale_date     = parse_date(row[field_col["sale_date"]])
            purchase_date = parse_date(row[field_col["purchase_date"]])
            if not sale_date or not purchase_date or sale_date <= purchase_date:
                continue

            sale_amt  = to_float(row[field_col["sale_amt"]].replace(",", ""))
            purch_amt = to_float(row[field_col["purchase_amt"]].replace(",", ""))
            if not sale_amt or sale_amt <= 0 or not purch_amt:
                continue

            qty        = to_float(row[field_col["qty"]].replace(",", "")) if "qty" in field_col else None
            sale_price = to_float(row[field_col["sale_price"]].replace(",", "")) if "sale_price" in field_col else None
            isin       = row[field_col["isin"]].strip() if "isin" in field_col else ""

            if "gain_type" in field_col:
                gt_raw = row[field_col["gain_type"]].lower()
                if "short" in gt_raw:
                    gain_type = "Short Term"
                elif "long" in gt_raw:
                    gain_type = "Long Term"
                else:
                    gain_type = determine_gain_type(purchase_date, sale_date, is_equity=True)
            else:
                gain_type = determine_gain_type(purchase_date, sale_date, is_equity=True)

            rows.append(make_row(
                self.source_name, name, isin,
                sale_date, qty, sale_price, round(sale_amt, 4),
                purchase_date, round(purch_amt, 4),
                gain_type, self.client_name,
            ))
        return rows

    def _detect_columns(self, header, data_rows):
        """Map column index → field type using header keywords then value samples."""
        n = len(header)
        samples = []
        for ci in range(n):
            samples.append([r[ci] for r in data_rows[:8] if ci < len(r) and r[ci]])

        col_map     = {}
        date_cols   = []
        amount_cols = []

        for ci, h in enumerate(header):
            if "isin" in h:
                col_map[ci] = "isin"
            elif any(k in h for k in ("security", "fund", "scheme", "stock", "scrip",
                                       "asset", "instrument", "company", "description")):
                if "name" not in col_map.values():
                    col_map[ci] = "name"
            elif "name" in h and "name" not in col_map.values():
                col_map[ci] = "name"
            elif any(k in h for k in ("qty", "quantity", "units", "shares", "no. of")):
                col_map[ci] = "qty"
            elif any(k in h for k in ("short", "long", "gain type", "type of gain", "term")):
                col_map[ci] = "gain_type"
            elif any(k in h for k in ("sale date", "sell date", "date of sale",
                                       "redemption date", "date of redemption")):
                col_map[ci] = "sale_date"
            elif any(k in h for k in ("purchase date", "buy date", "acquisition date",
                                       "date of purchase", "date of acquisition")):
                col_map[ci] = "purchase_date"
            elif "date" in h or h == "dt":
                date_cols.append(ci)
            elif any(k in h for k in ("sale amount", "sale value", "sell amount",
                                       "proceeds", "consideration", "redemption amount")):
                col_map[ci] = "sale_amt"
            elif any(k in h for k in ("sale price", "sell price", "nav on sale", "selling price")):
                col_map[ci] = "sale_price"
            elif any(k in h for k in ("purchase amount", "cost", "buy amount",
                                       "acquisition amount", "investment amount")):
                col_map[ci] = "purchase_amt"
            elif any(k in h for k in ("amount", "value", "price")):
                amount_cols.append(ci)

        # Value-based date detection for unclassified "date" columns
        unclassified_dates = [ci for ci in date_cols if ci not in col_map
                              and any(self._DATE_RE.match(v) for v in samples[ci])]
        if (len(unclassified_dates) >= 2
                and "purchase_date" not in col_map.values()
                and "sale_date" not in col_map.values()):
            meds = []
            for ci in unclassified_dates[:2]:
                parsed = [parse_date(v) for v in samples[ci] if self._DATE_RE.match(v)]
                parsed = [d for d in parsed if d]
                if parsed:
                    meds.append((sum(d.toordinal() for d in parsed) / len(parsed), ci))
            if len(meds) == 2:
                meds.sort()
                col_map[meds[0][1]] = "purchase_date"
                col_map[meds[1][1]] = "sale_date"

        # Value-based amount detection for unclassified "amount/value/price" columns
        unclassified_amts = [
            ci for ci in amount_cols if ci not in col_map
            and any(to_float(v.replace(",", "")) and to_float(v.replace(",", "")) > 100
                    for v in samples[ci])
        ]
        if (len(unclassified_amts) >= 2
                and "purchase_amt" not in col_map.values()
                and "sale_amt" not in col_map.values()):
            avgs = []
            for ci in unclassified_amts[:3]:
                vals = [to_float(v.replace(",", "")) for v in samples[ci]]
                vals = [v for v in vals if v and v > 0]
                avgs.append((sum(vals) / len(vals) if vals else 0, ci))
            avgs.sort(reverse=True)
            col_map[avgs[0][1]] = "sale_amt"
            if len(avgs) >= 2:
                col_map[avgs[1][1]] = "purchase_amt"

        # Last resort: find name column by text content
        if "name" not in col_map.values():
            for ci in range(n):
                if ci in col_map:
                    continue
                text_hits = sum(
                    1 for v in samples[ci]
                    if v and len(v) > 3
                    and not self._DATE_RE.match(v)
                    and not self._NUM_RE.match(v.replace(",", ""))
                    and re.search(r"[A-Za-z]{3,}", v)
                )
                if text_hits >= 2:
                    col_map[ci] = "name"
                    break

        return col_map


# ─────────────────────────────────────────────
# AI FALLBACK PARSER (Ollama, optional)
# Triggered only when PDFTableFallbackParser also returns 0 rows.
# Silently skipped if Ollama is not installed / not running.
# Configure via env: OLLAMA_URL, OLLAMA_MODEL
# ─────────────────────────────────────────────

class AIFallbackParser:
    """Sends broker statement text to a LOCAL Ollama LLM and asks it to extract
    capital gains rows as structured JSON.  Works for any broker format
    that the specialist parsers don't cover yet.

    Strategy:
      1. Extract full text from file
      2. Chunk it (per-page for PDFs, per-sheet for Excel)
      3. Send each chunk to Ollama and merge + dedupe results
    """

    PROMPT_TEMPLATE = """You are extracting Indian capital gains transactions from a broker/MF statement.

Return ONLY a JSON array of transactions. No explanation. No markdown. No code fences.

Each transaction object MUST have these exact keys:
{{
  "share_name": "Company or Fund Name",
  "isin": "INE..." or "INF..." or "",
  "sale_date": "DD/MM/YYYY",
  "purchase_date": "DD/MM/YYYY",
  "quantity": 100,
  "sale_price": 1500.00,
  "sale_amount": 150000.00,
  "purchase_amount": 120000.00,
  "gain_type": "Long Term" or "Short Term",
  "client_name": "NAME" or ""
}}

CRITICAL RULES:
- Output ONLY a valid JSON array starting with [ and ending with ]
- If no transactions found, output []
- Include EVERY individual lot/transaction (FIFO lots), not aggregates or totals
- Skip header rows, subtotal rows, "Total" rows, disclaimer text
- Equity (ISIN starts with INE) → Long Term if held > 365 days, else Short Term
- Debt/MF (ISIN starts with INF) → Short Term unless purchased before 31/03/2023 AND held > 730 days
- sale_amount = total sale value in INR (qty × sell price, before charges)
- purchase_amount = total cost in INR
- All amounts as plain numbers (no commas, no currency symbols)
- If ISIN not visible, set it to ""
- Dates strictly DD/MM/YYYY

Statement text:
{text}

JSON array:"""

    def parse(self, file_path, source_name=None):
        source = source_name or source_name_from_file(file_path)
        chunks = self._extract_chunks(file_path)
        if not chunks:
            return []

        all_rows = []
        seen_keys = set()  # for dedup across chunks
        for idx, chunk in enumerate(chunks):
            if not chunk or len(chunk.strip()) < 80:
                continue
            try:
                rows = self._call_ollama(chunk, source)
            except Exception as e:
                print(f"[WARN] Ollama chunk {idx+1}/{len(chunks)} failed: {e}", file=sys.stderr)
                continue
            for r in rows:
                # Dedup key: sale_date + name + sale_amount
                k = (
                    r.get("sale_date").strftime("%Y-%m-%d") if r.get("sale_date") else "",
                    (r.get("share_name") or "").strip().lower(),
                    round(float(r.get("sale_amt") or 0), 2),
                    (r.get("purchase_date").strftime("%Y-%m-%d") if r.get("purchase_date") else ""),
                )
                if k in seen_keys:
                    continue
                seen_keys.add(k)
                all_rows.append(r)

        if all_rows:
            print(f"[INFO] Ollama AI fallback extracted {len(all_rows)} rows from "
                  f"{Path(file_path).name} ({len(chunks)} chunk(s))", file=sys.stderr)
        return all_rows

    def _call_ollama(self, text, source, force_json_format=True):
        """One Ollama API call. Returns list of make_row dicts.
        Returns [] silently if Ollama is not installed / not running."""
        import urllib.request
        import urllib.error

        ollama_url   = os.environ.get('OLLAMA_URL', 'http://localhost:11434').rstrip('/')
        ollama_model = os.environ.get('OLLAMA_MODEL', 'llama3.1:8b')

        # Cap chunk text for context window safety
        if len(text) > 12000:
            text = text[:12000] + "\n[...truncated...]"

        prompt = self.PROMPT_TEMPLATE.format(text=text)
        body_dict = {
            "model": ollama_model,
            "prompt": prompt,
            "stream": False,
            "options": {"temperature": 0.0, "num_ctx": 16384}
        }
        if force_json_format:
            body_dict["format"] = "json"
        payload = json.dumps(body_dict).encode()

        req = urllib.request.Request(
            f"{ollama_url}/api/generate",
            data=payload,
            headers={"content-type": "application/json"},
            method="POST"
        )
        try:
            # Local LLMs are slower — give plenty of time
            with urllib.request.urlopen(req, timeout=300) as resp:
                body = json.loads(resp.read().decode())
        except (urllib.error.URLError, OSError):
            # Ollama not installed or not running — skip silently
            print(f"[INFO] Ollama not available — skipping AI fallback", file=sys.stderr)
            return []

        raw = (body.get("response") or "").strip()
        raw = re.sub(r'^```(?:json)?\s*', '', raw, flags=re.MULTILINE)
        raw = re.sub(r'\s*```\s*$', '', raw, flags=re.MULTILINE).strip()

        # Some models wrap the array in {"transactions": [...]} — handle that too
        txns = None
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                txns = parsed
            elif isinstance(parsed, dict):
                # find first list value
                for v in parsed.values():
                    if isinstance(v, list):
                        txns = v
                        break
                # Single transaction object returned instead of array
                if txns is None and 'sale_date' in parsed:
                    txns = [parsed]
        except Exception:
            # Try to recover by extracting the first [...] block
            m = re.search(r'\[[\s\S]*\]', raw)
            if m:
                try: txns = json.loads(m.group(0))
                except: txns = None

        if not isinstance(txns, list):
            return []

        rows = []
        for t in txns:
            if not isinstance(t, dict):
                continue
            sale_date     = parse_date(str(t.get('sale_date', '')))
            purchase_date = parse_date(str(t.get('purchase_date', '')))
            if not sale_date:
                continue
            sale_amt  = to_float(t.get('sale_amount'))
            purch_amt = to_float(t.get('purchase_amount')) or 0.0
            if not sale_amt or sale_amt <= 0:
                continue
            gt = str(t.get('gain_type', '') or '')
            if 'long' in gt.lower():
                gt = 'Long Term'
            elif 'short' in gt.lower():
                gt = 'Short Term'
            elif purchase_date and sale_date:
                # Auto-classify if AI didn't
                isin_raw = str(t.get('isin', '') or '').upper().strip()
                is_equity = isin_raw.startswith('INE')
                gt = determine_gain_type(purchase_date, sale_date, is_equity=is_equity)
            else:
                gt = 'Short Term'
            client = str(t.get('client_name', '') or '').strip()
            rows.append(make_row(
                source,
                str(t.get('share_name', '') or '').strip(),
                str(t.get('isin', '') or '').strip(),
                sale_date,
                to_float(t.get('quantity')),
                to_float(t.get('sale_price')),
                sale_amt,
                purchase_date,
                purch_amt,
                gt,
                client or None,
            ))
        return rows

    def _extract_chunks(self, file_path):
        """Return list of text chunks. PDF→per-page, Excel→per-sheet, CSV→single."""
        ext = Path(file_path).suffix.lower()
        chunks = []
        try:
            if ext == '.pdf':
                with open_pdf(file_path) as pdf:
                    for p in pdf.pages:
                        try:
                            t = p.extract_text() or ""
                        except Exception:
                            t = ""
                        if t.strip():
                            chunks.append(t)
                # If PDF has very few pages, also send a combined view for context
                if 1 < len(chunks) <= 3:
                    chunks = ["\n".join(chunks)]
            elif ext in ('.xlsx', '.xls'):
                xl = pd.ExcelFile(file_path)
                for sh in xl.sheet_names:
                    sl = sh.lower()
                    if any(x in sl for x in ['future','option','commodity','currency']):
                        continue
                    try:
                        df = pd.read_excel(file_path, sheet_name=sh, header=None)
                        chunks.append(f"[Sheet: {sh}]\n{df.to_string(index=False, na_rep='')}")
                    except Exception:
                        continue
            elif ext == '.csv':
                df = pd.read_csv(file_path, nrows=500, encoding='utf-8-sig',
                                 on_bad_lines='skip', header=None)
                chunks.append(df.to_string(index=False))
        except Exception as e:
            print(f"[WARN] AI fallback text extraction failed: {e}", file=sys.stderr)
        return chunks


def attempt_auto_parse(file_path, module_globals=None, extractor_path=None):
    """Auto-fallback called when a matched parser returns 0 rows.

    Order:
      1. PDFTableFallbackParser — pdfplumber native table extraction, no LLM, always works
      2. AIFallbackParser (Ollama) — only if step 1 returns nothing; silently skipped if Ollama not running
    """
    ext = Path(file_path).suffix.lower()

    # Step 1: pdfplumber table extraction (PDFs only, no LLM needed)
    if ext == ".pdf":
        try:
            tbl_parser = PDFTableFallbackParser()
            rows = tbl_parser.parse(file_path)
            if rows:
                print(f"[INFO] PDFTableFallbackParser extracted {len(rows)} rows from "
                      f"{Path(file_path).name}", file=sys.stderr)
                return rows, "PDFTableFallbackParser", None
        except Exception as e:
            print(f"[WARN] PDFTableFallbackParser error: {e}", file=sys.stderr)

    # Step 2: Ollama (optional, silently skipped if not running)
    try:
        parser = AIFallbackParser()
        rows = parser.parse(file_path)
        if rows:
            return rows, "AIFallbackParser(Ollama)", None
        return [], "AIFallbackParser(Ollama)", "No rows extracted (Ollama may not be running or format not recognised)"
    except Exception as e:
        return [], "AIFallbackParser(Ollama)", f"{type(e).__name__}: {e}"


class OCRPDFParser:
    """OCR-based fallback for CID-encoded PDFs.

    Runs two OCR passes (200 DPI for amounts, 300 DPI for dates in split rows),
    preprocesses text to remove table artifacts, then sends to Ollama.
    """

    def parse(self, file_path, source_name=None):
        source = source_name or source_name_from_file(file_path)
        try:
            from pdf2image import convert_from_path
            import pytesseract
        except ImportError as e:
            print(f"[WARN] OCRPDFParser: missing dependency: {e}", file=sys.stderr)
            return []

        try:
            images_200 = convert_from_path(file_path, dpi=200)
            images_300 = convert_from_path(file_path, dpi=300)
        except Exception as e:
            print(f"[WARN] OCRPDFParser: pdf2image failed: {e}", file=sys.stderr)
            return []

        all_200, all_300 = [], []
        for i, (img200, img300) in enumerate(zip(images_200, images_300)):
            try:
                all_200.append(self._preprocess_ocr(pytesseract.image_to_string(img200)))
                all_300.append(self._preprocess_ocr(pytesseract.image_to_string(img300)))
            except Exception as e:
                print(f"[WARN] OCRPDFParser: tesseract failed on page {i+1}: {e}", file=sys.stderr)
                all_200.append("")
                all_300.append("")

        full_200 = "\n\n".join(all_200)
        full_300 = "\n\n".join(all_300)

        # Try regex parser first — precise for known dense-table formats
        try:
            rows = self._regex_parse_cg_table(full_200, full_300, source)
        except Exception as e:
            print(f"[WARN] OCRPDFParser regex parse failed: {e}", file=sys.stderr)
            rows = []

        # Fall back to Ollama if regex got nothing
        if not rows:
            ai = AIFallbackParser()
            page_chunks = []
            for t200, t300 in zip(all_200, all_300):
                combined = t200
                if t300.strip() and t300.strip() != t200.strip():
                    combined = t200 + "\n\n[Higher-DPI pass:]\n" + t300
                has_date    = bool(re.search(r'\d{2}-[A-Za-z]{3}-\d{2}', combined))
                has_amounts = bool(re.search(r'[\d,]{4,}\.\d{2}', combined))
                if len(combined.strip()) >= 80 and has_date and has_amounts:
                    page_chunks.append(combined)
            if page_chunks:
                chunk = "\n\n".join(page_chunks)
                try:
                    rows = ai._call_ollama(chunk, source, force_json_format=False)
                except Exception as e:
                    print(f"[WARN] OCRPDFParser Ollama fallback failed: {e}", file=sys.stderr)

        if rows:
            print(f"[INFO] OCRPDFParser extracted {len(rows)} rows from {Path(file_path).name}", file=sys.stderr)
        return rows

    def _preprocess_ocr(self, text):
        """Clean common OCR artifacts from dense financial tables."""
        text = re.sub(r'(\d)[\]\|](\s)', r'\1\2', text)
        text = re.sub(r'(\d)[\]\|]$', r'\1', text, flags=re.MULTILINE)
        lines = text.split('\n')
        result = []
        i = 0
        while i < len(lines):
            ln = lines[i].strip()
            if re.match(r'^\d{8,}\s+[\d,]+\.\d+\s*$', ln):
                j = i + 1
                while j < len(lines) and not lines[j].strip():
                    j += 1
                if j < len(lines) and re.search(r'\d{2}-[A-Za-z]{3}-\d{2}', lines[j]):
                    result.append(ln + ' ' + lines[j].strip())
                    i = j + 1
                    continue
            result.append(lines[i])
            i += 1
        return '\n'.join(result)

    def _regex_parse_cg_table(self, text_200, text_300, source):
        """Regex-based parser for dense columnar CG tables (e.g. Axis Bank).

        Column order (200 DPI): folio qty purchase_date purchase_nav purchase_val
                                 sale_date [|] sale_nav stt sale_val ...
        Partial rows (split by OCR): folio qty purchase_date sale_val purchase_val
        300 DPI pass fills in missing sale dates for partial rows.
        """
        DATE_RE = r'\d{2}-[A-Za-z]{3}-\d{2,4}'
        FOLIO_QTY_RE = re.compile(r'^(\d{8,})\s+([\d,]+\.\d+)\s*[\|]?\s*(.+)$')

        # Regex for a complete row's "rest" (after folio+qty):
        # purchase_date  purchase_nav  purchase_val  sale_date  [| sep]  sale_nav  stt  sale_val
        FULL_REST_RE = re.compile(
            r'(\d{2}-[A-Za-z]{3}-\d{2,4})\s+'
            r'[\d.]+\s+'
            r'([\d,]+\.\d+)\s+'
            r'(\d{2}-[A-Za-z]{3}-\d{2,4})'
            r'[\s\|]+'
            r'[\d.]+\s+'
            r'[\d.]+\s+'
            r'([\d,]+\.\d+)'
        )
        # Regex for a partial row's "rest": purchase_date  sale_val  purchase_val
        PARTIAL_REST_RE = re.compile(
            r'(\d{2}-[A-Za-z]{3}-\d{2,4})\s+'
            r'([\d,]+\.\d+)\s+'
            r'([\d,]+\.\d+)'
        )
        # 300 DPI: folio qty [|] purchase_date sale_date (two dates on one row)
        DATE_ROW_300_RE = re.compile(
            r'\d{8,}\s+[\d,]+\.\d+\s*[\|]?\s*'
            r'(\d{2}-[A-Za-z]{3}-\d{2,4})\s+'
            r'(\d{2}-[A-Za-z]{3}-\d{2,4})'
        )

        # Build lookup: purchase_date → sale_date from 300 DPI
        dates_300 = {}
        for m in DATE_ROW_300_RE.finditer(text_300):
            pd_str, sd_str = m.groups()
            pd = parse_date(pd_str)
            sd = parse_date(sd_str)
            if pd and sd and pd != sd:
                dates_300[pd] = sd

        fund_name = None
        rows = []
        for line in text_200.split('\n'):
            ln = line.strip()
            if not ln:
                continue
            # Detect fund name: alphabetic line, no digits, reasonable length
            if re.match(r'^[A-Za-z]', ln) and not re.search(r'\d{2}-[A-Za-z]{3}', ln) \
                    and 10 < len(ln) < 100 and 'TOTAL' not in ln.upper():
                fund_name = ln
                continue
            m = FOLIO_QTY_RE.match(ln)
            if not m or not fund_name:
                continue
            folio, qty_str, rest = m.groups()
            qty = to_float(qty_str)
            # Try complete row first
            m2 = FULL_REST_RE.search(rest)
            if m2:
                pd_str, pv_str, sd_str, sv_str = m2.groups()
                rows.append((fund_name, qty, parse_date(pd_str), to_float(pv_str),
                              parse_date(sd_str), to_float(sv_str)))
                continue
            # Try partial row (one date, amounts but no sale date)
            # Column order: purchase_date  purchase_value  sale_value
            m3 = PARTIAL_REST_RE.search(rest)
            if m3:
                pd_str, purchase_val_str, sale_val_str = m3.groups()
                pd = parse_date(pd_str)
                sd = dates_300.get(pd)  # fill from 300 DPI
                rows.append((fund_name, qty, pd, to_float(purchase_val_str), sd, to_float(sale_val_str)))

        result = []
        seen = set()
        for (fund, qty, pd, pv, sd, sv) in rows:
            if not sd or not sv:
                continue
            k = (fund, str(pd), str(sd), round(sv, 2))
            if k in seen:
                continue
            seen.add(k)
            gt = determine_gain_type(pd, sd, is_equity=True)
            result.append(make_row(source, fund, '', sd, qty, None, sv, pd, pv or 0, gt, None))
        return result





# ─────────────────────────────────────────────
# ANAND RATHI / PRIMO DISTRIBUTION MF PARSER
# CAS-style PDF: Units Date NAV Amount [GF NAV GF Amt] Date NAV Amount LT ST Days
# ─────────────────────────────────────────────

class AnandRathiMFParser:
    """Parses Anand Rathi, Primo, Echowin, Alwin-S and similar CAS distributor
    Capital Gain PDFs. Handles three sub-formats:

    A) Anand Rathi "Family Report":
       Scheme: "SchemeName -FolioNo: 123 -ISIN: INF966L01234"
       Lot:    units DD Mon YYYY nav buy_amt 0.00 0.00 DD Mon YYYY nav sell_amt lt st days

    B) Echowin / Primo distributor:
       Scheme: "SchemeName\n(ISIN:" on separate lines
       Lot:    DD-MM-YYYY buy_amt nav DD-MM-YYYY sell_amt ...

    C) Asahoo / Alwin-S:
       Scheme: "SchemeName (ISIN:INFxxx)"
       Lot:    sno DD/MM/YYYY units buy_val acq_val DD/MM/YYYY sell_val stcg ltcg
    """

    _ISIN_RE = re.compile(r'(?:ISIN\s*:?\s*|\[ISIN:\s*)(INF[A-Z0-9]{6,12})', re.I)
    _GF_CUTOFF = datetime(2018, 1, 31)

    # AnandRathi / AssetPlus CAMS grand total: "Total  <ST>  <LT>"
    _TOTAL_RE = re.compile(r'^Total\s+([-\d,]+\.?\d*)\s+([-\d,]+\.?\d*)', re.M)

    def _extract_pdf_summary(self, pdf_path):
        """Read grand-total ST/LT from the last page of AnandRathi MF statements."""
        try:
            with open_pdf(pdf_path) as pdf:
                text = pdf.pages[-1].extract_text() or ''
            m = self._TOTAL_RE.search(text)
            if m:
                st = float(m.group(1).replace(',', ''))
                lt = float(m.group(2).replace(',', ''))
                return round(st, 2), round(lt, 2), round(st + lt, 2)
        except Exception:
            pass
        return None, None, None

    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        self.client_name = None
        with open_pdf(pdf_path) as pdf:
            pages_text = [p.extract_text() or "" for p in pdf.pages]
        full_text = "\n".join(pages_text)
        self.client_name = self._extract_client(pages_text[0] if pages_text else "")
        rows = self._parse_full_text(full_text)
        pdf_st, pdf_lt, pdf_total = self._extract_pdf_summary(pdf_path)
        self.last_validation = build_validation_report(
            rows, pdf_st=pdf_st, pdf_lt=pdf_lt, pdf_total=pdf_total,
            parse_failures=0, parser_name='AnandRathiMFParser',
        )
        return rows

    def _extract_client(self, text):
        SKIP = {'CAPITAL','GAIN','FINANCIAL','YEAR','REPORT','STATEMENT','PAN',
                'AMFI','ADDRESS','MOB','EMAIL','TEL','MUTUAL','FUND','FOLIO',
                'ALWIN','ECHOWIN','PRIMO','VALUEPLUS','TECHNOLOGY','DISTRIBUTOR'}
        for line in text.split("\n")[:10]:
            line = line.strip()
            if not line or len(line) < 4 or len(line) > 60: continue
            if re.match(r'^[A-Za-z][A-Za-z\s\.]+$', line):
                if not any(k in line.upper() for k in SKIP):
                    return line.title()
        m = re.search(r'(?:Client\s*Name|Name)\s*[:\-]\s*([A-Z][A-Z\s\.]{3,50})(?:\n|\(|PAN)',
                      text, re.I)
        if m: return m.group(1).strip().title()
        return None

    def _parse_full_text(self, text):
        rows = []
        current_scheme = None
        current_isin   = None
        current_asset  = "Equity"
        pending_name   = None      # scheme name fragment building up
        pending_lots   = []        # lot lines waiting for ISIN assignment

        for line in text.split("\n"):
            stripped = line.strip()
            if not stripped: continue
            su = stripped.upper()

            # ── Asset type section headers ────────────────────────────
            if re.match(r'^DEBT(\s|$)', su):
                current_asset = "Debt"; continue
            if re.match(r'^EQUITY(\s|$)', su):
                current_asset = "Equity"; continue
            if su in ('DEBT TOTAL','EQUITY TOTAL','GRAND TOTAL',
                      'DEBT SHORT TERM TOTAL','DEBT LONG TERM TOTAL',
                      'EQUITY SHORT TERM TOTAL','EQUITY LONG TERM TOTAL'):
                continue

            # ── ISIN detection → flush pending lots ──────────────────
            isin_m = self._ISIN_RE.search(stripped)
            if isin_m:
                current_isin = isin_m.group(1)
                # Name: text before ISIN marker on same line
                pre = stripped[:isin_m.start()].strip()
                pre = re.sub(r'\s*-?\s*FolioNo\s*:.*$','',pre,flags=re.I).strip()
                pre = re.sub(r'[-\s]+$','',pre).strip()
                if len(pre) > 4:
                    current_scheme = pre
                elif pending_name and len(pending_name) > 4:
                    current_scheme = pending_name
                pending_name = None
                # Infer asset type from fund name when no explicit DEBT/EQUITY section header
                if current_scheme:
                    _sname_u = current_scheme.upper()
                    _debt_kw = ('LIQUID','OVERNIGHT','ULTRA SHORT','SHORT DURATION',
                                'MONEY MARKET','GILT','G-SEC','DYNAMIC BOND',
                                'CORPORATE BOND','FLOATING RATE','FIXED MATURITY',
                                'FMP','BANKING AND PSU','CREDIT RISK','ARBITRAGE',
                                'SHORT TERM FUND','DEBT FUND')
                    _equity_kw = ('EQUITY','ELSS','LARGE CAP','MID CAP','SMALL CAP',
                                  'MULTI CAP','FLEXI CAP','BLUECHIP','BLUE CHIP',
                                  'FLEXICAP','BALANCED ADVANTAGE','HYBRID')
                    if any(k in _sname_u for k in _debt_kw):
                        current_asset = "Debt"
                    elif any(k in _sname_u for k in _equity_kw):
                        current_asset = "Equity"
                # Flush pending lots (from Echowin-style where data came before ISIN)
                for pending_item in pending_lots:
                    # pending_lots items can be 2-tuple (line, asset) or 3-tuple (line, asset, scheme)
                    if len(pending_item) == 3:
                        lot_line, lot_asset, lot_scheme = pending_item
                    else:
                        lot_line, lot_asset = pending_item
                        lot_scheme = None
                    use_scheme = current_scheme or lot_scheme or 'Unknown'
                    lot = self._parse_lot_line(lot_line, use_scheme,
                                               current_isin, lot_asset)
                    if lot: rows.append(lot)
                pending_lots = []
                pending_name = None    # Clear after ISIN assigned — next fund starts fresh
                # Also check if ISIN line itself contains data (merged line)
                rest = stripped[isin_m.end():].strip()
                # Strip bare ISIN continuation lines like "INF109K01431)"
                rest = re.sub(r'^INF[A-Z0-9]{6,12}\)?', '', rest, flags=re.I).strip()
                if rest and re.search(r'\d{2}[-/]\d{2}[-/]\d{2,4}', rest):
                    lot = self._parse_lot_line(rest, current_scheme or 'Unknown',
                                               current_isin, current_asset)
                    if lot: rows.append(lot)
                continue

            # ── Skip obvious non-data lines ───────────────────────────
            SKIP_KW = ['TOTAL','PURCHASE','GRANDFATHERED','UNITS DATE NAV',
                       'DATE NAV AMOUNT','DISCLAIMER','PAGE NO','REPORT -',
                       'FAMILY REPORT','GAIN/LOSS PERIOD','(DAYS)',
                       'ASSET TYPE','PUR. DATE','INITIAL PUR','S.NO',
                       'ACQUISITION VALUE','VALUEPLUS',
                       'COST OF ACQ','ST TOTAL','LG TOTAL','ACTUAL G/L',
                       'STT','TDS','GAIN/LOSS GAIN/LOSS','SUMMARY',
                       'TRANSACTIONS','REDEMPTION','FOLIO NUMBER',
                       'STCG','LTCG','FUND TOTAL','SHORT TERM','LONG TERM',
                       'PAGE NO','(B)','(A)','(C)','(D)','(E)','(F)','(G)',
                       'INITIAL AMOUNT','DUTY)','(31-01-18)','GAIN LOSS','DUTY']
            if any(kw in su for kw in SKIP_KW):
                pending_name = None
                continue

            # ── Check if lot line (has dates + multiple numbers) ─────
            date_pat = r'\d{1,2}[-/]\d{2}[-/]\d{2,4}|\d{1,2}\s+[A-Za-z]{3}\s+\d{4}'
            first_date_m = re.search(date_pat, stripped)
            has_many_nums = len(re.findall(r'[\d,]+\.\d+', stripped)) >= 2

            if first_date_m and has_many_nums:
                # Check for name prefix before the first date (Primo/Sridhar format)
                # e.g. "Regular Plan - Growth(7771751171) 22-08-2023 76109.01..."
                name_prefix = stripped[:first_date_m.start()].strip()
                # Clean junk from prefix: strip folio numbers, trailing punctuation
                name_prefix = re.sub(r'\(\d{5,}\)', '', name_prefix).strip()
                name_prefix = re.sub(r'\s*\(\s*ISIN\s*:?.*$', '', name_prefix, flags=re.I).strip()
                name_prefix = name_prefix.rstrip('-(, ').strip()
                
                # Only use prefix if it looks like a real fund name fragment (has letters, not pure number)
                # Pure numbers before a date = units (e.g. Anand Rathi format "138.28 29 Nov 2023 ...")
                is_real_name = (bool(re.search(r'[A-Za-z]{3,}', name_prefix))
                                and not re.match(r'^[\d,\.\s]+$', name_prefix))
                
                if is_real_name and len(name_prefix) > 3:
                    if pending_name:
                        full_name = pending_name.rstrip('- ').strip() + ' ' + name_prefix
                    else:
                        full_name = name_prefix
                    # Use this as current_scheme for this lot (before ISIN known)
                    current_scheme = full_name
                    pending_name = None
                
                # Parse just the data part (from first date onwards)
                data_part = stripped[first_date_m.start():]
                
                if current_scheme and current_isin:
                    lot = self._parse_lot_line(data_part, current_scheme,
                                               current_isin, current_asset)
                    if lot: rows.append(lot)
                else:
                    # Buffer with the scheme name captured at this moment
                    pending_lots.append((data_part, current_asset, current_scheme))
                continue

            # ── Accumulate scheme name fragments (no date on this line) ──
            if not first_date_m and len(stripped) > 3:
                # Skip bare ISIN-only lines like "INF109K01431)" that leaked out of ISIN detection
                if re.match(r'^INF[A-Z0-9]{6,12}\)?', stripped, re.I):
                    continue
                if pending_name:
                    fragment = stripped.lstrip('- ').strip()
                    pending_name = pending_name + ' ' + fragment
                else:
                    pending_name = stripped
                # Promote to current_scheme if no ISIN yet
                if not current_isin and len(pending_name) > 4:
                    current_scheme = pending_name

        return rows

    def _parse_lot_line(self, line, scheme, isin, asset_type):
        """
        All formats share: (at least) 2 dates, buy_amount between them, sell_amount after.
        Layout:
          A) [units] buy_date [nav] buy_amt [0.00 0.00] sell_date [nav] sell_amt [lt] [st] [days]
          B) buy_date buy_amt [nav] sell_date sell_amt ...
          C) [sno] buy_date [units] buy_val acq_val sell_date sell_val [stcg] [ltcg]
        """
        date_pat = (r'\d{1,2}\s+[A-Za-z]{3}\s+\d{4}'    # "13 Jan 2025"
                    r'|\d{2}[-/]\d{2}[-/]\d{2,4}')         # "14-10-2024"
        dates = re.findall(date_pat, line)
        if len(dates) < 2:
            return None

        buy_date  = parse_date(dates[0])
        sell_date = parse_date(dates[-1])
        if not buy_date or not sell_date or sell_date <= buy_date:
            return None

        # Split line at dates
        pos_d0  = line.find(dates[0])
        pos_d1  = line.rfind(dates[-1])
        before  = line[:pos_d0]
        between = line[pos_d0 + len(dates[0]) : pos_d1]
        after   = line[pos_d1 + len(dates[-1]):]

        def extract_nums(seg):
            return [to_float(n) for n in re.findall(r'-?[\d,]+\.?\d*', seg)
                    if to_float(n) is not None]

        before_nums  = extract_nums(before)
        between_nums = extract_nums(between)
        after_nums   = extract_nums(after)

        # qty: last number before buy_date (may be preceded by serial number)
        qty = None
        if before_nums:
            qty = before_nums[-1]
            # If multiple numbers before date (e.g. "1 149.925"), skip the serial
            if (len(before_nums) >= 2 and before_nums[0] == int(before_nums[0])
                    and before_nums[0] < 1000):
                qty = before_nums[1]

        # Values in between section: [units] [buy_val] [acq_val] (positional order)
        # Layout: units always comes first (it's a decimal like 149.925)
        #         buy_val and acq_val are larger round amounts (like 5000.00)
        # Strategy: 
        #   - Collect all between values > 10 in order
        #   - The FIRST value that is NOT a whole round integer OR is < 1000 = likely units
        #   - Skip it; the remaining 2 values are buy_val and acq_val
        all_between_pos = [n for n in between_nums if n > 0]  # all positional

        # Detect and skip units: units have decimal parts like 149.925
        # Amount values are typically >> units or are whole-number multiples like 5000.00
        amount_between = []
        units_found = False
        for n in all_between_pos:
            if not units_found and n < 10000:
                # Check if this looks like units (decimal precision, not a round rupee amount)
                # Units: 149.925, 68.130 etc. Buy amounts: 5000.00, 4999.99
                # Heuristic: if n has 3 decimal places OR n < 500 with decimal part, it's units
                n_str = str(n)
                decimal_places = len(n_str.split('.')[-1]) if '.' in n_str else 0
                is_units_candidate = (decimal_places >= 1 and n < 5000) or n < 100
                if is_units_candidate and not amount_between:
                    units_found = True
                    # Update qty with this more accurate value
                    if qty is None or qty < 2:  # qty from before was just sno
                        qty = n
                    continue  # skip units
            amount_between.append(n)

        # Large values in after section (positional)
        large_after = [n for n in after_nums if n > 0]
        # Filter out zeros and small gains (< 0 possible, but sell_amt is always positive)
        positive_after = [n for n in large_after if n > 100]

        buy_amt  = amount_between[0] if amount_between else None
        # Skip leading per-unit NAV: if first value is < 10% of second, it's a unit price not amount
        if len(positive_after) >= 2 and positive_after[0] < positive_after[1] * 0.1:
            sell_amt = positive_after[1]
        else:
            sell_amt = positive_after[0] if positive_after else None

        # Fallback: if no sell_amt in after, use last amount from between
        if not sell_amt and len(amount_between) >= 2:
            sell_amt = amount_between[-1]

        if not buy_amt or not sell_amt:
            return None
        if sell_amt <= 0 or buy_amt <= 0:
            return None

        # Grandfathering: Asahoo/Kuvera: [buy_val, acq_val] in positional order
        # Pre-2018: acq_val = grandfathered cost (Jan-2018 FMV), may differ from buy_val
        # Post-2018: acq_val == buy_val (no grandfathering)
        fmv_total = None
        effective_purchase = buy_amt
        if len(amount_between) >= 2:
            acq_val = amount_between[1]  # positionally AFTER buy_val
            if abs(acq_val - buy_amt) > 0.02 and acq_val > 0:
                if buy_date and buy_date <= self._GF_CUTOFF:
                    # Pre-2018: acq_val is grandfathered cost
                    fmv_total = acq_val
                    effective_purchase = acq_val
                # Post-2018: no grandfathering, buy_amt stays

        is_equity = (asset_type or "").lower() not in ("debt",)
        gain_type = determine_gain_type(buy_date, sell_date, is_equity)
        sell_price = (sell_amt / qty) if (qty and qty > 0) else None

        return make_row(
            self.source_name, scheme, isin or '',
            sell_date, qty, sell_price, sell_amt,
            buy_date, effective_purchase, gain_type, self.client_name,
            asset_type=asset_type,
            total_market_value_31jan2018=fmv_total,
        )

class GrowwStocksFlatParser:
    """Parses Groww Stocks Capital Gains Report XLSX with flat section layout.

    Sections: "Intraday trades" / "Short Term trades" / "Long Term trades" / "Buyback trades"
    Header row: Stock name | ISIN | Quantity | Buy date | Buy price | Buy value |
                Sell date | Sell price | Sell value | Realised P&L | Remark

    Rules:
     - Intraday trades → marked Semi-Processed (sale_date == purchase_date, filtered by existing rule)
     - Buyback trades → skipped (not standard CG)
     - Short/Long Term → use declared section for gain_type
    """

    SKIP_SECTIONS = {'intraday trades', 'buyback trades'}
    SECTION_MAP   = {
        'short term trades': 'Short Term',
        'long term trades':  'Long Term',
    }

    def parse(self, file_path, source_name=None):
        self.source_name = source_name or source_name_from_file(file_path)
        self.client_name = None
        rows = []

        try:
            df = pd.read_excel(file_path, header=None)
        except Exception as e:
            print(f"[WARN] GrowwStocksFlatParser: {e}", file=sys.stderr)
            return rows

        # Extract client name
        for i in range(min(5, len(df))):
            for j in range(min(3, df.shape[1])):
                cell = str(df.iloc[i, j]) if pd.notna(df.iloc[i, j]) else ''
                if cell.lower() == 'name':
                    nj = j + 1
                    if nj < df.shape[1] and pd.notna(df.iloc[i, nj]):
                        self.client_name = str(df.iloc[i, nj]).strip().title()
                    break

        current_section = None
        current_gain_type = None
        in_skip_section = False
        col = {}   # column index map from header row

        for pos in range(len(df)):
            row = df.iloc[pos]
            first = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''

            # Detect section headers
            if first in self.SKIP_SECTIONS:
                in_skip_section = True
                current_section = first
                col = {}
                continue
            if first in self.SECTION_MAP:
                in_skip_section = False
                current_section = first
                current_gain_type = self.SECTION_MAP[first]
                col = {}
                continue

            # Detect header row within a section
            vals_lower = [str(v).strip().lower() if pd.notna(v) else '' for v in row]
            if 'stock name' in vals_lower and 'sell date' in vals_lower:
                col = {}
                for i, h in enumerate(vals_lower):
                    if h == 'stock name':       col['name'] = i
                    elif h == 'isin':           col['isin'] = i
                    elif h == 'quantity':       col['qty'] = i
                    elif h == 'buy date':       col['buy_date'] = i
                    elif h == 'buy price':      col['buy_price'] = i
                    elif h == 'buy value':      col['buy_value'] = i
                    elif h == 'sell date':      col['sell_date'] = i
                    elif h == 'sell price':     col['sell_price'] = i
                    elif h == 'sell value':     col['sell_value'] = i
                    elif h == 'realised p&l':   col['pnl'] = i
                    elif h == 'remark':         col['remark'] = i
                continue

            if in_skip_section or not col or not current_gain_type:
                continue

            def g(field):
                idx = col.get(field)
                if idx is None or idx >= len(row): return None
                v = row.iloc[idx]
                return None if pd.isna(v) else v

            name = str(g('name') or '').strip()
            if not name or name.lower() in ('nan', '', 'stock name'):
                continue
            if 'total' in name.lower() or 'disclaimer' in name.lower():
                continue

            # Check remark — skip intraday tagged rows
            remark = str(g('remark') or '').strip().lower()
            if 'intraday' in remark:
                continue

            sell_date = parse_date(str(g('sell_date') or ''))
            buy_date  = parse_date(str(g('buy_date')  or ''))
            if not sell_date:
                continue

            qty        = to_float(g('qty'))
            sell_value = to_float(g('sell_value'))
            buy_value  = to_float(g('buy_value'))
            sell_price = to_float(g('sell_price'))
            buy_price  = to_float(g('buy_price'))

            if sell_value is None and qty and sell_price:
                sell_value = qty * sell_price
            if buy_value is None and qty and buy_price:
                buy_value = qty * buy_price

            if not sell_value or not buy_value:
                continue

            rows.append(make_row(
                self.source_name, name, str(g('isin') or '').strip(),
                sell_date, qty, sell_price, sell_value,
                buy_date, buy_value, current_gain_type, self.client_name
            ))

        return rows


# ─────────────────────────────────────────────
# KUVERA CAPITAL GAINS PDF PARSER
# Format: SNo Units Date Value Acquisition Value * Date Value STCG LTCG
# ─────────────────────────────────────────────

class KuveraParser:
    """Parses Kuvera Capital Gains Statement PDFs.

    Fund block header: "SchemeName [ISIN: INFxxx] (Debt/Equity)"
    Data row: sno units buy_date buy_value acq_value * sell_date sell_value stcg ltcg
    Dates format: "Jan, 22 2025" or "Jan 22 2025"
    """

    _ISIN_RE  = re.compile(r'\[ISIN:\s*(INF[A-Z0-9]+)\]', re.I)
    _GF_CUTOFF = datetime(2018, 1, 31)
    # "Short Term Capital Gains ₹ 2,800.87" or "Total ₹ 2,800.87 ₹ 0"
    _ST_RE = re.compile(r'Short\s+Term\s+Capital\s+Gains\s+[₹Rs.]*\s*([-\d,]+\.?\d*)', re.I)
    _LT_RE = re.compile(r'Long\s+Term\s+Capital\s+Gains\s+[₹Rs.]*\s*([-\d,]+\.?\d*)', re.I)
    _TOT_RE = re.compile(r'^Total\s+[₹Rs.]*\s*([-\d,]+\.?\d*)\s+[₹Rs.]*\s*([-\d,]+\.?\d*)', re.M)

    def _extract_pdf_summary(self, pdf_path):
        """Extract ST/LT totals from Kuvera summary page."""
        try:
            with open_pdf(pdf_path) as pdf:
                text = "\n".join(p.extract_text() or "" for p in pdf.pages)
            ms = self._ST_RE.search(text)
            ml = self._LT_RE.search(text)
            if ms and ml:
                st = float(ms.group(1).replace(',', ''))
                lt = float(ml.group(1).replace(',', ''))
                return round(st, 2), round(lt, 2), round(st + lt, 2)
            mt = self._TOT_RE.search(text)
            if mt:
                st = float(mt.group(1).replace(',', ''))
                lt = float(mt.group(2).replace(',', ''))
                return round(st, 2), round(lt, 2), round(st + lt, 2)
        except Exception:
            pass
        return None, None, None

    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        self.client_name = None
        with open_pdf(pdf_path) as pdf:
            pages_text = [p.extract_text() or "" for p in pdf.pages]
        full_text = "\n".join(pages_text)
        m = re.search(r'Name\s+([A-Za-z][A-Za-z\s\.]+?)\n', full_text)
        if m: self.client_name = m.group(1).strip().title()
        rows = self._parse_text(full_text)
        pdf_st, pdf_lt, pdf_total = self._extract_pdf_summary(pdf_path)
        self.last_validation = build_validation_report(
            rows, pdf_st=pdf_st, pdf_lt=pdf_lt, pdf_total=pdf_total,
            parse_failures=0, parser_name='KuveraParser',
        )
        return rows

    def _parse_text(self, text):
        rows = []
        current_scheme = None
        current_isin   = None
        current_asset  = "Equity"

        for line in text.split("\n"):
            stripped = line.strip()
            if not stripped: continue
            su = stripped.upper()

            # Fund header: SchemeName [ISIN: INFxxx] (Debt)
            isin_m = self._ISIN_RE.search(stripped)
            if isin_m:
                current_isin = isin_m.group(1)
                name_part = stripped[:isin_m.start()].strip().rstrip('[-')
                current_scheme = name_part if len(name_part) > 3 else current_scheme
                current_asset  = "Debt" if "DEBT" in su else "Equity"
                continue

            # Skip totals/headers
            if any(k in su for k in ['TOTAL','NOTE','DISCLAIMER','ACQUISITION',
                                      'SDATE','STATEMENT PERIOD','SUB TOTAL',
                                      'TRANSACTIONS','PURCHASE','REDEMPTION',
                                      'EQUITY SUB','DEBT SUB','STCG','LTCG',
                                      'FOLIO NO','FUND TOTAL','₹']):
                continue

            if not current_scheme: continue

            lot = self._parse_lot_line(stripped, current_scheme, current_isin, current_asset)
            if lot: rows.append(lot)
        return rows

    def _parse_lot_line(self, line, scheme, isin, asset_type):
        """sno units buy_date buy_value acq_value sell_date sell_value stcg ltcg
        Kuvera date format: "Jan, 22 2025" or "Jan 22 2025"
        """
        date_pat = r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[,\s]+\d{1,2}[,\s]+\d{4}'
        raw_dates = re.findall(date_pat, line, re.I)
        dates = []
        for d in raw_dates:
            m = re.match(r'([A-Za-z]+)[,\s]+(\d{1,2})[,\s]+(\d{4})', d.strip())
            if m:
                dates.append(f"{m.group(2)} {m.group(1)} {m.group(3)}")
            else:
                dates.append(d)
        if len(dates) < 2:
            dates = re.findall(r'\d{2}/\d{2}/\d{4}', line)
            if len(dates) < 2:
                return None

        buy_date  = parse_date(dates[0])
        sell_date = parse_date(dates[1])
        if not buy_date or not sell_date or sell_date <= buy_date:
            return None

        # Strip date strings from line BEFORE extracting numbers
        # This prevents year digits (e.g. 2025) being parsed as amounts
        line_no_dates = re.sub(date_pat, ' ', line, flags=re.I)
        nums = [to_float(n) for n in re.findall(r'[\d,]+\.?\d*', line_no_dates)
                if to_float(n) is not None and to_float(n) > 0]
        if len(nums) < 3: return None

        # Layout: [sno] units buy_value acq_value sell_value [stcg] [ltcg]
        idx = 0
        if nums[0] == int(nums[0]) and nums[0] < 500: idx = 1  # skip serial
        if idx >= len(nums): return None

        units    = nums[idx]
        buy_val  = nums[idx+1] if len(nums) > idx+1 else None
        acq_val  = nums[idx+2] if len(nums) > idx+2 else None
        sell_val = nums[idx+3] if len(nums) > idx+3 else None

        if not sell_val or not buy_val: return None
        if sell_val <= 0: return None

        purchase_amt = acq_val if acq_val else buy_val
        fmv_total    = None
        if (asset_type == "Equity" and acq_val and buy_val
                and abs(acq_val - buy_val) > 0.01
                and buy_date and buy_date <= self._GF_CUTOFF):
            fmv_total = acq_val

        gain_type  = determine_gain_type(buy_date, sell_date, asset_type != "Debt")
        sell_price = (sell_val / units) if (units and units > 0) else None

        return make_row(
            self.source_name, scheme, isin or '',
            sell_date, units, sell_price, sell_val,
            buy_date, purchase_amt, gain_type, self.client_name,
            asset_type=asset_type,
            total_market_value_31jan2018=fmv_total,
        )

# ─────────────────────────────────────────────
# CAMS SECTION A/B/C PDF PARSER
# ─────────────────────────────────────────────
class CAMSSectionABCParser:
    """Parses CAMS-generated capital gain PDFs where each row contains
    both Section A (purchase) and Section B (redemption) inline.

    Used by: Motilal Oswal MF, UTI MF, and other CAMS-platform AMCs
    that generate 'Section A : Subscriptions | Section B : Redemptions | Section C : Gains' layout.

    Row format (buy keyword + date + nums + Redemption + date + nums):
      [Buy type] [DD-MM-YYYY] units units nav buy_amt [nav31jan gf_cost]×2
      Redemption [DD-MM-YYYY] units sell_amt sell_nav [...gains]
    """

    _BUY_KW  = re.compile(
        r'\b(Purchase|Systematic|Sys\.\s*Investment|Switch\s*In|Div\.?\s*Reinv)\b', re.I)
    _DATE_RE = re.compile(r'\d{2}-\d{2}-\d{4}')
    _NUM_RE  = re.compile(r'[\d,]+\.?\d*')
    _GF_DATE = datetime(2018, 1, 31)

    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        self.client_name = None
        rows = []

        with open_pdf(pdf_path) as pdf:
            pages_text = [p.extract_text() or '' for p in pdf.pages]
        full_text = '\n'.join(pages_text)

        # Client name
        m = re.search(
            r'Name\s*:\s*([A-Z][A-Z\s]+?)(?:\s+Status|\s+S/O|\s+Folio|$)',
            full_text, re.I | re.M)
        if m:
            self.client_name = ' '.join(m.group(1).split()).title()

        # PAN
        pan = None
        m = re.search(r'PAN\s*(?:No\s*)?:?\s*([A-Z]{5}\d{4}[A-Z])', full_text, re.I)
        if m:
            pan = m.group(1).upper()

        # Split into scheme blocks on "Section A : Subscriptions"
        parts = re.split(r'Section A\s*:\s*Subscriptions', full_text, flags=re.I)

        for i in range(1, len(parts)):
            scheme, isin = self._extract_scheme_isin(parts[i - 1])
            if not scheme:
                continue

            is_equity = not any(k in scheme.lower() for k in
                                ['debt', 'liquid', 'bond fund', 'money market',
                                 'overnight', 'gilt', 'arbitrage'])

            for line in parts[i].split('\n'):
                r = self._parse_data_line(
                    line.strip(), scheme, isin, is_equity, pan)
                if r:
                    rows.append(r)

        return rows

    def _extract_scheme_isin(self, preceding_text):
        """Get scheme name and ISIN from text before 'Section A'."""
        lines = [l.strip() for l in preceding_text.split('\n') if l.strip()]
        scheme = ''
        isin   = ''
        for line in reversed(lines[-8:]):
            m = re.search(r'(INF[A-Z0-9]{6,12})', line)
            if m:
                isin = m.group(1)
            # Remove ISIN, trailing stats, leading EQUITY/DEBT label
            cand = re.sub(r'\(?\s*INF[A-Z0-9]+\s*\)?|ISIN\s*:\s*INF[A-Z0-9]+', '', line).strip()
            cand = re.sub(r'\s+\d[\d,.]*(?:\s+\d[\d,.]*)+\s*$', '', cand).strip()
            cand = re.sub(r'^(EQUITY|DEBT)\s+', '', cand, flags=re.I).strip()
            cand = cand.strip(' -–')
            if 10 < len(cand) < 120 and not cand[0].isdigit():
                scheme = cand
                break
        return scheme, isin

    def _parse_data_line(self, line, scheme, isin, is_equity, pan):
        """Parse one merged Section A+B row."""
        red_m = re.search(r'\bRedemption\b', line, re.I)
        if not red_m:
            return None
        if not self._BUY_KW.search(line[:red_m.start()]):
            return None

        buy_part  = line[:red_m.start()]
        sell_part = line[red_m.start():]

        # Buy date
        buy_dates = self._DATE_RE.findall(buy_part)
        if not buy_dates:
            return None
        buy_date = parse_date(buy_dates[0])
        if not buy_date:
            return None

        # Sell date
        sell_dates = self._DATE_RE.findall(sell_part)
        if not sell_dates:
            return None
        sell_date = parse_date(sell_dates[0])
        if not sell_date:
            return None

        # Buy numbers: units units nav buy_amount [nav31jan gf_cost nav31jan gf_cost]
        clean_buy = re.sub(r'\d{2}-\d{2}-\d{4}', '', buy_part)
        buy_nums = [to_float(n) for n in self._NUM_RE.findall(clean_buy)
                    if to_float(n) is not None and to_float(n) > 0]
        if len(buy_nums) < 4:
            return None
        buy_amount = buy_nums[3]  # 4th number = original purchase amount

        # Grandfathering: pre-2018 purchases have NAV31Jan + GF cost as columns 5,6
        fmv_total = None
        if buy_date <= self._GF_DATE and len(buy_nums) >= 6:
            gf_cost = buy_nums[5]
            nav_31jan = buy_nums[4]
            # GF cost is meaningful when NAV31Jan ≠ original NAV (appreciation)
            if nav_31jan > 0 and abs(nav_31jan - buy_nums[2]) > 0.001:
                fmv_total = gf_cost

        # Sell numbers: units sell_amount sell_nav [...gains]
        clean_sell = re.sub(r'Redemption\s*', '', sell_part, flags=re.I)
        clean_sell = re.sub(r'\d{2}-\d{2}-\d{4}', '', clean_sell)
        sell_nums = [to_float(n) for n in self._NUM_RE.findall(clean_sell)
                     if to_float(n) is not None and to_float(n) > 0]
        if len(sell_nums) < 2:
            return None
        sell_amount = sell_nums[1]   # 2nd number = sell amount (1st = units)

        gain_type = determine_gain_type(buy_date, sell_date, is_equity=is_equity)

        r = make_row(
            self.source_name, scheme, isin,
            sell_date, None, None, sell_amount,
            buy_date, buy_amount, gain_type, self.client_name,
            total_market_value_31jan2018=fmv_total,
            asset_type='Equity' if is_equity else 'Debt',
        )
        if pan:
            r['pan'] = pan
        return r


class SBIMFParser:
    """Parses SBI Mutual Fund Investment Gain/Loss Statement PDFs.

    Format per scheme block:
      INF200K01T51 / SBI Small Cap Fund - Direct Plan - Growth
      Purchase - DD-Mon-YYYY  units  nav  indexed_nav
      Redemption DD-Mon-YYYY  total_units_redeemed  sell_price  [sell_amt]  [st_gain lt_gain]

    The PDF uses a complex two-column layout. We parse lot-level:
    each Purchase line feeds a lot; each Redemption line creates a row
    using the most recent purchase lot as the cost basis.
    """

    _ISIN_SCHEME_RE = re.compile(
        r'(INF[A-Z0-9]{6,12})\s*/\s*(.+)', re.I)
    _GF_CUTOFF = datetime(2018, 1, 31)

    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        self.client_name = None
        self.pan = None
        with open_pdf(pdf_path) as pdf:
            pages_text = [p.extract_text() or "" for p in pdf.pages]
        full_text = "\n".join(pages_text)
        m = re.search(r'Name:\s*([A-Z][A-Z\s]+?)(?:\s+Status|\s+PAN|$)', full_text, re.I)
        if m: self.client_name = m.group(1).strip().title()
        mp = re.search(r'\bPAN\s*:\s*([A-Z]{5}\d{4}[A-Z])', full_text, re.I)
        if mp: self.pan = mp.group(1).upper()
        return self._parse_text(full_text)

    def _parse_text(self, text):
        """SBI MF statement: reads totals directly from the PDF summary line.
        
        Strategy: collect all Redemption lines (sell amounts + sell dates),
        then use the PDF-provided Total line to get:
          - total_purchase (printed directly)
          - LT_without_index gain
        Then distribute across redemptions proportionally.
        """
        rows = []
        current_isin   = None
        current_scheme = None
        lots = []          # (buy_date, units, nav)
        redemptions = []   # (sell_date, sell_amt)
        total_purchase = None
        lt_gain_no_idx = None

        lines = text.split("\n")
        # Two-pass: first collect all data, then build rows
        i = 0
        while i < len(lines):
            stripped = lines[i].strip()
            i += 1
            if not stripped: continue

            # ── ISIN / Scheme header ──────────────────────────────────
            m = self._ISIN_SCHEME_RE.match(stripped)
            if m:
                # Flush previous scheme
                if current_scheme and redemptions:
                    rows.extend(self._build_rows(
                        current_scheme, current_isin,
                        redemptions, lots, total_purchase, lt_gain_no_idx))
                current_isin   = m.group(1)
                current_scheme = m.group(2).strip()
                # Fund name may wrap to next line (e.g. "SBI Energy Fund -\nDirect Growth")
                if current_scheme.endswith('-') and i < len(lines):
                    nxt = lines[i].strip()
                    if nxt and not self._ISIN_SCHEME_RE.match(nxt) and not re.search(r'\d{2}-[A-Za-z]', nxt):
                        current_scheme = (current_scheme + ' ' + nxt).strip()
                        i += 1
                lots = []; redemptions = []
                total_purchase = None; lt_gain_no_idx = None
                continue

            su = stripped.upper()
            if any(k in su for k in ['DESCRIPTION','AMOUNT(RS','LONG TERM',
                                      'SHORT TERM','INDEXED','WITHOUT INDEX',
                                      'COST *','TDS(IF','DEDUCTED','DISCLAIMER',
                                      'NOTE :','UNIT COST','INDEXED COST',
                                      'SUMMARY OF DIVIDEND','B: CORRESPONDING',
                                      'A: REDEMPTION','FOR THE PERIOD']):
                continue
            if not current_scheme: continue

            # ── Total summary line → derive total_purchase ──────────
            # Format: "Total [units] [total_sell] [other] [0] [total_gain] ..."
            if stripped.upper().startswith('TOTAL'):
                tnums = [to_float(n) for n in re.findall(r'[\d,]+\.?\d*', stripped)
                         if to_float(n) is not None]  # keep zeros
                if len(tnums) >= 5 and tnums[1] > 1000:
                    total_purchase = tnums[1] - tnums[4]   # total_sell − total_gain
                elif len(tnums) >= 2 and tnums[1] > 1000:
                    total_purchase = tnums[1]
                continue

            # ── Skip standalone numbers (folio IDs, page numbers) ───
            nums_raw = re.findall(r'[\d,]+\.?\d*', stripped)
            if (len(nums_raw) <= 2 and not re.search(r'[A-Za-z]', stripped)):
                continue

            # ── Redemption line ───────────────────────────────────────
            if re.search(r'Redemption', stripped, re.I):
                dates = re.findall(r'\d{2}-[A-Za-z]{3}-\d{4}|\d{2}-\d{2}-\d{4}', stripped)
                if not dates: continue
                sell_date = parse_date(dates[0])
                if not sell_date: continue
                # Numbers AFTER the date: [units, sell_amt_or_nav, nav, ...]
                after_date = re.sub(r'^.*?' + re.escape(dates[0]), '', stripped)
                after_nums = [to_float(n) for n in re.findall(r'[\d,]+\.?\d*', after_date)
                              if to_float(n) is not None and to_float(n) > 0]
                if len(after_nums) < 2: continue
                units_r, second = after_nums[0], after_nums[1]
                # If second >> units, it's the explicit sell amount; else compute units × nav
                if second > units_r * 1.5:
                    sell_amt = second
                else:
                    sell_amt = round(units_r * second, 2)
                if sell_amt > 100:
                    redemptions.append((sell_date, sell_amt))
                continue

            # ── Purchase line ─────────────────────────────────────────
            if re.search(r'Purchase|Systematic|Switch\s*In', stripped, re.I):
                dates = re.findall(r'\d{2}-[A-Za-z]{3}-\d{4}|\d{2}-\d{2}-\d{4}', stripped)
                if not dates: continue
                buy_date = parse_date(dates[0])
                if not buy_date: continue
                nums_all = [to_float(n) for n in re.findall(r'[\d,]+\.?\d*', stripped)
                            if to_float(n) is not None and to_float(n) > 0]
                if len(nums_all) < 2: continue
                lots.append((buy_date, nums_all[0], nums_all[1]))

        # Flush last scheme
        if current_scheme and redemptions:
            rows.extend(self._build_rows(
                current_scheme, current_isin,
                redemptions, lots, total_purchase, lt_gain_no_idx))
        return rows

    def _build_rows(self, scheme, isin, redemptions, lots, total_purchase, lt_gain_no_idx):
        """Build rows for one scheme using proportional purchase allocation with FIFO lot dating."""
        rows = []
        if not redemptions: return rows

        total_sell = sum(amt for _, amt in redemptions)
        is_eq = not any(k in scheme.lower() for k in
                        ['debt','liquid','bond fund','money market',
                         'overnight','gilt','arbitrage'])

        # Sort lots FIFO (earliest first); estimate total units for FIFO positioning
        fifo_lots = sorted(lots, key=lambda x: x[0]) if lots else []
        total_lot_units = sum(u for _, u, _ in fifo_lots) if fifo_lots else 1.0
        units_consumed = 0.0

        for sell_date, sell_amt in sorted(redemptions, key=lambda x: x[0]):
            proportion = sell_amt / total_sell if total_sell > 0 else 1.0
            if total_purchase and total_purchase > 0:
                buy_amt = round(total_purchase * proportion, 2)
            elif lt_gain_no_idx and lt_gain_no_idx > 0:
                buy_amt = round((total_sell - lt_gain_no_idx) * proportion, 2)
            else:
                buy_amt = round(sell_amt * 0.7, 2)

            # FIFO lot dating: find which lot covers the current redemption position
            redemp_units = total_lot_units * proportion
            buy_date = fifo_lots[0][0] if fifo_lots else sell_date
            if fifo_lots:
                cursor = 0.0
                for lot_date, lot_units, _ in fifo_lots:
                    cursor += lot_units
                    if cursor > units_consumed:
                        buy_date = lot_date
                        break
            units_consumed += redemp_units

            gain_type = determine_gain_type(buy_date, sell_date, is_equity=is_eq)

            r = make_row(
                self.source_name, scheme, isin or '',
                sell_date, None, None, sell_amt,
                buy_date, buy_amt, gain_type, self.client_name,
            )
            if getattr(self, 'pan', None):
                r['pan'] = self.pan
            rows.append(r)
        return rows

def validate_output_rows(rows):
    """
    Run post-extraction validation checks on all rows.
    Returns list of warning dicts: {row_idx, share_name, issue}

    Checks (matching client-reported issues):
    1. Sale value taken entirely as gain (purchase_amount is 0 or near-zero)
    2. Share name is blank or looks like a date/number
    3. Sale or purchase amount is zero/None
    4. Grandfathering: LT equity pre-2018 purchase but fmv_total is None
    5. Intraday: sale_date == purchase_date (should have been filtered)
    6. Gain type mismatch vs holding period
    7. Unlisted classification: ISIN starts with INF (MF) but share_type is Unlisted Security
    """
    warnings = []
    for i, r in enumerate(rows):
        name      = r.get("share_name","") or ""
        sale_amt  = r.get("sale_amount") or 0
        buy_amt   = r.get("purchase_amount") or 0
        sale_date = r.get("sale_date")
        buy_date  = r.get("purchase_date")
        gain_type = r.get("gain_type","")
        isin      = r.get("isin","") or ""
        share_type= r.get("share_type","")

        # Check 1: entire sale taken as gain
        if sale_amt > 0 and (buy_amt == 0 or buy_amt is None):
            warnings.append({
                "row": i+1, "name": name,
                "issue": "WARN-01: Purchase amount is zero — entire sale may be taken as gain",
                "severity": "HIGH"
            })

        # Check 2: share name looks wrong (date/number/too short)
        if not name or len(name) < 3:
            warnings.append({"row": i+1, "name": name,
                "issue": "WARN-02: Share name missing or too short", "severity": "HIGH"})
        elif re.match(r'^\d{2}[\s/-]\w+[\s/-]\d{4}$', name.strip()):
            warnings.append({"row": i+1, "name": name,
                "issue": "WARN-02: Share name appears to be a date (parser error)", "severity": "HIGH"})

        # Check 3: zero/missing amounts
        if not sale_amt or sale_amt == 0:
            warnings.append({"row": i+1, "name": name,
                "issue": "WARN-03: Sale amount is zero", "severity": "MEDIUM"})

        # Check 4: Missing grandfathering for pre-2018 LT equity
        GF_CUTOFF = datetime(2018, 1, 31)
        if (gain_type == "Long Term"
                and buy_date and buy_date <= GF_CUTOFF
                and isin and not isin.startswith("INF")   # stocks not MF
                and not r.get("total_market_value_31jan2018")):
            warnings.append({"row": i+1, "name": name,
                "issue": "WARN-04: Pre-Jan-2018 LT equity purchase — grandfathering (31-Jan-2018 FMV) not set",
                "severity": "MEDIUM"})

        # Check 5: Intraday not filtered
        if sale_date and buy_date and sale_date == buy_date:
            warnings.append({"row": i+1, "name": name,
                "issue": "WARN-05: Intraday transaction (sale_date == purchase_date) not filtered",
                "severity": "HIGH"})

        # Check 6: Gain type mismatch — threshold differs for equity vs debt
        if sale_date and buy_date:
            days = (sale_date - buy_date).days
            row_atype = str(r.get("asset_type","")).lower()
            lt_threshold = 730 if "debt" in row_atype else 365
            expected_lt = days > lt_threshold
            if gain_type == "Long Term" and not expected_lt:
                warnings.append({"row": i+1, "name": name,
                    "issue": f"WARN-06: Marked Long Term but held only {days} days",
                    "severity": "MEDIUM"})
            elif gain_type == "Short Term" and expected_lt:
                warnings.append({"row": i+1, "name": name,
                    "issue": f"WARN-06: Marked Short Term but held {days} days (>{lt_threshold} day threshold)",
                    "severity": "LOW"})

        # Check 7: MF ISIN but classified as Unlisted Security (only warn for obvious equity funds)
        if isin.startswith("INF") and share_type == "Unlisted Security":
            _debt_kw = ["debt", "liquid", "overnight", "duration", "money market",
                        "credit risk", "floater", "gilt", "banking and psu",
                        "ultra short", "short term", "ultra-short", "low duration",
                        "fixed maturity", "corporate bond", "income", "arbitrage",
                        "bond fund", "bond ", " bond", "all seasons"]
            if not any(kw in name.lower() for kw in _debt_kw):
                warnings.append({"row": i+1, "name": name,
                    "issue": "WARN-07: MF ISIN (INF...) classified as Unlisted Security — check if equity MF",
                    "severity": "LOW"})

    return warnings



# ─────────────────────────────────────────────
# ANGEL ONE EXCEL PARSER
# ─────────────────────────────────────────────
class AngelOneExcelParser:
    """Parses Angel One capital gains XLSX files.
    Sheet: Equity+Bonds+SGB Trade Details
    Delivery P&L section columns: ISIN, Scrip Name, Qty, Buy Date, Sell Date,
      Avg Buy Price, Buy Value, Avg Sell Price, Sell Value, Cost Of Acquisition,
      Charges and Statutory Levies, STT, Net Profit/Loss,
      Long term taxable income, Short term taxable income, Purchase Type, Type of instrument
    """
    def parse(self, file_path, source_name=None):
        self.source_name = source_name or source_name_from_file(file_path)
        self.client_name = None
        rows = []
        try:
            import warnings as _w
            with _w.catch_warnings():
                _w.simplefilter("ignore")
                wb = openpyxl.load_workbook(file_path, data_only=True)

            # Try to extract client name from Summary sheet
            if 'Summary' in wb.sheetnames:
                ws_sum = wb['Summary']
                for row in ws_sum.iter_rows(values_only=True):
                    for cell in row:
                        if cell and 'client name' in str(cell).lower():
                            idx = list(row).index(cell)
                            if idx + 1 < len(row) and row[idx + 1]:
                                self.client_name = str(row[idx + 1]).strip().title()
                                break
                    if self.client_name:
                        break

            # Also try "Client Name" row in the equity sheet header
            target_sheet = None
            for sname in wb.sheetnames:
                if 'equity' in sname.lower() or 'trade' in sname.lower():
                    target_sheet = sname
                    break
            if not target_sheet:
                return rows

            ws = wb[target_sheet]
            all_rows = [list(r) for r in ws.iter_rows(values_only=True)]

            # Find client name from early rows
            for r in all_rows[:15]:
                cells = [str(c).strip() if c is not None else '' for c in r]
                for i, c in enumerate(cells):
                    if 'client name' in c.lower() and i + 1 < len(cells) and cells[i+1]:
                        self.client_name = cells[i+1].title()

            # Find the Delivery P&L header row
            delivery_hdr_idx = None
            col = {}
            for i, r in enumerate(all_rows):
                cells = [str(c).strip().lower() if c is not None else '' for c in r]
                if 'buy date' in cells and 'sell date' in cells and 'buy value' in cells:
                    delivery_hdr_idx = i
                    for j, h in enumerate(cells):
                        if h == 'isin':                             col.setdefault('isin', j)
                        elif 'scrip name' in h:                     col.setdefault('name', j)
                        elif h == 'qty' or h == 'quantity':         col.setdefault('qty', j)
                        elif h == 'buy date':                       col.setdefault('buy_date', j)
                        elif h == 'sell date':                      col.setdefault('sell_date', j)
                        elif h == 'avg buy price':                  col.setdefault('buy_rate', j)
                        elif h == 'buy value':                      col.setdefault('buy_value', j)
                        elif h == 'avg sell price':                 col.setdefault('sell_rate', j)
                        elif h == 'sell value':                     col.setdefault('sell_value', j)
                        elif 'charges and statutory' in h:         col.setdefault('charges', j)
                        elif h == 'stt':                            col.setdefault('stt', j)
                        elif 'long term taxable' in h:              col.setdefault('lt_income', j)
                        elif 'short term taxable' in h:             col.setdefault('st_income', j)
                        elif 'type of instrument' in h:             col.setdefault('instrument', j)
                    break

            if delivery_hdr_idx is None:
                return rows

            def gv(r, field):
                idx = col.get(field)
                if idx is None or idx >= len(r): return None
                v = r[idx]
                return None if (v is None or str(v).strip() in ('', 'None')) else v

            for r in all_rows[delivery_hdr_idx + 1:]:
                cells_str = [str(c).strip() if c is not None else '' for c in r]
                # Stop when we reach an empty section or "Sub total" / next header
                first = cells_str[0] if cells_str else ''
                if any(k in first.lower() for k in ('sub total', 'grand total', 'intraday', 'unrealised')):
                    break
                # Check if this looks like a header row (no ISIN)
                isin_val = str(gv(r, 'isin') or '').strip()
                if not isin_val or not isin_val.startswith('IN'):
                    continue

                name_val     = str(gv(r, 'name') or '').strip()
                qty          = to_float(gv(r, 'qty'))
                buy_date     = parse_date(str(gv(r, 'buy_date') or ''))
                sell_date    = parse_date(str(gv(r, 'sell_date') or ''))
                buy_rate     = to_float(gv(r, 'buy_rate'))
                buy_value    = to_float(gv(r, 'buy_value'))
                sell_rate    = to_float(gv(r, 'sell_rate'))
                sell_value   = to_float(gv(r, 'sell_value'))
                charges      = to_float(gv(r, 'charges')) or 0
                lt_income    = to_float(gv(r, 'lt_income'))
                st_income    = to_float(gv(r, 'st_income'))

                if not sell_date or not name_val:
                    continue
                if not sell_value:
                    if qty and sell_rate: sell_value = qty * sell_rate
                    else: continue
                if not buy_value or buy_value == 0:
                    if qty and buy_rate: buy_value = qty * buy_rate
                    else: buy_value = 0.0

                # Determine gain type from the "Long/Short term taxable income" columns
                if lt_income is not None and lt_income != 0:
                    gain_type = 'Long Term'
                elif st_income is not None and st_income != 0:
                    gain_type = 'Short Term'
                else:
                    gain_type = determine_gain_type(buy_date, sell_date, True)

                # Charges and Statutory Levies = transfer expenses (brokerage + GST + SEBI charges)
                transfer_exp = charges if charges else None

                rows.append(make_row(
                    self.source_name, name_val, isin_val,
                    sell_date, qty, sell_rate, sell_value,
                    buy_date, buy_value, gain_type, self.client_name,
                    transfer_expenses=transfer_exp,
                ))
        except Exception as e:
            print(f"[WARN] AngelOneExcelParser: {e}", file=sys.stderr)
        return rows


# ─────────────────────────────────────────────
# ICICI SECURITIES EXCEL PARSER
# ─────────────────────────────────────────────
class ICICISecuritiesExcelParser:
    """Parses ICICI Securities capital gain Excel files (.xlsx).

    Sheet: contains 'CapitalGains' or truncated variant.
    Columns (row 4, 0-indexed):
      Stock Symbol | ISIN | Qty | Sale Date | Sale Rate | Sale Value |
      Sale Expenses | Purchase Date | Purchase Rate |
      Price as on 31st Jan 2018 | Purchase Price Considered |
      Purchase Value | Purchase Expenses | Profit/Loss(-)

    Sections: 'Short Term Capital Gain (STT paid)' / 'Long Term Capital Gain (STT paid)'
    """

    _GF_DATE = datetime(2018, 1, 31)

    def parse(self, file_path, source_name=None):
        self.source_name = source_name or source_name_from_file(file_path)
        self.client_name = None
        rows = []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            print(f"[ERROR] ICICISecuritiesExcelParser: {e}", file=sys.stderr)
            return rows

        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))

        current_gain_type = None   # 'Short Term' or 'Long Term'

        for row in all_rows:
            # Client name row
            if row[0] == 'Name' and row[1]:
                self.client_name = str(row[1]).strip().title()
                continue

            # Section headers
            if isinstance(row[0], str):
                low = row[0].lower()
                if 'short term capital gain' in low:
                    current_gain_type = 'Short Term'
                    continue
                elif 'long term capital gain' in low:
                    current_gain_type = 'Long Term'
                    continue
                # Skip header and note rows
                if row[0] in ('Stock Symbol', 'Account', 'Capital Gain') or 'note:' in low:
                    continue
                # Disclaimer / notes block → stop parsing
                if 'icici securities' in low and 'tool' in low:
                    break

            # Data row: must have a valid sale_date (datetime) in col 3
            if not isinstance(row[3], datetime):
                continue

            stock_sym  = str(row[0]).strip() if row[0] else ''
            isin       = str(row[1]).strip() if row[1] else ''
            qty        = to_float(row[2])
            sale_date  = row[3]
            sale_amt   = to_float(row[5]) or 0
            sale_exp   = to_float(row[6]) or 0
            purch_date = row[7] if isinstance(row[7], datetime) else None
            price_31jan = row[9]   # 'NA' or float
            purch_amt  = to_float(row[11]) or 0
            purch_exp  = to_float(row[12]) or 0

            if not stock_sym or sale_amt == 0:
                continue

            # Grandfathering
            fmv_total = None
            if (purch_date and purch_date <= self._GF_DATE
                    and price_31jan not in (None, 'NA', 'na', '')):
                fmv_per_unit = to_float(price_31jan)
                if fmv_per_unit and qty:
                    fmv_total = round(qty * fmv_per_unit, 2)

            transfer_exp = round(sale_exp + purch_exp, 2) if (sale_exp or purch_exp) else None

            # Gain type from section header; fallback to holding period
            if current_gain_type:
                gain_type = current_gain_type
            else:
                gain_type = determine_gain_type(purch_date, sale_date, is_equity=True)

            rows.append(make_row(
                self.source_name, stock_sym, isin,
                sale_date, qty, None, sale_amt,
                purch_date, purch_amt, gain_type, self.client_name,
                total_market_value_31jan2018=fmv_total,
                transfer_expenses=transfer_exp,
            ))

        return rows


# ─────────────────────────────────────────────
# ICICI DIRECT TAB-SEPARATED (.xls) PARSER
# ─────────────────────────────────────────────
class ICICIDirectTSVParser:
    """Parses ICICIDirect EQ Capital Gains files saved as tab-separated .xls.
    Columns: Stock Symbol | ISIN | Qty | Sale Date | Sale Rate | Sale Value |
             Sale Expenses | Purchase Date | Purchase Rate |
             Price as on 31st Jan 2018 | Purchase Price Considered |
             Purchase Value | Purchase Expenses | Profit/Loss(-)
    """
    def parse(self, file_path, source_name=None):
        self.source_name = source_name or source_name_from_file(file_path)
        self.client_name = None
        rows = []
        GF_CUTOFF = datetime(2018, 1, 31)

        try:
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
        except Exception:
            return rows

        lines = [l.rstrip('\r') for l in content.split('\n')]

        # Extract client name
        for line in lines[:5]:
            parts = line.split('\t')
            if len(parts) >= 2 and parts[0].strip().lower() == 'name':
                self.client_name = parts[1].strip().title()
                break

        # Find header row
        col = {}
        hdr_idx = None
        for i, line in enumerate(lines):
            parts = [p.strip().lower() for p in line.split('\t')]
            if 'stock symbol' in parts and 'isin' in parts:
                hdr_idx = i
                for j, h in enumerate(parts):
                    if h == 'stock symbol':                     col.setdefault('name', j)
                    elif h == 'isin':                           col['isin'] = j
                    elif h == 'qty':                            col['qty'] = j
                    elif h == 'sale date':                      col['sale_date'] = j
                    elif h == 'sale rate':                      col['sale_rate'] = j
                    elif h == 'sale value':                     col['sale_value'] = j
                    elif h == 'sale expenses':                  col['sale_exp'] = j
                    elif h == 'purchase date':                  col['buy_date'] = j
                    elif h == 'purchase rate':                  col['buy_rate'] = j
                    elif '31st jan' in h or '31/01/2018' in h: col['fmv_price'] = j
                    elif h == 'purchase value':                 col['buy_value'] = j
                    elif h == 'purchase expenses':              col['buy_exp'] = j
                break
        if hdr_idx is None:
            return rows

        def g(parts, field):
            idx = col.get(field)
            if idx is None or idx >= len(parts): return None
            v = parts[idx].strip()
            return v if v and v.upper() != 'NA' else None

        current_gain_type = 'Short Term'
        for line in lines[hdr_idx + 1:]:
            parts = line.split('\t')
            first = parts[0].strip() if parts else ''

            if 'short term' in first.lower():   current_gain_type = 'Short Term';  continue
            if 'long term'  in first.lower():   current_gain_type = 'Long Term';   continue
            if not first or 'sub total' in first.lower() or 'total' in first.lower() or 'note' in first.lower():
                continue

            name = first
            if not name or name.lower() in ('nan', ''): continue

            sale_date  = parse_date(g(parts, 'sale_date') or '')
            buy_date   = parse_date(g(parts, 'buy_date')  or '')
            if not sale_date: continue

            qty        = to_float(g(parts, 'qty'))
            sale_rate  = to_float(g(parts, 'sale_rate'))
            sale_value = to_float(g(parts, 'sale_value'))
            buy_rate   = to_float(g(parts, 'buy_rate'))
            buy_value  = to_float(g(parts, 'buy_value'))
            sale_exp   = to_float(g(parts, 'sale_exp'))  or 0
            buy_exp    = to_float(g(parts, 'buy_exp'))   or 0
            fmv_price  = to_float(g(parts, 'fmv_price'))
            isin_val   = g(parts, 'isin') or ''

            if not sale_value:
                if qty and sale_rate: sale_value = qty * sale_rate
                else: continue
            if not buy_value or buy_value == 0:
                if qty and buy_rate and buy_rate != 0: buy_value = qty * buy_rate
                else: buy_value = 0.0

            fmv_total = None
            if fmv_price and fmv_price > 0 and qty and buy_date and buy_date <= GF_CUTOFF:
                fmv_total = round(fmv_price * qty, 4)

            transfer_exp = (sale_exp + buy_exp) if (sale_exp or buy_exp) else None

            rows.append(make_row(
                self.source_name, name, isin_val,
                sale_date, qty, sale_rate, sale_value,
                buy_date, buy_value, current_gain_type, self.client_name,
                market_price_31jan2018=fmv_price if fmv_total else None,
                total_market_value_31jan2018=fmv_total,
                transfer_expenses=transfer_exp,
            ))
        return rows


# ─────────────────────────────────────────────
# ICICI DIRECT EQUITY PDF PARSER
# ─────────────────────────────────────────────
class ICICIDirectEquityPDFParser:
    """Parses ICICIDirect Equity Capital Gain PDF statements.
    Format: table with columns Description, Stocks, ISIN, Qty,
            Sale Date, Sale Rate, Sale Value, Sale Expenses,
            Purchase Date, Purchase Rate, Purchase Value, Purchase Expenses, Profit/Loss
    Sections: Short Term Capital Gain, Long Term Capital Gain, Speculation Income
    """
    _ISIN_RE = re.compile(r'\b(IN[EF][A-Z0-9]{9,12})\b')
    _DATE_RE = re.compile(r'\d{2}-(?:\d{2}|\w{3})-\d{2,4}', re.I)

    def parse(self, pdf_path, source_name=None):
        self.source_name = source_name or source_name_from_file(pdf_path)
        self.client_name = None
        rows = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                pages_text = [p.extract_text() or "" for p in pdf.pages]
            full_text = "\n".join(pages_text)

            # Client name: look for "Account" or "Name" in first page
            m = re.search(r'Account\s*\d+\s+Name\s+([A-Z][A-Z\s]+?)(?:\n|Capital)', full_text[:500])
            if m: self.client_name = m.group(1).strip().title()

            rows = self._parse_text(full_text)
        except Exception as e:
            print(f"[WARN] ICICIDirectEquityPDFParser: {e}", file=sys.stderr)
        return rows

    def _parse_text(self, text):
        rows = []
        current_section = 'Short Term'
        lines = text.split('\n')

        for line in lines:
            lu = line.upper()

            # Section markers
            if 'SHORT TERM' in lu and 'CAPITAL GAIN' in lu:
                current_section = 'Short Term'
            elif 'LONG TERM' in lu and 'CAPITAL GAIN' in lu:
                current_section = 'Long Term'
            elif 'SPECULATION' in lu:
                current_section = None  # skip speculation/intraday

            if current_section is None:
                # Check if we're back to a capital gain section
                if 'SHORT TERM' in lu or 'LONG TERM' in lu:
                    pass  # handled above
                continue

            isin_m = self._ISIN_RE.search(line)
            if not isin_m:
                continue

            isin = isin_m.group(1)

            # Symbol: last UPPERCASE word before ISIN
            before = line[:isin_m.start()]
            before_clean = re.sub(
                r'Short\s+Term|Long\s+Term|Speculation|Capital\s+Gain|\(STT\s+paid\)',
                '', before, flags=re.I
            ).strip()
            parts_before = before_clean.split()
            symbol = parts_before[-1] if parts_before else ''
            if not symbol or not symbol.replace('-','').isalnum():
                continue

            # After ISIN: QTY SALE_DATE SALE_RATE SALE_VALUE SALE_EXP BUY_DATE BUY_RATE BUY_VALUE BUY_EXP P/L
            after = line[isin_m.end():].strip()
            dates = self._DATE_RE.findall(after)
            if len(dates) < 2:
                continue

            after_no_dates = self._DATE_RE.sub(' ', after)
            raw_nums = re.findall(r'-?[\d,]+\.?\d*', after_no_dates)
            floats = []
            for n in raw_nums:
                try:
                    v = float(n.replace(',', ''))
                    floats.append(v)
                except ValueError:
                    pass

            if len(floats) < 6:
                continue

            try:
                qty        = floats[0]
                sale_rate  = floats[1]
                sale_value = floats[2]
                sale_exp   = floats[3] if floats[3] < sale_value else 0  # sanity check
                buy_rate   = floats[4] if len(floats) > 4 else 0
                buy_value  = floats[5] if len(floats) > 5 else 0
                buy_exp    = floats[6] if len(floats) > 6 and floats[6] < buy_value else 0
            except IndexError:
                continue

            sale_date = parse_date(dates[0])
            buy_date  = parse_date(dates[1])
            if not sale_date or not buy_date:
                continue
            if not sale_value or sale_value <= 0:
                continue

            transfer_exp = (sale_exp + buy_exp) if (sale_exp or buy_exp) else None

            rows.append(make_row(
                self.source_name, symbol, isin,
                sale_date, qty, sale_rate, sale_value,
                buy_date, buy_value, current_section, self.client_name,
                transfer_expenses=transfer_exp,
            ))

        # Deduplicate
        seen, unique = set(), []
        for r in rows:
            k = (str(r.get('sale_date')), str(r.get('purchase_date')),
                 (r.get('share_name') or '').lower()[:20],
                 round(float(r.get('sale_amount') or 0), 0))
            if k not in seen:
                seen.add(k)
                unique.append(r)
        return unique


# ─────────────────────────────────────────────
# CAMS XLS PARSER (Excel .xls with TRXN_DETAILS)
# ─────────────────────────────────────────────
class CAMSXLSParser:
    """Parses CAMS Capital Gain / Loss Statement in XLS format.
    Sheet: TRXN_DETAILS — each row is one purchase-lot matched to a redemption.
    Columns: AMC Name, Folio No, ASSET CLASS, NAME, STATUS, PAN, GUARDIAN_PAN,
             Scheme Name, Desc, Date (sale), Units, Amount (sale), Price (sale NAV), STT,
             Desc_1 (buy type), Date_1 (buy), PurhUnit, RedUnits, Unit Cost,
             Indexed Cost, Grandfathered Units, NAV 31/01/2018,
             Market Value 31/01/2018, Short Term, LT With Index, LT Without Index, ...
    """
    _ISIN_RE = re.compile(r'ISIN\s*:\s*(IN[EF][A-Z0-9]{9,12})', re.I)
    _GF_CUTOFF = datetime(2018, 1, 31)

    def parse(self, file_path, source_name=None):
        self.source_name = source_name or source_name_from_file(file_path)
        self.client_name = None
        rows = []
        try:
            import xlrd
            wb = xlrd.open_workbook(file_path)

            # Client name from INVESTOR_DETAILS
            if 'INVESTOR_DETAILS' in wb.sheet_names():
                ws_inv = wb.sheet_by_name('INVESTOR_DETAILS')
                for ri in range(ws_inv.nrows):
                    row_vals = ws_inv.row_values(ri)
                    if 'INV_NAME' in str(row_vals):
                        continue
                    if len(row_vals) >= 2 and row_vals[1] and str(row_vals[1]).strip():
                        self.client_name = str(row_vals[1]).strip().title()
                        break

            if 'TRXN_DETAILS' not in wb.sheet_names():
                return rows
            ws = wb.sheet_by_name('TRXN_DETAILS')

            # Find header row
            hdr_idx = None
            col = {}
            for ri in range(min(10, ws.nrows)):
                row_vals = ws.row_values(ri)
                row_str = [str(v).strip() for v in row_vals]
                if 'Scheme Name' in row_str or 'AMC Name' in row_str:
                    hdr_idx = ri
                    for j, h in enumerate(row_str):
                        hl = h.lower().strip()
                        if hl == 'amc name':            col['amc'] = j
                        elif hl == 'asset class':       col['asset_class'] = j
                        elif hl == 'name':              col['inv_name'] = j
                        elif hl == 'scheme name':       col['scheme'] = j
                        elif hl == 'desc':              col['desc'] = j
                        elif hl == 'date':              col['sale_date'] = j
                        elif hl == 'units':             col['units'] = j
                        elif hl == 'amount':            col['amount'] = j
                        elif hl == 'price':             col['price'] = j
                        elif hl == 'stt':               col['stt'] = j
                        elif hl == 'desc_1':            col['desc_1'] = j
                        elif hl == 'date_1':            col['buy_date'] = j
                        elif hl == 'purhunit':          col['purch_unit'] = j
                        elif hl == 'redunits':          col['red_units'] = j
                        elif hl == 'unit cost':         col['unit_cost'] = j
                        elif hl == 'indexed cost':      col['indexed_cost'] = j
                        elif 'grandfathered units' in hl: col['gf_units'] = j
                        elif 'nav as on' in hl or 'grandfathered nav' in hl: col['gf_nav'] = j
                        elif 'market value as on' in hl: col['gf_value'] = j
                        elif hl == 'short term':        col['st_gain'] = j
                        elif 'long term with' in hl:    col['lt_with'] = j
                        elif 'long term without' in hl: col['lt_without'] = j
                    break
            if hdr_idx is None:
                return rows

            def gv(row_vals, field):
                idx = col.get(field)
                if idx is None or idx >= len(row_vals): return None
                v = row_vals[idx]
                if v == '' or v is None: return None
                return v

            for ri in range(hdr_idx + 1, ws.nrows):
                row_vals = ws.row_values(ri)
                if not any(str(v).strip() for v in row_vals):
                    continue

                scheme_raw = str(gv(row_vals, 'scheme') or '')
                if not scheme_raw.strip():
                    continue

                # Extract ISIN from scheme name
                isin_m = self._ISIN_RE.search(scheme_raw)
                isin = isin_m.group(1) if isin_m else ''
                # Clean scheme name (remove ISIN suffix)
                scheme_name = re.sub(r',?\s*ISIN\s*:.*', '', scheme_raw, flags=re.I).strip()

                asset_class = str(gv(row_vals, 'asset_class') or '').upper()
                is_equity = 'EQUITY' in asset_class

                desc = str(gv(row_vals, 'desc') or '').strip()
                # Only process redemptions/switch-outs (not purchases)
                if not any(k in desc.lower() for k in ('redemption', 'switch', 'dividend payout')):
                    continue

                sale_date_raw = gv(row_vals, 'sale_date')
                buy_date_raw  = gv(row_vals, 'buy_date')
                sale_date = parse_date(str(sale_date_raw or ''))
                buy_date  = parse_date(str(buy_date_raw  or ''))
                if not sale_date:
                    continue

                price     = to_float(gv(row_vals, 'price'))   or 0
                red_units = to_float(gv(row_vals, 'red_units')) or 0
                unit_cost = to_float(gv(row_vals, 'unit_cost')) or 0
                gf_value  = to_float(gv(row_vals, 'gf_value'))

                if red_units <= 0:
                    continue

                sale_amount     = round(red_units * price, 4)
                purchase_amount = round(red_units * unit_cost, 4)

                # Grandfathering for equity pre-2018
                fmv_total = None
                if (is_equity and gf_value and gf_value > 0
                        and buy_date and buy_date <= self._GF_CUTOFF):
                    fmv_total = round(gf_value, 4)

                gain_type = determine_gain_type(buy_date, sale_date, is_equity)

                # Client name
                if not self.client_name:
                    inv = str(gv(row_vals, 'inv_name') or '').strip()
                    if inv:
                        self.client_name = inv.title()

                rows.append(make_row(
                    self.source_name, scheme_name, isin,
                    sale_date, red_units, price, sale_amount,
                    buy_date, purchase_amount, gain_type, self.client_name,
                    asset_type='Equity' if is_equity else 'Debt',
                    total_market_value_31jan2018=fmv_total,
                ))

        except ImportError:
            print("[WARN] CAMSXLSParser: xlrd not installed. Run: pip install xlrd", file=sys.stderr)
        except Exception as e:
            print(f"[WARN] CAMSXLSParser: {e}", file=sys.stderr)
        return rows


def detect_parser(file_path):
    ext = Path(file_path).suffix.lower()

    if ext in ('.xlsx', '.xls'):
        # Check if file is actually an ASCII text/TSV file (ICICIDirect disguises .xls as TSV)
        try:
            with open(file_path, 'rb') as _f:
                _sig = _f.read(8)
            # Real XLS = D0CF (CFBF), real XLSX = PK (zip). Anything else = text/HTML
            if not (_sig[:2] == b'\xD0\xCF' or _sig[:2] == b'PK'):
                with open(file_path, 'r', encoding='utf-8', errors='replace') as _tf:
                    _snip = _tf.read(300)
                if ('Stock Symbol' in _snip or 'Capital Gain' in _snip or
                        ('Account' in _snip and 'ISIN' in _snip)):
                    return ICICIDirectTSVParser()
        except Exception:
            pass

        try:
            xl = pd.ExcelFile(file_path)
            sheets = xl.sheet_names
            sheets_lower = [s.lower() for s in sheets]

            # Angel One — "Equity+Bonds+SGB Trade Details" sheet
            if any('equity+bonds+sgb' in s or 'equity+bonds' in s for s in sheets_lower):
                return AngelOneExcelParser()

            # CAMS XLS — TRXN_DETAILS sheet
            if 'trxn_details' in sheets_lower:
                return CAMSXLSParser()

            # Groww Stocks — Trade Level + Scrip Level
            if 'trade level' in sheets_lower and 'scrip level' in sheets_lower:
                return GrowwStocksParser()

            # Groww Stocks — Flat format (single sheet with sections)
            df_peek0 = pd.read_excel(file_path, sheet_name=sheets[0], header=None, nrows=5)
            peek_text = ' '.join(str(v) for row in df_peek0.values for v in row if pd.notna(v)).lower()
            if ('unique client code' in peek_text or 'groww' in peek_text or
                    'stock name' in peek_text) and 'capital gains' in peek_text:
                return GrowwStocksFlatParser()

            # Sharekhan — Assets + Equity + Mutual Fund + Derivatives
            if ('assets' in sheets_lower and 'equity' in sheets_lower
                    and 'mutual fund' in sheets_lower and 'derivatives' in sheets_lower):
                return SharekhanSummaryParser()

            # Consolidated MF — TransactionDetails sheet
            if 'transactiondetails' in sheets_lower:
                return ConsolidatedMFParser()

            # MF Capital Gains with Grandfathered Nav column
            cg_sheet = next((s for s in sheets if 'capital gain' in s.lower()), None)
            if cg_sheet:
                df_peek = pd.read_excel(file_path, sheet_name=cg_sheet, header=None, nrows=25)
                full_text = ' '.join(str(v) for row in df_peek.values for v in row if pd.notna(v))
                if 'grandfathered nav' in full_text.lower() or 'Grandfathered Nav' in full_text:
                    return MFCapGainsParser()

            # ICICI Securities equity capital gain Excel
            df_peek_ici = pd.read_excel(file_path, sheet_name=sheets[0], header=None, nrows=6)
            peek_ici = ' '.join(str(v) for row in df_peek_ici.values for v in row if pd.notna(v)).lower()
            if ('stock symbol' in peek_ici and 'sale value' in peek_ici
                    and 'purchase value' in peek_ici):
                return ICICISecuritiesExcelParser()

            # Opulence Wealth — Cost of Acquisition column in header
            df_peek = pd.read_excel(file_path, sheet_name=sheets[0], header=None, nrows=6)
            top_text = ' '.join(str(v) for row in df_peek.values for v in row if pd.notna(v))
            if ('OPULENCE WEALTH' in top_text or 'Cost of Acquisition' in top_text
                    or 'Realized Capital Gain Report' in top_text):
                return OpulenceWealthParser()

            # Dezerv/PMS consolidated format
            if 'consolidated' in sheets_lower or any('capital gain' in s for s in sheets_lower):
                df = pd.read_excel(file_path, sheet_name=sheets[0], header=None)
                title = str(df.iloc[0,0]) if pd.notna(df.iloc[0,0]) else ""
                if 'consolidated capital gain' in title.lower() or 'capital gain / loss' in title.lower():
                    return DezervExcelParser()

            if any('equit' in s for s in sheets_lower) and any('summary' in s for s in sheets_lower):
                return UpstoxExcelParser()
            if any(any(k in s for k in ['tradewise','exit','p&l','pl']) for s in sheets_lower):
                return ZerodhaExcelParser()
            if any(any(k in s for k in ['equity','mutual']) for s in sheets_lower):
                return ZerodhaExcelParser()
        except Exception:
            pass
        return GenericExcelParser()

    if ext == '.csv':
        try:
            df_peek = pd.read_csv(file_path, nrows=6, header=None, encoding='utf-8-sig',
                                  on_bad_lines='skip')
            top_text = ' '.join(str(v) for row in df_peek.values for v in row if pd.notna(v))
            if '31st jan' in top_text.lower() or 'sale expenses' in top_text.lower():
                return AngelBrokingCSVParser()
        except Exception:
            pass
        return GenericExcelParser()

    if ext == '.pdf':
        try:
            with open_pdf(file_path) as pdf:
                text = (pdf.pages[0].extract_text() or "").upper()
        except Exception:
            return GenericPDFParser()

        # ── SPECIFIC FORMAT DETECTIONS (must run FIRST, before broad keyword checks) ──
        # These do a full-document scan to correctly identify formats that share
        # keywords (GRANDFATHER, CAPITAL GAIN, ISIN) with other parsers.
        try:
            with open_pdf(file_path) as _pdf2:
                _full2 = "\n".join((_p.extract_text() or "") for _p in _pdf2.pages)
            _full2_up = _full2.upper()
            # Kuvera: bracket-ISIN format "[ISIN: INFxxx]" unique to kuvera exports
            if "[ISIN:" in _full2 and "CAPITAL GAINS STATEMENT" in _full2_up:
                return KuveraParser()
            # SBI Mutual Fund: "SBI Mutual Fund Investment Gain/Loss Statement"
            if "SBI MUTUAL FUND" in _full2_up and "INVESTMENT GAIN" in _full2_up:
                return SBIMFParser()
            # Anand Rathi family report format
            if "FAMILY REPORT" in _full2_up and "CAPITAL GAIN" in _full2_up:
                return AnandRathiMFParser()
            # CAS-distributor format: "Capital Gain Statement For Financial Year" header
            # with per-folio lot-level table (Primo, Echowin, Alwin-S, similar AMFi distributors)
            if ("CAPITAL GAIN STATEMENT FOR FINANCIAL YEAR" in _full2_up
                    and "(ISIN:" in _full2):
                return AnandRathiMFParser()
            # Asahoo / Alwin-S: "Folio Number:" variant
            if ("(ISIN:" in _full2 and "FOLIO NUMBER" in _full2_up
                    and "CAPITAL GAIN" in _full2_up):
                return AnandRathiMFParser()
            # CAMS Section A/B/C format (Motilal MF, UTI MF, other CAMS-platform AMCs)
            if ("SECTION A : SUBSCRIPTIONS" in _full2_up
                    and "SECTION B : REDEMPTIONS" in _full2_up):
                return CAMSSectionABCParser()
            # Motilal Oswal P&L Transaction Statement (equity, lot-level)
            if ("motilaloswal" in _full2.lower()
                    and "PROFIT AND LOSS TRANSACTION STATEMENT" in _full2_up):
                return MotilalOswalPnLParser()
        except Exception as e:
            print(f"[WARN] detect_parser: specific-format detection failed: {e}", file=sys.stderr)

        # ── STANDARD CHECKS (keyword-based on page 1 only) ──
        # ICICIDirect equity P&L PDF: "Equity Capital Gain" with Sale/Purchase Expenses columns
        if "EQUITY CAPITAL GAIN" in text and "SALE" in text and "PURCHASE" in text and "EXPENSES" in text:
            return ICICIDirectEquityPDFParser()
        if "ICICI PRUDENTIAL PMS" in text or ("STATEMENT OF CAPITAL GAIN" in text and "ICICI" in text):
            return ICICIPMSParser()
        if "MUTUALFUNDSADDA" in text or ("DEBT SHORT TERM" in text and "STP" in text):
            return MFAddaParser()
        if "NIRMAL BANG" in text:
            return NirmalBangParser()
        if "MOTILAL OSWAL" in text or "MOFSL" in text or (
                "GRANDFATHER" in text and "GAINLOSS" in text):
            return MotilalOswalParser()
        if "HDFC BANK" in text or ("REALIZED GAIN" in text and "CUSTOMER ID" in text) or \
                ("REALIZED GAIN" in text and "RISK PROFILE" in text):
            return HDFCBankParser()
        if "PURNARTHA" in text:
            return PurnarthaParser()
        if "MIRAE ASSET" in text or "KFINTECH" in text:
            return MiraeAssetParser()
        if "CAMS" in text or "COMPUTER AGE" in text:
            return CAMSParser()
        if "FRANKLIN TEMPLETON" in text or (
                "FRANKLIN" in text and "INVESTMENT GAIN" in text):
            return FranklinParser()
        if "CAPITAL GAIN" in text and "ISIN" in text:
            return CAMSParser()

        # ── NEW FORMAT DETECTIONS (multi-page scan) ──────────────────
        # Read full text of all pages for more specific matching
        return GenericPDFParser()

    return None


# ─────────────────────────────────────────────
# ZIP HANDLER — password = last segment before final extension
# ─────────────────────────────────────────────

def extract_zip(zip_path):
    stem = Path(zip_path).stem  # e.g. "vansh.1.pass"
    passwords_to_try = []
    # Last dot-segment before extension
    if '.' in stem:
        passwords_to_try.append(stem.rsplit('.', 1)[-1])
    # Last space/underscore word
    parts = re.split(r'[\s_\-]+', stem.replace('.', ' '))
    if parts: passwords_to_try.append(parts[-1])
    passwords_to_try.append(None)

    extracted = []
    with zipfile.ZipFile(zip_path) as zf:
        for pwd in passwords_to_try:
            try:
                pwd_bytes = pwd.encode() if pwd else None
                for name in zf.namelist():
                    if name.endswith('/') or name.startswith('__MACOSX'): continue
                    ext = Path(name).suffix.lower()
                    if ext not in ('.pdf','.xlsx','.xls','.csv'): continue
                    data = zf.read(name, pwd=pwd_bytes)
                    tmp_path = f"/tmp/cg_zip_{Path(zip_path).stem}_{name.replace('/','_')}"
                    with open(tmp_path,'wb') as f: f.write(data)
                    extracted.append(tmp_path)
                if extracted:
                    print(f"[INFO] ZIP extracted {len(extracted)} files", file=sys.stderr)
                    return extracted
            except: continue
    print(f"[WARN] Could not extract ZIP: {zip_path}", file=sys.stderr)
    return []


# ─────────────────────────────────────────────
# DATE RANGE FILTER
# ─────────────────────────────────────────────

def filter_by_date_range(rows, fy_start, fy_end):
    """Filter rows by financial year sale date range."""
    if not fy_start and not fy_end: return rows, []
    filtered = []
    out_of_range = []
    for r in rows:
        sd = r.get('sale_date')
        if sd is None: filtered.append(r); continue
        if fy_start and sd < fy_start: out_of_range.append(r); continue
        if fy_end and sd > fy_end: out_of_range.append(r); continue
        filtered.append(r)
    return filtered, out_of_range


# ─────────────────────────────────────────────
# EXCEL OUTPUT WRITER
# ─────────────────────────────────────────────

HEADERS = [
    "Source\n(Statement Name)", "Share Name\n(Select/Fill share name)",
    "Share Type\n(Listed/unlisted,etc.)", "ISIN Code",
    "Sale Date\n(dd/mm/yyyy) Format", "Quantity", "Sale Price per share",
    "Sale Amount\n", "Market price per share on 31/01/2018",
    "Total Market Value as on 31/01/2018\n (In case of Long Term Transaction Tax)",
    "Transfer\nExpenses\n", "Purchase\nDate (dd/mm/yyyy) Format",
    "Purchase\nAmt", "Fair Market Value in case of Unquated Shares (Only for Non Resident)",
    "Gain Type\n(Short/Long Term)", "Sale Period\n(Do not fill if sale date is filled)",
    "Purchase Yr.\n(Do not fill if purch. Date is filled)", "1st Proviso u/s 48 applied",
    "Other Special Rate For NRI", "Bonds & Debenture etc.. (No Indexing)",
    "Mutual Funds", "select Section for 94(7)/97(8) applicable on short term only",
    "Dividend\n (applicable on short term only)", ".", "Capital Gain / Loss",
    "AIS / TIS Reconciliation",
]

def _build_row_values(row, r_idx):
    """Return the full 26-column values list + cg_formula for one data row."""
    sale_val     = row.get("sale_amount") or 0
    purch_val    = row.get("purchase_amount") or 0
    fmv_price    = row.get("market_price_31jan2018")
    fmv_total    = row.get("total_market_value_31jan2018")
    transfer_exp = row.get("transfer_expenses")
    H = f"H{r_idx}"; J = f"J{r_idx}"; M = f"M{r_idx}"; K = f"K{r_idx}"
    cg_formula = f"=+{H}-IF(ISBLANK({J}),{M},MAX(MIN({H},{J}),{M}))-{K}"
    values = [
        row.get("source",""),
        row.get("share_name",""),
        row.get("share_type","Transaction Tax"),
        row.get("isin",""),
        row.get("sale_date"),
        row.get("quantity"),
        row.get("sale_price"),
        sale_val,
        fmv_price,
        fmv_total,
        transfer_exp,
        row.get("purchase_date"),
        purch_val,
        None,
        row.get("gain_type",""),
        None,
        None,
        "No",
        "No",
        "No",
        row.get("mutual_fund_col","Equity Mutual Fund"),
        row.get("section","112A"),
        None,
        ".",
        cg_formula,
        None,
    ]
    return values, apply_grandfathering(sale_val, fmv_total, purch_val, transfer_exp)


def _write_verification_sheet(wb, rows, file_meta):
    """Add a Verification sheet to workbook wb using file_meta for process status.

    file_meta is a list of dicts:
      { source, client_name, pan, process_status, oor_count, remark }
    process_status: 'Processed' | 'Semi-Processed' | 'Not Processed'
    """
    from collections import defaultdict

    def _is_equity(r):
        # share_type is the authoritative classification set by _determine_share_type
        if r.get("share_type") == "Unlisted Security":
            return False
        sec    = str(r.get("section","")).lower()
        sctype = str(r.get("scheme_type","")).lower()
        mfcol  = str(r.get("mutual_fund_col","")).lower()
        atype  = str(r.get("asset_type","")).lower()
        # Check explicit debt from parser FIRST — make_row hardcodes scheme_type="Equity"
        # so checking atype/mfcol before sctype prevents debt rows being miscategorised.
        if "debt" in atype or "debt" in mfcol: return False
        if "equity" in sctype or "equity" in mfcol or "equity" in atype: return True
        if "112a" in sec or sec == "111a": return True
        if "debt" in sctype: return False
        return True

    # Aggregate rows by source
    src_rows_map = defaultdict(list)
    for r in rows:
        src_rows_map[r.get("source","")].append(r)

    ws = wb.create_sheet("Verification")
    VER_HEADERS = [
        "Source Name", "Process Status", "Client Name", "PAN",
        "Sale Value", "Purchase Value",
        "Capital Gains - LT Equity", "Capital Gains - ST Equity",
        "Capital Gains - LT Debt", "Capital Gains - ST Debt",
        "Capital Gains - Others", "Remark",
        # PDF validation columns
        "Val Status", "PDF ST Gain", "PDF LT Gain", "PDF Total Gain",
        "Extracted ST", "Extracted LT", "Extracted Total",
        "Gap ST (₹)", "Gap ST %", "Gap LT (₹)", "Gap LT %",
        "Gap Total (₹)", "Gap Total %", "Causes",
    ]
    vh_font  = Font(bold=True, size=9)
    vh_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(VER_HEADERS, 1):
        cell = ws.cell(1, ci, h)
        cell.font = vh_font; cell.alignment = vh_align
    ws.row_dimensions[1].height = 30

    # Status badge colours
    STATUS_COLOURS = {
        "Processed":      "C6EFCE",   # green
        "Semi-Processed": "FFEB9C",   # yellow
        "Not Processed":  "FFC7CE",   # red
    }
    alt_fill = PatternFill("solid", fgColor="EBF3FB")

    # Build rows — one row per file_meta entry
    data_rows = len(file_meta)
    for vi, fm in enumerate(file_meta, 2):
        src    = fm["source"]
        status = fm["process_status"]
        src_r  = src_rows_map.get(src, [])
        client = fm.get("client_name") or next(
            (r.get("client_name") for r in src_r if r.get("client_name")), "—")
        pan    = fm.get("pan") or next(
            (r.get("pan") for r in src_r if r.get("pan")), "—")
        remark = fm.get("remark","")
        if fm.get("oor_count",0) > 0 and not remark:
            remark = f"{fm['oor_count']} transaction(s) excluded (out of FY range)"

        sale_val = purch_val = lt_eq = st_eq = lt_debt = st_debt = others = 0.0
        for r in src_r:
            sale_val  += r.get("sale_amount") or 0
            purch_val += r.get("purchase_amount") or 0
            cg = apply_grandfathering(
                r.get("sale_amount") or 0,
                r.get("total_market_value_31jan2018"),
                r.get("purchase_amount") or 0,
                r.get("transfer_expenses"),
            )
            gt = r.get("gain_type","Short Term")
            eq = _is_equity(r)
            if eq and gt == "Long Term":    lt_eq   += cg
            elif eq and gt == "Short Term": st_eq   += cg
            elif gt == "Long Term":         lt_debt += cg
            elif gt == "Short Term":        st_debt += cg
            else:                           others  += cg

        # Pull pdf_validation from file_meta entry (raw build_validation_report() dict)
        _pv = fm.get("pdf_validation") or {}
        if _pv:
            _gap_pct     = _pv.get("gap_pct") or 0
            _gap_st_pct  = _pv.get("gap_st_pct") or 0
            _gap_lt_pct  = _pv.get("gap_lt_pct") or 0
            _worst       = max(_gap_pct, _gap_st_pct, _gap_lt_pct)
            val_status   = "OK" if _worst < 0.5 else ("WARN" if _worst < 3.0 else "REVIEW")
            causes_text  = "; ".join(c.get("note","") for c in (_pv.get("causes") or []))
        else:
            val_status  = "NO_VAL"
            causes_text = ""

        ver_values = [
            src, status, client, pan,
            round(sale_val,2), round(purch_val,2),
            round(lt_eq,2), round(st_eq,2),
            round(lt_debt,2), round(st_debt,2),
            round(others,2), remark,
            # validation columns (raw build_validation_report keys)
            val_status,
            _pv.get("pdf_st_gain"),    _pv.get("pdf_lt_gain"),    _pv.get("pdf_total_gain"),
            _pv.get("extracted_st_gain"), _pv.get("extracted_lt_gain"), _pv.get("extracted_total_gain"),
            _pv.get("gap_st"),  _pv.get("gap_st_pct"),
            _pv.get("gap_lt"),  _pv.get("gap_lt_pct"),
            _pv.get("gap"),     _pv.get("gap_pct"),
            causes_text,
        ]
        VAL_STATUS_COLOURS = {"OK": "C6EFCE", "WARN": "FFEB9C", "REVIEW": "FFC7CE", "NO_VAL": "EDEDED"}
        # column indices for number formatting (1-based)
        MONEY_COLS = {5,6,7,8,9,10,11, 14,15,16,17,18,19,20,22,24}
        PCT_COLS   = {21, 23, 25}
        row_fill = PatternFill("solid", fgColor="EBF3FB") if vi % 2 == 0 else None
        for ci, val in enumerate(ver_values, 1):
            cell = ws.cell(vi, ci, val)
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if row_fill: cell.fill = row_fill
            if ci in MONEY_COLS and val is not None: cell.number_format = '#,##0.00'
            if ci in PCT_COLS   and val is not None: cell.number_format = '0.00"%"'
            # Process Status colour
            if ci == 2:
                sc = STATUS_COLOURS.get(status)
                if sc:
                    cell.fill = PatternFill("solid", fgColor=sc)
                    cell.font = Font(bold=True, size=9)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            # Val Status colour (col 13)
            if ci == 13:
                vc = VAL_STATUS_COLOURS.get(val_status, "EDEDED")
                cell.fill = PatternFill("solid", fgColor=vc)
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Red text for gap cells that are non-trivial
            if ci in {21, 23, 25} and val is not None and abs(val) >= 1.0:
                cell.font = Font(size=9, color="C00000", bold=True)
            # Causes cell wrap
            if ci == 26:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Grand total row (original cols only)
    last_r = data_rows + 2
    ws.cell(last_r, 1, "GRAND TOTAL").font = Font(bold=True, size=9)
    for ci in range(5, 12):
        col_letter = get_column_letter(ci)
        formula = f"=SUM({col_letter}2:{col_letter}{last_r-1})"
        cell = ws.cell(last_r, ci, formula)
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True, size=9)

    ver_widths = [30, 16, 22, 14, 16, 16, 22, 22, 18, 18, 18, 30,
                  10, 14, 14, 14, 14, 14, 14, 13, 9, 13, 9, 13, 9, 45]
    for ci, w in enumerate(ver_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


def _inject_pivot_table(wb, summary_ws, rows, pt_start_row):
    """No-op — pivot injection is handled by _post_save_inject_pivot after wb.save()."""
    pass


# ── Template XML strings copied verbatim from the user's demo file ──
# ── Template XML strings copied from user's demo file ──
# (cache definition is built dynamically per-run with exact row count)

_PT_CACHE_RECORDS = '''\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotCacheRecords xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" count="0"/>'''

_PT_CACHE_DEF_RELS = '''\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
 <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheRecords" Target="pivotCacheRecords1.xml"/>
</Relationships>'''

# PT1: Gain Type → Source → Share Type  (left table, cols A-D)
# No pre-populated <items>/<rowItems>/<colItems> — Excel fills these on refresh.
# Removed non-standard dataField="1" attr and baseField/baseItem on dataFields.
_PT1_XML = '''\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotTableDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 name="PivotTable1" cacheId="11"
 applyNumberFormats="0" applyBorderFormats="0" applyFontFormats="0"
 applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1"
 dataCaption="Values" createdVersion="6" updatedVersion="6" minRefreshableVersion="3"
 useAutoFormatting="1" itemPrintTitles="1"
 indent="0" outline="1" outlineData="1" multipleFieldFilters="0">
 <location ref="A3:D50" firstHeaderRow="1" firstDataRow="2" firstDataCol="1" rowPageCount="0" colPageCount="0"/>
 <pivotFields count="25">
  <pivotField axis="axisRow" showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField axis="axisRow" showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField axis="axisRow" showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField/>
 </pivotFields>
 <rowFields count="3"><field x="14"/><field x="0"/><field x="2"/></rowFields>
 <colFields count="1"><field x="-2"/></colFields>
 <dataFields count="3">
  <dataField name="Sum of Sale Amount" fld="7" numFmtId="2"/>
  <dataField name="Sum of Purchase Amt" fld="12" numFmtId="2"/>
  <dataField name="Sum of Capital Gain / Loss" fld="24" numFmtId="2"/>
 </dataFields>
 <pivotTableStyleInfo name="PivotStyleLight16" showRowHeaders="1" showColHeaders="1"
  showRowStripes="0" showColStripes="0" showLastColumn="1"/>
</pivotTableDefinition>'''

# PT2: Gain Type → Share Type  (right table, cols F-I, no Source level)
_PT2_XML = '''\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotTableDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 name="PivotTable2" cacheId="11"
 applyNumberFormats="0" applyBorderFormats="0" applyFontFormats="0"
 applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1"
 dataCaption="Values" createdVersion="6" updatedVersion="6" minRefreshableVersion="3"
 useAutoFormatting="1" itemPrintTitles="1"
 indent="0" outline="1" outlineData="1" multipleFieldFilters="0">
 <location ref="F3:I50" firstHeaderRow="1" firstDataRow="2" firstDataCol="1" rowPageCount="0" colPageCount="0"/>
 <pivotFields count="25">
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField axis="axisRow" showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField axis="axisRow" showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField showAll="0" defaultSubtotal="0"/>
  <pivotField/>
 </pivotFields>
 <rowFields count="2"><field x="14"/><field x="2"/></rowFields>
 <colFields count="1"><field x="-2"/></colFields>
 <dataFields count="3">
  <dataField name="Sum of Sale Amount" fld="7" numFmtId="2"/>
  <dataField name="Sum of Purchase Amt" fld="12" numFmtId="2"/>
  <dataField name="Sum of Capital Gain / Loss" fld="24" numFmtId="2"/>
 </dataFields>
 <pivotTableStyleInfo name="PivotStyleLight16" showRowHeaders="1" showColHeaders="1"
  showRowStripes="0" showColStripes="0" showLastColumn="1"/>
</pivotTableDefinition>'''

_PT_TABLE_RELS = '''\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
 <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" Target="../pivotCache/pivotCacheDefinition1.xml"/>
</Relationships>'''


def _post_save_inject_pivot(output_path, row_count):
    """Patch the saved xlsx zip to inject live PivotTable XML.

    Called once after openpyxl's wb.save(). Adds:
      - pivotCacheDefinition1.xml  (source = Capital Gains!A2:Y{last_row})
      - pivotCacheRecords1.xml     (empty — Excel refreshes on open)
      - pivotTable1.xml            (Gain Type → Source → Share Type, cols A-D)
      - pivotTable2.xml            (Gain Type → Share Type, cols F-I)
      - all .rels files
    And patches workbook.xml, workbook.xml.rels, sheet2.xml.rels,
    [Content_Types].xml to wire everything together.
    """
    import zipfile, shutil, re, os, string as _string

    # Capital Gains sheet: row 2 = column headers, rows 3..N = data
    last_data_row = row_count + 2   # header on row 2, data rows 3..(row_count+2)

    # numFmtId per column A-Y (0-indexed): 0=general, 2=number, 14=date
    _CACHE_NUM_FMTS = [0, 0, 0, 0, 14, 0, 2, 2, 2, 2, 2, 14, 2, 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2]

    tmp = output_path + "._pivot_tmp"
    shutil.copy2(output_path, tmp)

    try:
        with zipfile.ZipFile(tmp, "r") as zin:
            names = set(zin.namelist())

            wb_xml       = zin.read("xl/workbook.xml").decode("utf-8")
            wb_rels      = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            ct_xml       = zin.read("[Content_Types].xml").decode("utf-8")
            styles_xml   = zin.read("xl/styles.xml").decode("utf-8")
            sh1_xml      = zin.read("xl/worksheets/sheet1.xml").decode("utf-8")

            # ── Read actual column headers from sheet1 row 2 (A-Y) ───────
            # Build cache field names dynamically so they always match the
            # real sheet headers — hardcoded names caused Excel repair dialogs
            # when trailing spaces / newlines differed by even one character.
            _row2_m = re.search(r'<row r="2".*?</row>', sh1_xml, re.DOTALL)
            _col_map = {}
            if _row2_m:
                for _ref, _content in re.findall(r'<c r="([A-Y])2"[^>]*>(.*?)</c>',
                                                  _row2_m.group(0), re.DOTALL):
                    _t = re.search(r'<t[^>]*>(.*?)</t>', _content, re.DOTALL)
                    _col_map[_ref] = _t.group(1) if _t else ''
            # Encode cell text as OOXML cache field name (newlines → _x000a_)
            _actual_headers = [
                _col_map.get(c, '').replace('\n', '_x000a_').replace('\r', '')
                for c in _string.ascii_uppercase[:25]
            ]
            _cache_fields_xml = '\n'.join(
                f'  <cacheField name="{h}" numFmtId="{nf}"><sharedItems containsBlank="1"/></cacheField>'
                for h, nf in zip(_actual_headers, _CACHE_NUM_FMTS)
            )
            cache_def = f'''\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotCacheDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
 r:id="rId1" createdVersion="6" refreshedVersion="6" minRefreshableVersion="3"
 recordCount="0" refreshOnLoad="1">
 <cacheSource type="worksheet">
  <worksheetSource ref="A2:Y{last_data_row}" sheet="Capital Gains"/>
 </cacheSource>
 <cacheFields count="25">
{_cache_fields_xml}
 </cacheFields>
</pivotCacheDefinition>'''

            sh2_rels_path = "xl/worksheets/_rels/sheet2.xml.rels"
            sh2_rels = zin.read(sh2_rels_path).decode("utf-8") if sh2_rels_path in names else \
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>'

            # ── Clear static cell data from Summary sheet ─────────────
            # openpyxl writes a pre-built static table into sheet2 cells.
            # Those cells conflict with the pivot table ranges (A3:D50, F3:I50),
            # causing Excel's "found a problem / repaired content" dialog.
            # Wipe sheetData so the pivot tables own their ranges cleanly.
            sh2_path = "xl/worksheets/sheet2.xml"
            if sh2_path in names:
                sh2_xml = zin.read(sh2_path).decode("utf-8")
                sh2_xml = re.sub(r'<dimension[^/]*/>', '<dimension ref="A1"/>', sh2_xml)
                sh2_xml = re.sub(r'<sheetData>.*?</sheetData>', '<sheetData/>', sh2_xml, flags=re.DOTALL)
            else:
                sh2_xml = None

            # ── Patch styles.xml ─────────────────────────────────
            # openpyxl omits <name> from font elements it writes — OOXML requires it.
            # Also fix bare <patternFill/> which needs explicit patternType attribute.
            def _add_font_name(m):
                inner = m.group(1)
                if '<name ' in inner:
                    return m.group(0)
                return '<font><name val="Calibri"/>' + inner + '</font>'
            styles_xml = re.sub(r'<font>(.*?)</font>', _add_font_name, styles_xml, flags=re.DOTALL)
            styles_xml = styles_xml.replace('<patternFill/>', '<patternFill patternType="none"/>')

            # ── Patch workbook.xml ────────────────────────────────
            # Remove empty <workbookProtection/> — invalid without attributes in strict OOXML.
            wb_xml = wb_xml.replace('<workbookProtection/>', '')
            if "<pivotCaches>" not in wb_xml:
                wb_xml = wb_xml.replace(
                    "</workbook>",
                    '<pivotCaches>'
                    '<pivotCache cacheId="11" '
                    'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
                    'r:id="rId_ptcache1"/>'
                    '</pivotCaches></workbook>'
                )

            # ── Patch workbook.xml.rels ───────────────────────────
            if "pivotCacheDefinition" not in wb_rels:
                wb_rels = wb_rels.replace(
                    "</Relationships>",
                    '<Relationship Id="rId_ptcache1" '
                    'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" '
                    'Target="pivotCache/pivotCacheDefinition1.xml"/></Relationships>'
                )

            # ── Patch sheet2.xml.rels ─────────────────────────────
            sh2_rels = re.sub(
                r'<Relationships([^>]*)/>\s*$',
                r'<Relationships\1></Relationships>',
                sh2_rels.strip()
            )
            for rid, tgt in [("rId_pt1", "../pivotTables/pivotTable1.xml"),
                              ("rId_pt2", "../pivotTables/pivotTable2.xml")]:
                if tgt not in sh2_rels:
                    sh2_rels = sh2_rels.replace(
                        "</Relationships>",
                        f'<Relationship Id="{rid}" '
                        f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable" '
                        f'Target="{tgt}"/></Relationships>'
                    )

            # ── Patch [Content_Types].xml ─────────────────────────
            new_ct_parts = {
                "/xl/pivotCache/pivotCacheDefinition1.xml":
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml",
                "/xl/pivotCache/pivotCacheRecords1.xml":
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheRecords+xml",
                "/xl/pivotTables/pivotTable1.xml":
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml",
                "/xl/pivotTables/pivotTable2.xml":
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml",
            }
            for pn, ct in new_ct_parts.items():
                if pn not in ct_xml:
                    ct_xml = ct_xml.replace(
                        "</Types>",
                        f'<Override PartName="{pn}" ContentType="{ct}"/></Types>'
                    )

            # ── Write new zip ─────────────────────────────────────
            with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    fn = item.filename
                    if fn == "xl/workbook.xml":
                        zout.writestr(item, wb_xml.encode("utf-8"))
                    elif fn == "xl/_rels/workbook.xml.rels":
                        zout.writestr(item, wb_rels.encode("utf-8"))
                    elif fn == "[Content_Types].xml":
                        zout.writestr(item, ct_xml.encode("utf-8"))
                    elif fn == "xl/styles.xml":
                        zout.writestr(item, styles_xml.encode("utf-8"))
                    elif fn == sh2_rels_path:
                        zout.writestr(item, sh2_rels.encode("utf-8"))
                    elif fn == sh2_path and sh2_xml is not None:
                        zout.writestr(item, sh2_xml.encode("utf-8"))
                    elif fn.startswith("xl/worksheets/") and fn.endswith(".xml"):
                        ws_xml = zin.read(fn).decode("utf-8")
                        # Strip empty <v></v> openpyxl writes for uncalculated formula cells
                        ws_xml = re.sub(r'<v/>', '', ws_xml)
                        ws_xml = re.sub(r'<v></v>', '', ws_xml)
                        # Fix inlineStr cells with no <is> child — missing child is invalid
                        # OOXML and triggers the "found a problem" repair dialog every open.
                        ws_xml = re.sub(
                            r'(<c[^>]*t="inlineStr"[^>]*>)\s*</c>',
                            r'\1<is><t/></is></c>',
                            ws_xml
                        )
                        zout.writestr(item, ws_xml.encode("utf-8"))
                    else:
                        zout.writestr(item, zin.read(fn))

                if sh2_rels_path not in names:
                    zout.writestr(sh2_rels_path, sh2_rels.encode("utf-8"))

                zout.writestr("xl/pivotCache/pivotCacheDefinition1.xml",
                               cache_def.encode("utf-8"))
                zout.writestr("xl/pivotCache/pivotCacheRecords1.xml",
                               _PT_CACHE_RECORDS.encode("utf-8"))
                zout.writestr("xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels",
                               _PT_CACHE_DEF_RELS.encode("utf-8"))
                zout.writestr("xl/pivotTables/pivotTable1.xml",
                               _PT1_XML.encode("utf-8"))
                zout.writestr("xl/pivotTables/pivotTable2.xml",
                               _PT2_XML.encode("utf-8"))
                zout.writestr("xl/pivotTables/_rels/pivotTable1.xml.rels",
                               _PT_TABLE_RELS.encode("utf-8"))
                zout.writestr("xl/pivotTables/_rels/pivotTable2.xml.rels",
                               _PT_TABLE_RELS.encode("utf-8"))

        os.remove(tmp)
        print(f"[INFO] PivotTable injected into {output_path}", file=sys.stderr)

    except Exception as e:
        if os.path.exists(tmp):
            shutil.move(tmp, output_path)
        print(f"[WARN] PivotTable injection failed: {e}", file=sys.stderr)


def write_output_excel(rows, output_path, file_meta=None):
    """Write the full output Excel (cols A-Z) with Capital Gains, Summary, Verification sheets."""
    from collections import defaultdict

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Capital Gains"

    total_sale = sum(r.get("sale_amount") or 0 for r in rows)
    total_purchase = sum(r.get("purchase_amount") or 0 for r in rows)
    total_gain = sum(
        apply_grandfathering(
            r.get("sale_amount") or 0,
            r.get("total_market_value_31jan2018"),
            r.get("purchase_amount") or 0,
            r.get("transfer_expenses"),
        )
        for r in rows
    )

    ws.cell(1,1).value = None
    ws.cell(1,2).value = "For Computax Template copy from Column B to W"
    ws.cell(1,8).value = total_sale
    ws.cell(1,13).value = total_purchase
    ws.cell(1,25).value = total_gain

    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF", size=9)
    hdr_align = Alignment(wrap_text=True, horizontal="center", vertical="center")
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(2, col, header)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = hdr_align
    ws.row_dimensions[2].height = 45

    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    date_fmt = "dd/mm/yyyy"
    for r_idx, row in enumerate(rows, 3):
        fill = alt_fill if r_idx % 2 == 0 else None
        values, _ = _build_row_values(row, r_idx)
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(r_idx, c_idx, val)
            if fill: cell.fill = fill
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if c_idx in (5,12) and isinstance(val, datetime): cell.number_format = date_fmt
            if c_idx in (6,7,8,13,25): cell.number_format = '#,##0.00'

    widths = [14,42,18,16,14,10,14,14,16,18,12,14,14,20,14,12,12,10,10,10,18,10,10,4,16,10]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    # ── Summary Sheet — real Excel PivotTable ─────────────────────
    ws2 = wb.create_sheet("Summary")
    # Row 1 intentionally left blank — pivot tables start at row 3

    # The PivotTable reads from the Capital Gains sheet.
    # openpyxl cannot build a native PT cache; instead we write the PT XML
    # definition directly into the workbook parts so Excel renders it live.
    # We also write a static "screenshot" of the pivot below the instruction
    # so the data is visible even without refreshing.

    # ── static pivot (always visible, matches PT data) ────────────
    HDR_FILL  = PatternFill("solid", fgColor="2E75B6")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    GT_FILL   = PatternFill("solid", fgColor="D6E4F0")
    GT_FONT   = Font(bold=True, size=10)
    SRC_FONT  = Font(bold=True, size=9)
    ST_FONT   = Font(size=9)
    TOT_FONT  = Font(bold=True, size=10)
    GRAND_FONT= Font(bold=True, size=11)
    NUM_FMT   = '#,##0.00'
    CGAIN_NUM = '#,##0.00;[Red]-#,##0.00'

    def _hdr_row(ws, start_col, row=3):
        labels = ["Row Labels","Sum of Sale Amount","Sum of PurchaseAmt","Sum of Capital Gain / Loss"]
        for i, lbl in enumerate(labels):
            c = ws.cell(row, start_col + i, lbl)
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _num(ws, row, col, val, is_cg=False):
        c = ws.cell(row, col, round(val, 2))
        c.number_format = CGAIN_NUM if is_cg else NUM_FMT
        return c

    # Build in-memory pivot: gain_type → source → share_type → {sale, purchase}
    pivot = defaultdict(lambda: defaultdict(lambda: defaultdict(
        lambda: {"sale": 0.0, "purchase": 0.0})))
    for r in rows:
        gt  = r.get("gain_type", "Short Term")
        src = r.get("source", "")
        st  = r.get("share_type", "Transaction Tax")
        pivot[gt][src][st]["sale"]     += r.get("sale_amount") or 0
        pivot[gt][src][st]["purchase"] += r.get("purchase_amount") or 0

    gain_order = ["Short Term", "Long Term"]

    # ── LEFT TABLE: Gain Type → Source → Share Type ───────────────
    _hdr_row(ws2, start_col=1)
    rn = 4
    g_sale = g_purch = 0.0
    for gt in gain_order:
        if gt not in pivot:
            continue
        gt_sale = gt_purch = 0.0
        c = ws2.cell(rn, 1, gt); c.font = GT_FONT; c.fill = GT_FILL
        rn += 1
        for src, stypes in sorted(pivot[gt].items()):
            src_sale = src_purch = 0.0
            ws2.cell(rn, 1, f"  {src}").font = SRC_FONT
            rn += 1
            for st, vals in stypes.items():
                cg = vals["sale"] - vals["purchase"]
                ws2.cell(rn, 1, f"    {st}").font = ST_FONT
                _num(ws2, rn, 2, vals["sale"])
                _num(ws2, rn, 3, vals["purchase"])
                _num(ws2, rn, 4, cg, is_cg=True)
                src_sale += vals["sale"]; src_purch += vals["purchase"]
                rn += 1
            gt_sale += src_sale; gt_purch += src_purch
        c2 = ws2.cell(rn, 1, f"{gt} Total"); c2.font = TOT_FONT
        _num(ws2, rn, 2, gt_sale); _num(ws2, rn, 3, gt_purch)
        _num(ws2, rn, 4, gt_sale - gt_purch, is_cg=True)
        g_sale += gt_sale; g_purch += gt_purch
        rn += 1

    c3 = ws2.cell(rn, 1, "Grand Total"); c3.font = GRAND_FONT
    _num(ws2, rn, 2, g_sale); _num(ws2, rn, 3, g_purch)
    _num(ws2, rn, 4, g_sale - g_purch, is_cg=True)

    # ── RIGHT TABLE: Gain Type → Share Type (consolidated) ────────
    RIGHT_COL = 6
    _hdr_row(ws2, start_col=RIGHT_COL)
    cpivot = defaultdict(lambda: defaultdict(lambda: {"sale": 0.0, "purchase": 0.0}))
    for gt, srcs in pivot.items():
        for src, stypes in srcs.items():
            for st, vals in stypes.items():
                cpivot[gt][st]["sale"]     += vals["sale"]
                cpivot[gt][st]["purchase"] += vals["purchase"]

    rn2 = 4
    cg_sale = cg_purch = 0.0
    for gt in gain_order:
        if gt not in cpivot:
            continue
        gt_sale = gt_purch = 0.0
        c = ws2.cell(rn2, RIGHT_COL, gt); c.font = GT_FONT; c.fill = GT_FILL
        rn2 += 1
        for st, vals in cpivot[gt].items():
            cg = vals["sale"] - vals["purchase"]
            ws2.cell(rn2, RIGHT_COL, f"  {st}").font = ST_FONT
            _num(ws2, rn2, RIGHT_COL + 1, vals["sale"])
            _num(ws2, rn2, RIGHT_COL + 2, vals["purchase"])
            _num(ws2, rn2, RIGHT_COL + 3, cg, is_cg=True)
            gt_sale += vals["sale"]; gt_purch += vals["purchase"]
            rn2 += 1
        cg_sale += gt_sale; cg_purch += gt_purch

    c4 = ws2.cell(rn2, RIGHT_COL, "Grand Total"); c4.font = GRAND_FONT
    _num(ws2, rn2, RIGHT_COL + 1, cg_sale)
    _num(ws2, rn2, RIGHT_COL + 2, cg_purch)
    _num(ws2, rn2, RIGHT_COL + 3, cg_sale - cg_purch, is_cg=True)

    # Column widths + freeze
    for col, w in [(1,35),(2,20),(3,20),(4,24),(5,4),(6,26),(7,20),(8,20),(9,24)]:
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[3].height = 22
    ws2.freeze_panes = "B4"

    # ── Inject a real Excel PivotTable via XML ────────────────────
    # This makes the Summary sheet contain a live, refreshable PivotTable
    # in addition to the static snapshot above (below row rn+2).
    try:
        _inject_pivot_table(wb, ws2, rows, rn + 3)
    except Exception:
        pass  # static table is always there as fallback


    # ── Verification Sheet ──
    _write_verification_sheet(wb, rows, file_meta or [])

    wb.save(output_path)
    # Fix openpyxl quirks that cause Excel "repair" dialogs:
    #  - bare <patternFill/> → must have patternType attribute
    #  - missing <name> in <font> elements → required by OOXML spec
    try:
        import zipfile as _zf, shutil as _sh
        _tmp = output_path + "._fix_tmp"
        _sh.copy2(output_path, _tmp)
        with _zf.ZipFile(_tmp, "r") as _zin, _zf.ZipFile(output_path, "w", _zf.ZIP_DEFLATED) as _zout:
            for _item in _zin.infolist():
                _data = _zin.read(_item.filename)
                if _item.filename == "xl/styles.xml":
                    _xml = _data.decode("utf-8")
                    _xml = _xml.replace('<patternFill/>', '<patternFill patternType="none"/>')
                    def _add_font_name(_m):
                        _inner = _m.group(1)
                        if '<name ' in _inner: return _m.group(0)
                        return '<font><name val="Calibri"/>' + _inner + '</font>'
                    _xml = re.sub(r'<font>(.*?)</font>', _add_font_name, _xml, flags=re.DOTALL)
                    _data = _xml.encode("utf-8")
                _zout.writestr(_item, _data)
        os.remove(_tmp)
    except Exception as _e:
        print(f"[WARN] styles fix skipped: {_e}", file=sys.stderr)


def write_computax_excel(rows, output_path, file_meta=None, fy_start=None, fy_end=None):
    """Write Computax-ready Excel — columns B-W only.
    Output format: Microsoft Excel 97-2003 (.xls) via xlwt.
    Falls back to .xlsx if xlwt is unavailable.
    No Verification sheet in the computax file.
    """
    COMPUTAX_HEADERS = HEADERS[1:23]   # B..W header labels

    total_sale     = sum(r.get("sale_amount") or 0 for r in rows)
    total_purchase = sum(r.get("purchase_amount") or 0 for r in rows)

    try:
        import xlwt

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet("Capital Gains")

        # Header styles
        hdr_style = xlwt.easyxf(
            'font: bold true, colour white, height 180;'
            'pattern: pattern solid, fore_colour dark_blue;'
            'alignment: wrap true, horiz centre, vert centre;'
        )
        data_style = xlwt.easyxf('font: height 160;')
        data_alt   = xlwt.easyxf(
            'font: height 160;'
            'pattern: pattern solid, fore_colour ice_blue;'
        )
        num_style  = xlwt.easyxf('font: height 160;', num_format_str='#,##0.00')
        num_alt    = xlwt.easyxf(
            'font: height 160;'
            'pattern: pattern solid, fore_colour ice_blue;',
            num_format_str='#,##0.00'
        )
        date_style = xlwt.easyxf(
            'font: height 160;', num_format_str='DD/MM/YYYY')
        date_alt   = xlwt.easyxf(
            'font: height 160;'
            'pattern: pattern solid, fore_colour ice_blue;',
            num_format_str='DD/MM/YYYY'
        )

        # Blank columns (0-indexed within computax cols A-V):
        # O=14, P=15, Q=16, R=17, S=18, U=20, V=21
        BLANK_COLS = {14, 15, 16, 17, 18, 20, 21}

        # Assessment Year = FY end year + 1  (FY 2024-25 → AY 2025-26)
        if fy_start and fy_end:
            ay_label = f"{fy_end.year}-{str(fy_end.year + 1)[-2:]}"
        elif fy_end:
            ay_label = f"{fy_end.year}-{str(fy_end.year + 1)[-2:]}"
        else:
            sale_dates = [r["sale_date"] for r in rows if r.get("sale_date")]
            if sale_dates:
                mx = max(sale_dates)
                fy_end_yr = mx.year if mx.month <= 3 else mx.year + 1
                ay_label = f"{fy_end_yr}-{str(fy_end_yr + 1)[-2:]}"
            else:
                ay_label = ""

        # Row 0 (Row 1 in Excel): Assessment Year label + FY period only
        ay_style = xlwt.easyxf('font: bold true, height 180;')
        ws.write(0, 0, "Assessment Year", ay_style)
        ws.write(0, 1, ay_label, ay_style)

        # Row 1 (Row 2 in Excel): column headers
        ws.row(1).height_mismatch = True
        ws.row(1).height = 900   # ~45pt
        for ci, h in enumerate(COMPUTAX_HEADERS):
            ws.write(1, ci, h, hdr_style)

        # Rows 2+: data — skip blank columns
        for r_idx, row in enumerate(rows):
            xls_row = r_idx + 2
            alt = (xls_row % 2 == 0)
            all_vals, _ = _build_row_values(row, r_idx + 3)
            computax_vals = all_vals[1:23]
            for ci, val in enumerate(computax_vals):
                if ci in BLANK_COLS:
                    ws.write(xls_row, ci, "",
                             data_alt if alt else data_style)
                    continue
                if isinstance(val, datetime):
                    st = date_alt if alt else date_style
                    ws.write(xls_row, ci, val, st)
                elif isinstance(val, (int, float)) and ci in (4, 5, 6, 11):
                    ws.write(xls_row, ci, val, num_alt if alt else num_style)
                else:
                    ws.write(xls_row, ci, val if val is not None else "",
                             data_alt if alt else data_style)

        # Column widths (in units of 1/256 of char width)
        ct_widths = [42,18,16,14,10,14,14,16,18,12,14,14,20,14,12,12,10,10,10,18,10,10]
        for ci, w in enumerate(ct_widths):
            ws.col(ci).width = w * 256

        wb.save(output_path)

    except ImportError:
        # Fallback: write as .xlsx when xlwt not yet installed
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Capital Gains"
        # Blank columns (1-indexed within computax, matching col letters O-V)
        BLANK_COLS_OX = {15, 16, 17, 18, 19, 21, 22}  # 1-based col indices

        # Assessment Year = FY end year + 1  (FY 2024-25 → AY 2025-26)
        if fy_start and fy_end:
            ay_label = f"{fy_end.year}-{str(fy_end.year + 1)[-2:]}"
        elif fy_end:
            ay_label = f"{fy_end.year}-{str(fy_end.year + 1)[-2:]}"
        else:
            sale_dates = [r["sale_date"] for r in rows if r.get("sale_date")]
            if sale_dates:
                mx = max(sale_dates)
                fy_end_yr = mx.year if mx.month <= 3 else mx.year + 1
                ay_label = f"{fy_end_yr}-{str(fy_end_yr + 1)[-2:]}"
            else:
                ay_label = ""

        ay_font = Font(bold=True, size=11)
        ws.cell(1, 1, "Assessment Year").font = ay_font
        ws.cell(1, 2, ay_label).font = ay_font

        hdr_fill  = PatternFill("solid", fgColor="1F4E79")
        hdr_font  = Font(bold=True, color="FFFFFF", size=9)
        hdr_align = Alignment(wrap_text=True, horizontal="center", vertical="center")
        for col, header in enumerate(COMPUTAX_HEADERS, 1):
            cell = ws.cell(2, col, header)
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = hdr_align
        ws.row_dimensions[2].height = 45
        date_fmt = "dd/mm/yyyy"
        for r_idx, row in enumerate(rows, 3):
            all_vals, _ = _build_row_values(row, r_idx)
            computax_vals = all_vals[1:23]
            for c_idx, val in enumerate(computax_vals, 1):
                if c_idx in BLANK_COLS_OX:
                    ws.cell(r_idx, c_idx, "")
                    continue
                cell = ws.cell(r_idx, c_idx, val)
                cell.font = Font(size=9)
                if c_idx in (4, 11) and isinstance(val, datetime):
                    cell.number_format = date_fmt
                if c_idx in (5, 6, 7, 12): cell.number_format = '#,##0.00'
        ct_widths = [42,18,16,14,10,14,14,16,18,12,14,14,20,14,12,12,10,10,10,18,10,10]
        for i, w in enumerate(ct_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A3"
        wb.save(output_path)




# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def process_file(file_path):
    """Returns (rows, had_error) tuple.
    When a file cannot be processed, records the EXACT tripping reason
    in builtins._cge_process_errors[file_path].
    """
    import builtins
    _errors = []
    had_error = False

    ext = Path(file_path).suffix.lower()
    source_name = source_name_from_file(file_path)

    if ext == '.zip':
        extracted = extract_zip(file_path)
        all_rows = []
        any_sub_error = False
        for ef in extracted:
            try:
                sub_rows, sub_err = process_file(ef)
                all_rows.extend(sub_rows)
                if sub_err:
                    any_sub_error = True
            except Exception as e:
                print(f"[WARN] {ef}: {e}", file=sys.stderr)
                any_sub_error = True
            finally:
                try: os.unlink(ef)
                except: pass
        return all_rows, any_sub_error and len(all_rows) == 0

    parser = detect_parser(file_path)
    if not parser:
        # No specialist parser matched — try Ollama AI fallback before giving up
        print(f"[AUTO-PARSER] {Path(file_path).name}: no specialist parser matched, trying Ollama...",
              file=sys.stderr)
        try:
            auto_rows, _, auto_err = attempt_auto_parse(file_path)
            if auto_rows:
                print(f"[AUTO-PARSER] Ollama extracted {len(auto_rows)} rows", file=sys.stderr)
                # apply same post-processing as matched-parser path
                auto_rows = [
                    r for r in auto_rows
                    if not (r.get("sale_date") and r.get("purchase_date")
                            and r["sale_date"] == r["purchase_date"])
                ]
                if not hasattr(builtins, '_cge_process_errors'):
                    builtins._cge_process_errors = {}
                builtins._cge_process_errors[file_path] = []
                return auto_rows, False
            msg = (f"No parser matched for file type: {ext}. "
                   f"Ollama fallback also returned no rows: {auto_err}")
        except Exception as _e:
            msg = (f"No parser matched for file type: {ext}. "
                   f"Ollama fallback raised: {type(_e).__name__}: {_e}")
        _errors.append(f"NOT_PROCESSED: {msg}")
        print(f"[ERROR] {Path(file_path).name}: {msg}", file=sys.stderr)
        if not hasattr(builtins, '_cge_process_errors'):
            builtins._cge_process_errors = {}
        builtins._cge_process_errors[file_path] = _errors
        return [], True

    try:
        rows = parser.parse(file_path, source_name=source_name)

        # Rule: If 0 rows, diagnose the exact reason
        if len(rows) == 0:
            trip_reason = f"Parser {parser.__class__.__name__} matched but extracted 0 rows."
            try:
                with open_pdf(file_path) as _chk:
                    # Check for completely image-based (scanned) PDF — no text on any page
                    all_text = " ".join((_p.extract_text() or "") for _p in _chk.pages).strip()
                    sample = (_chk.pages[0].extract_text() or '')[:300]
                    if not all_text:
                        trip_reason = (
                            "SCANNED_IMAGE_PDF: This PDF appears to be a scanned image — "
                            "it has no readable text layer. The tool cannot extract data from "
                            "image-only PDFs.\n"
                            "What to do: Log in to the mutual fund / broker portal and download "
                            "a fresh digital statement (not a scan). Digital statements have "
                            "selectable text. If you only have a scanned copy, open it in "
                            "Adobe Acrobat, use 'Recognize Text (OCR)', then re-save and upload."
                        )
                    elif '(cid:' in sample:
                        trip_reason = (
                            "CID_ENCODED_FONT: PDF uses a proprietary embedded font that cannot "
                            "be decoded by standard text extraction. This is common with Axis Bank "
                            "and Canara Robeco statements.\n"
                            "What to do: Open this PDF in Adobe Acrobat or Mac Preview, "
                            "then File → Print → Save as PDF to create a re-encoded copy, "
                            "and upload that instead."
                        )
            except Exception:
                pass

            if 'CID-encoded' in trip_reason:
                rows = OCRPDFParser().parse(file_path, source_name=source_name)
                if rows:
                    trip_reason = None

            # Try auto-parser (Claude API) only if NOT CID-encoded and OCR didn't already get rows
            is_cid_encoded = trip_reason is not None and ('(cid:' in trip_reason or 'CID-encoded' in trip_reason)
            if not is_cid_encoded and len(rows) == 0:
                try:
                    print(f"[AUTO-PARSER] {Path(file_path).name}: 0 rows from {parser.__class__.__name__}, trying auto-generation...", file=sys.stderr)
                    # Use module globals for the auto-parser scope
                    import sys as _sys
                    _module = _sys.modules[__name__] if __name__ in _sys.modules else None
                    _g = vars(_module) if _module else globals()
                    auto_rows, auto_cls, auto_err = attempt_auto_parse(
                        file_path, _g, extractor_path=__file__)
                    if auto_rows:
                        rows = auto_rows
                        print(f"[AUTO-PARSER] Generated parser extracted {len(rows)} rows", file=sys.stderr)
                        trip_reason = None  # success
                    else:
                        _errors.append(f"AUTO_PARSER_FAILED: {auto_err}")
                        trip_reason = f"{trip_reason} Auto-parser also failed: {auto_err}"
                except Exception as _ae:
                    _errors.append(f"AUTO_PARSER_EXCEPTION: {_ae}")
                    trip_reason = f"{trip_reason} Auto-parser exception: {_ae}"

            if trip_reason and len(rows) == 0:
                _errors.append(f"NOT_PROCESSED: {trip_reason}")
                print(f"[ERROR] {Path(file_path).name}: {trip_reason}", file=sys.stderr)
                had_error = True

        # Rule 5: skip intra-day transactions (sale_date == purchase_date)
        intraday_rows = [
            r for r in rows
            if r.get("sale_date") and r.get("purchase_date")
            and r["sale_date"] == r["purchase_date"]
        ]
        if intraday_rows:
            if not hasattr(builtins, '_cge_intraday_counts'):
                builtins._cge_intraday_counts = {}
            src = source_name_from_file(file_path)
            builtins._cge_intraday_counts[src] = len(intraday_rows)
        rows = [r for r in rows if r not in intraday_rows]
        # Rule 6: clean share names
        for r in rows:
            name = r.get("share_name") or ""
            name = re.sub(r"\s*[\(\[]?\s*INF[A-Z0-9]{6,12}\s*[\)\]]?\s*$", "", name, flags=re.I)
            name = re.sub(r"\s*-\s*$", "", name).strip()
            name = re.sub(r"\s*\(\d{6,}\)\s*$", "", name).strip()
            if name:
                r["share_name"] = name

        # Capture validation report after all fallbacks — so OCR/auto-parser rows are included
        _val_report = None
        if hasattr(parser, 'last_validation') and parser.last_validation:
            _val_report = parser.last_validation
        elif rows:
            _val_report = build_validation_report(
                rows, pdf_st=None, pdf_lt=None, pdf_total=None,
                parse_failures=0, parser_name=parser.__class__.__name__,
            )
        if _val_report:
            if not hasattr(builtins, '_cge_validation_reports'):
                builtins._cge_validation_reports = {}
            builtins._cge_validation_reports[file_path] = _val_report

        if rows:
            print(f"[INFO] Extracted {len(rows)} rows from {Path(file_path).name}", file=sys.stderr)

        if not hasattr(builtins, '_cge_process_errors'):
            builtins._cge_process_errors = {}
        builtins._cge_process_errors[file_path] = _errors
        return rows, had_error or len(rows) == 0

    except Exception as e:
        trip = f"Exception during parsing with {parser.__class__.__name__}: {type(e).__name__}: {e}"
        # Check if failure is because PDF is image-only (scanned)
        if ext == '.pdf':
            try:
                import pdfplumber as _pl
                with _pl.open(file_path) as _chk:
                    all_text = " ".join((_p.extract_text() or "") for _p in _chk.pages).strip()
                if not all_text:
                    trip = (
                        "SCANNED_IMAGE_PDF: This PDF appears to be a scanned image — "
                        "it has no readable text layer. The tool cannot extract data from "
                        "image-only PDFs.\n"
                        "What to do: Log in to the mutual fund / broker portal and download "
                        "a fresh digital statement (not a scan). Digital statements have "
                        "selectable text. If you only have a scanned copy, open it in "
                        "Adobe Acrobat, use 'Recognize Text (OCR)', then re-save and upload."
                    )
            except Exception:
                pass
        _errors.append(f"NOT_PROCESSED: {trip}")
        print(f"[ERROR] {Path(file_path).name}: {trip.split(chr(10))[0]}", file=sys.stderr)
        if not hasattr(builtins, '_cge_process_errors'):
            builtins._cge_process_errors = {}
        builtins._cge_process_errors[file_path] = _errors
        return [], True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("files", nargs="+")
    ap.add_argument("-o","--output", required=True)
    ap.add_argument("--computax-output", default=None,
                    help="If given, also write a Computax-ready (B-W only) Excel here")
    ap.add_argument("--fy-start", help="FY start date YYYY-MM-DD (e.g. 2025-04-01)", default=None)
    ap.add_argument("--fy-end",   help="FY end date YYYY-MM-DD (e.g. 2026-03-31)",   default=None)
    args = ap.parse_args()

    fy_start = parse_date(args.fy_start) if args.fy_start else None
    fy_end   = parse_date(args.fy_end)   if args.fy_end   else None

    all_rows  = []
    clients   = {}
    file_meta = []   # one entry per input file
    import builtins as builtins  # needed for validation report registry

    for f in args.files:
        source = source_name_from_file(f)
        rows, had_error = process_file(f)
        all_rows.extend(rows)
        for r in rows:
            cn = r.get("client_name")
            if cn: clients.setdefault(cn, set()).add(source)
        _val = (getattr(builtins, '_cge_validation_reports', {}) or {}).get(f)
        file_meta.append({
            "source":       source,
            "_file_path":   f,              # for error lookup
            "had_error":    had_error,
            "raw_count":    len(rows),   # before date filter
            "oor_count":    0,
            "client_name":  next((r.get("client_name") for r in rows if r.get("client_name")), ""),
            "pan":          next((r.get("pan")         for r in rows if r.get("pan")),          ""),
            "remark":       "",
            "process_status": "",        # filled after filter
            "pdf_validation": _val,      # validation report (parsers that support it)
        })

    # ── date range filter ──
    out_of_range_count = 0
    out_of_range_items = []
    if fy_start or fy_end:
        all_rows, out_of_range = filter_by_date_range(all_rows, fy_start, fy_end)
        out_of_range_count = len(out_of_range)
        out_of_range_items = [
            {
                "name":      r.get("share_name",""),
                "sale_date": r["sale_date"].strftime("%d/%m/%Y") if r.get("sale_date") else "",
                "gain_type": r.get("gain_type","")
            }
            for r in out_of_range if r.get("sale_date")
        ][:30]

        # Count OOR per source
        oor_by_source = {}
        for r in out_of_range:
            src = r.get("source","")
            oor_by_source[src] = oor_by_source.get(src, 0) + 1
        for fm in file_meta:
            fm["oor_count"] = oor_by_source.get(fm["source"], 0)

    # ── determine process_status per file ──
    rows_by_source = {}
    for r in all_rows:
        rows_by_source.setdefault(r.get("source",""), []).append(r)

    intraday_counts = getattr(builtins, '_cge_intraday_counts', {})
    for fm in file_meta:
        remaining = len(rows_by_source.get(fm["source"], []))
        if fm["had_error"] or (remaining == 0 and fm["oor_count"] == 0):
            fm["process_status"] = "Not Processed"
        elif fm["oor_count"] > 0 or intraday_counts.get(fm["source"], 0) > 0:
            fm["process_status"] = "Semi-Processed"
        else:
            fm["process_status"] = "Processed"

    if not all_rows:
        import builtins as _b0
        _pe0 = getattr(_b0, '_cge_process_errors', {})
        _fmo0 = []
        for _fm0 in file_meta:
            _fmo0.append({
                "source":               _fm0["source"],
                "process_status":       _fm0["process_status"],
                "client_name":          _fm0.get("client_name", ""),
                "rows_extracted":       0,
                "oor_count":            _fm0.get("oor_count", 0),
                "sale_amount":          0,
                "purchase_amount":      0,
                "capital_gain":         0,
                "remark":               _fm0.get("remark", ""),
                "not_processed_reason": next(
                    (e.replace("NOT_PROCESSED: ", "")
                     for e in _pe0.get(_fm0.get("_file_path", ""), [])
                     if e.startswith("NOT_PROCESSED:")), ""
                ),
                "pdf_validation": None,
            })
        result = {"total_rows":0,"total_sale_amount":0,"total_purchase_amount":0,
                  "total_capital_gain":0,"output_file":args.output,"clients":{},
                  "out_of_range_count":out_of_range_count,
                  "out_of_range_items":out_of_range_items,
                  "file_meta": _fmo0}
        print(json.dumps(result)); return

    write_output_excel(all_rows, args.output, file_meta)

    # ── Self-validation: run checks on extracted rows ──
    validation_warnings = validate_output_rows(all_rows)
    if validation_warnings:
        high = [w for w in validation_warnings if w.get("severity") == "HIGH"]
        med  = [w for w in validation_warnings if w.get("severity") == "MEDIUM"]
        print(f"[VALIDATION] {len(validation_warnings)} issues found "
              f"({len(high)} HIGH, {len(med)} MEDIUM)", file=sys.stderr)
        for w in validation_warnings[:20]:   # limit stderr spam
            print(f"[VALIDATION-{w['severity']}] Row {w['row']} ({w['name'][:40]}): {w['issue']}",
                  file=sys.stderr)
    else:
        print("[VALIDATION] All checks passed — no issues detected", file=sys.stderr)

    computax_path = args.computax_output or args.output.replace(".xlsx", "_computax.xls")
    write_computax_excel(all_rows, computax_path, file_meta, fy_start=fy_start, fy_end=fy_end)

    total_sale     = sum(r.get("sale_amount") or 0 for r in all_rows)
    total_purchase = sum(r.get("purchase_amount") or 0 for r in all_rows)
    total_cg       = sum(
        apply_grandfathering(
            r.get("sale_amount") or 0,
            r.get("total_market_value_31jan2018"),
            r.get("purchase_amount") or 0,
            r.get("transfer_expenses"),
        )
        for r in all_rows
    )

    # ── add per-file totals to file_meta for portal verification tab ──
    for fm in file_meta:
        src_rows = rows_by_source.get(fm["source"], [])
        fm["rows_in_range"]    = len(src_rows)
        fm["sale_amount"]      = round(sum(r.get("sale_amount") or 0 for r in src_rows), 2)
        fm["purchase_amount"]  = round(sum(r.get("purchase_amount") or 0 for r in src_rows), 2)
        fm["capital_gain"]     = round(sum(
            apply_grandfathering(
                r.get("sale_amount") or 0,
                r.get("total_market_value_31jan2018"),
                r.get("purchase_amount") or 0,
                r.get("transfer_expenses"),
            ) for r in src_rows
        ), 2)

    # Serialise file_meta (remove non-serialisable had_error bool is fine, keep the rest)
    import builtins
    _process_errors = getattr(builtins, '_cge_process_errors', {})

    file_meta_out = []
    for fm in file_meta:
        _val = fm.get("pdf_validation") or {}
        # Build clear, human-readable validation summary for this file
        _val_summary = None
        if _val:
            gap = _val.get("gap")
            gap_pct = _val.get("gap_pct")
            gap_st = _val.get("gap_st")
            gap_lt = _val.get("gap_lt")
            gap_st_pct = _val.get("gap_st_pct")
            gap_lt_pct = _val.get("gap_lt_pct")
            max_cat_gap = max(gap_st_pct or 0, gap_lt_pct or 0)
            worst_gap = max(gap_pct or 0, max_cat_gap)
            _val_summary = {
                "pdf_stated": {
                    "st_gain":    _val.get("pdf_st_gain"),
                    "lt_gain":    _val.get("pdf_lt_gain"),
                    "total_gain": _val.get("pdf_total_gain"),
                },
                "extracted": {
                    "st_gain":    _val.get("extracted_st_gain"),
                    "lt_gain":    _val.get("extracted_lt_gain"),
                    "total_gain": _val.get("extracted_total_gain"),
                },
                "gap":          gap,
                "gap_pct":      gap_pct,
                "gap_st":       gap_st,
                "gap_lt":       gap_lt,
                "gap_st_pct":   gap_st_pct,
                "gap_lt_pct":   gap_lt_pct,
                "status":    "OK" if worst_gap < 0.5
                             else ("WARN" if worst_gap < 3.0 else "REVIEW"),
                "rows_before_filters": _val.get("rows_extracted"),
                "collapsed_parse_failures": _val.get("collapsed_parse_failures", 0),
                "causes":    _val.get("causes", []),
            }
        file_meta_out.append({
            "source":              fm["source"],
            "process_status":      fm["process_status"],
            "client_name":         fm.get("client_name", ""),
            "rows_extracted":      fm.get("rows_in_range", 0),
            "oor_count":           fm.get("oor_count", 0),
            "sale_amount":         fm.get("sale_amount", 0),
            "purchase_amount":     fm.get("purchase_amount", 0),
            "capital_gain":        fm.get("capital_gain", 0),
            "remark":              fm.get("remark", ""),
            "not_processed_reason": next(
                (e.replace("NOT_PROCESSED: ", "") for e in _process_errors.get(fm.get("_file_path",""), [])
                 if e.startswith("NOT_PROCESSED:")), ""
            ),
            "pdf_validation":      _val_summary,
        })

    result = {
        "total_rows":            len(all_rows),
        "total_sale_amount":     round(total_sale, 2),
        "total_purchase_amount": round(total_purchase, 2),
        "total_capital_gain":    round(total_cg, 2),
        "output_file":           args.output,
        "computax_file":         computax_path,
        "clients":               {k: list(v) for k, v in clients.items()},
        "out_of_range_count":    out_of_range_count,
        "out_of_range_items":    out_of_range_items,
        "file_meta":             file_meta_out,
        "validation_warnings":   validation_warnings[:50] if validation_warnings else [],
    }
    print(json.dumps(result))

if __name__ == "__main__":
    main()
